#!/usr/bin/env python3
"""Build an in-house DTER xlsx report (en-US + nl-NL corpora).

Layout:
    Sheet `inhouse_dter`:
  Row 2  : Header | Baseline | <model-label-1> | ... | WERR(s)
  Row 3  : Column | A        | B               | ... | A->B, ...
  Row 4-6 : en-US datasets
  Row 7   : en-US avg
  Row 8-10: nl-NL datasets
  Row 11  : nl-NL avg
  Row 12  : overall avg
  WERR columns: 1 - <model>/Baseline per non-baseline model column.

    Sheet `overall_improve_degrade`:
    One row per non-baseline model, sorted by overall WERR versus baseline.

Metrics are micro-DTER (sum_edits / sum_tokens), recovered per corpus from verl
`val-aux/<corpus>/dter_n_err/mean@1` and `val-aux/<corpus>/dter_n_ref/mean@1`.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")  # light blue
LANG_AVG_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # light yellow
OVERALL_AVG_FILL = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")  # light green
IMPROVE_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")  # green
DEGRADE_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # red

# ---------------------------------------------------------------------------
# Fixed schema: (locale, [canonical dataset names in report])
# ---------------------------------------------------------------------------
INHOUSE_GROUPS: List[Tuple[str, List[str]]] = [
    ("en-US", [
        "Conversation_DTEST_FY21Q1_en-US",
        "Conversation_OnlineMeetings_DTEST_FY25Q3_en-US_DTEST_OfflineDataCollection",
        "Dictation_Commonset_OfficeOffline_FY24Q3_en-US_DTEST_OfflineDataCollection",
    ]),
    ("nl-NL", [
        "Conversation_DTEST_FY23Q2_nl-NL_DTEST",
        "Conversation_OnlineMeetings_DTEST_FY23Q1_nl-NL_DTEST",
        "Dictation_DTEST_L_D_FY23Q4_nl-NL_DTEST",
    ]),
]

BASELINE_LABEL = "Qwen3.5-audio"
# Baseline = fast-llm-2605-qwen3-5-9b-s2-st-example-r2 @ step 90000.
# Values supplied in the original report (percent → fractions).
BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q1_en-US": 0.1863,
    "Conversation_OnlineMeetings_DTEST_FY25Q3_en-US_DTEST_OfflineDataCollection": 0.1374,
    "Dictation_Commonset_OfficeOffline_FY24Q3_en-US_DTEST_OfflineDataCollection": 0.1010,
    "Conversation_DTEST_FY23Q2_nl-NL_DTEST": 0.2476,
    "Conversation_OnlineMeetings_DTEST_FY23Q1_nl-NL_DTEST": 0.2422,
    "Dictation_DTEST_L_D_FY23Q4_nl-NL_DTEST": 0.1570,
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_enus_seg (segmented long-audio eval).
# 5 en-US TER corpora (the two CustomerSpeechDomainSet_* Entity sets are excluded).
# data_source keys are the short corpus names emitted by recipe.phimm long-audio eval.
# ---------------------------------------------------------------------------
ENUS_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("en-US", [
        "Conversation_DTEST_FY21Q1",
        "Conversation_OnlineMeetings_DTEST_FY25Q3",
        "Dictation_Commonset_OfficeOffline_FY24Q3",
        "OnlineMeetings_CS_Product_FY22_FullMeeting",
        "OnlineMeetings_CS_Shiproom_FY22",
    ]),
]

# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_enus_seg), micro-DTER = n_err / n_ref.
ENUS_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q1": 7766 / 41826,                  # 0.18567
    "Conversation_OnlineMeetings_DTEST_FY25Q3": 6395 / 47149,   # 0.13563
    "Dictation_Commonset_OfficeOffline_FY24Q3": 3888 / 38383,   # 0.10129
    "OnlineMeetings_CS_Product_FY22_FullMeeting": 8165 / 35020,  # 0.23315
    "OnlineMeetings_CS_Shiproom_FY22": 10158 / 39042,           # 0.26018
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_nlnl (segmented long-audio eval).
# 3 nl-NL TER corpora (the two Conversation_DomainSet_*_Entity_* sets are excluded).
# data_source keys are the short corpus names emitted by recipe.phimm long-audio eval.
# ---------------------------------------------------------------------------
NLNL_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("nl-NL", [
        "Conversation_DTEST_FY23Q2",
        "Conversation_OnlineMeetings_DTEST_FY23Q1",
        "Dictation_DTEST_L_D_FY23Q4",
    ]),
]

# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_nlnl), micro-DTER = n_err / n_ref.
NLNL_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY23Q2": 11359 / 45958,                 # 0.24716
    "Conversation_OnlineMeetings_DTEST_FY23Q1": 11148 / 46574,  # 0.23936
    "Dictation_DTEST_L_D_FY23Q4": 6391 / 40642,                 # 0.15725
}

# Registry of selectable schemas: name -> (groups, baseline_metrics).
# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_dadk (segmented long-audio eval).
# 3 da-DK TER corpora. data_source keys are the short corpus names.
# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_dadk), micro-DTER = n_err / n_ref.
# ---------------------------------------------------------------------------
DADK_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("da-DK", [
        "Conversation_DTEST_FY21Q3",
        "Conversation_OnlineMeetings_DTEST_FY23Q1",
        "Dictation_DTEST_L_D_FY23Q4",
    ]),
]

DADK_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q3": 13725 / 58149,                 # 0.23603
    "Conversation_OnlineMeetings_DTEST_FY23Q1": 11911 / 48951,  # 0.24332
    "Dictation_DTEST_L_D_FY23Q4": 10003 / 44482,                # 0.22488
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_huhu (segmented long-audio eval).
# 3 hu-HU TER corpora. data_source keys are the short corpus names.
# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_huhu), micro-DTER = n_err / n_ref.
# ---------------------------------------------------------------------------
HUHU_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("hu-HU", [
        "Conversation_DTEST_FY22Q4",
        "Conversation_OnlineMeetings_DTEST_FY24Q2",
        "Dictation_DTEST_L_D_FY25Q2",
    ]),
]

HUHU_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY22Q4": 7655 / 33569,                  # 0.22804
    "Conversation_OnlineMeetings_DTEST_FY24Q2": 7676 / 35005,   # 0.21928
    "Dictation_DTEST_L_D_FY25Q2": 7557 / 31098,                 # 0.24301
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_nbno (segmented long-audio eval).
# 3 nb-NO TER corpora. data_source keys are the short corpus names.
# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_nbno), micro-DTER = n_err / n_ref.
# ---------------------------------------------------------------------------
NBNO_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("nb-NO", [
        "Conversation_DTEST_FY21Q3",
        "Conversation_OnlineMeetings_DTEST_FY23Q1",
        "Dictation_DTEST_L_D_FY23Q4",
    ]),
]

NBNO_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q3": 10345 / 47273,                 # 0.21884
    "Conversation_OnlineMeetings_DTEST_FY23Q1": 7803 / 37996,   # 0.20536
    "Dictation_DTEST_L_D_FY23Q4": 8584 / 40601,                 # 0.21142
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_cscz (segmented long-audio eval).
# 3 cs-CZ TER corpora. Uses new-style (slug, display) tuples so the internal
# key matches the per-corpus slug emitted by the long-audio eval (summary line
# `[cscz_conv_fy23q2] DTER: ...`, blob dir `cscz_conv_fy23q2/measures.json`, and
# `val-aux/cscz_conv_fy23q2/dter_n_err/mean@1`).
# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_cscz), micro-DTER = n_err / n_ref.
# ---------------------------------------------------------------------------
CSCZ_SEG_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    ("cs-CZ", [
        ("cscz_conv_fy23q2", "Conversation_DTEST_FY23Q2"),
        ("cscz_conv_om_fy24q2", "Conversation_OnlineMeetings_DTEST_FY24Q2"),
        ("cscz_dict_fy24q2", "Dictation_DTEST_L_D_FY24Q2"),
    ]),
]

CSCZ_SEG_BASELINE_METRICS: Dict[str, float] = {
    "cscz_conv_fy23q2": 10218 / 43021,     # 0.23751
    "cscz_conv_om_fy24q2": 5514 / 38232,   # 0.14422
    "cscz_dict_fy24q2": 5212 / 39881,      # 0.13069
}

# ---------------------------------------------------------------------------
# Combined schema: inhouse_2605_5lang_seg_v2 (6 locales × 3 corpora).
#
# Internal keys are the per-corpus slug directory names produced by the
# segmented long-audio eval (e.g. `enus_conv_fy21q1`). Each slug carries its
# own `measures.json` with `dter_n_err` / `dter_n_ref`, so slugs are unique
# across locales (whereas the short corpus names collide, e.g.
# `Conversation_DTEST_FY21Q3` exists for both da-DK and nb-NO).
#
# Display labels are the short corpus names prefixed with the locale code
# (e.g. `en-US_Conversation_DTEST_FY21Q1`) so each row is unambiguous and sorts
# by locale. The 3 en-US corpora here are the "DTEST" subset
# used in the cross-locale report — the two `OnlineMeetings_CS_*` Entity sets
# from the en-US-only schema are intentionally excluded.
# ---------------------------------------------------------------------------
ALL_SEG_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    ("en-US", [
        ("enus_conv_fy21q1", "en-US_Conversation_DTEST_FY21Q1"),
        ("enus_conv_om_fy25q3", "en-US_Conversation_OnlineMeetings_DTEST_FY25Q3"),
        ("enus_dict_office_fy24q3", "en-US_Dictation_Commonset_OfficeOffline_FY24Q3"),
    ]),
    ("nl-NL", [
        ("nlnl_conv_fy23q2", "nl-NL_Conversation_DTEST_FY23Q2"),
        ("nlnl_conv_om_fy23q1", "nl-NL_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("nlnl_dict_fy23q4", "nl-NL_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("da-DK", [
        ("dadk_conv_fy21q3", "da-DK_Conversation_DTEST_FY21Q3"),
        ("dadk_conv_om_fy23q1", "da-DK_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("dadk_dict_fy23q4", "da-DK_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("hu-HU", [
        ("huhu_conv_fy22q4", "hu-HU_Conversation_DTEST_FY22Q4"),
        ("huhu_conv_om_fy24q2", "hu-HU_Conversation_OnlineMeetings_DTEST_FY24Q2"),
        ("huhu_dict_fy25q2", "hu-HU_Dictation_DTEST_L_D_FY25Q2"),
    ]),
    ("nb-NO", [
        ("nbno_conv_fy21q3", "nb-NO_Conversation_DTEST_FY21Q3"),
        ("nbno_conv_om_fy23q1", "nb-NO_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("nbno_dict_fy23q4", "nb-NO_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("cs-CZ", [
        ("cscz_conv_fy23q2", "cs-CZ_Conversation_DTEST_FY23Q2"),
        ("cscz_conv_om_fy24q2", "cs-CZ_Conversation_OnlineMeetings_DTEST_FY24Q2"),
        ("cscz_dict_fy24q2", "cs-CZ_Dictation_DTEST_L_D_FY24Q2"),
    ]),
]

# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_5lang_seg_v2), micro-DTER per
# slug. Numbers match the per-locale embedded baselines above (en-US uses the
# 3 DTEST corpora; the OnlineMeetings_CS_* sets are excluded here).
ALL_SEG_BASELINE_METRICS: Dict[str, float] = {
    # en-US (subset of ENUS_SEG_BASELINE_METRICS)
    "enus_conv_fy21q1": ENUS_SEG_BASELINE_METRICS["Conversation_DTEST_FY21Q1"],
    "enus_conv_om_fy25q3": ENUS_SEG_BASELINE_METRICS["Conversation_OnlineMeetings_DTEST_FY25Q3"],
    "enus_dict_office_fy24q3": ENUS_SEG_BASELINE_METRICS["Dictation_Commonset_OfficeOffline_FY24Q3"],
    # nl-NL
    "nlnl_conv_fy23q2": NLNL_SEG_BASELINE_METRICS["Conversation_DTEST_FY23Q2"],
    "nlnl_conv_om_fy23q1": NLNL_SEG_BASELINE_METRICS["Conversation_OnlineMeetings_DTEST_FY23Q1"],
    "nlnl_dict_fy23q4": NLNL_SEG_BASELINE_METRICS["Dictation_DTEST_L_D_FY23Q4"],
    # da-DK
    "dadk_conv_fy21q3": DADK_SEG_BASELINE_METRICS["Conversation_DTEST_FY21Q3"],
    "dadk_conv_om_fy23q1": DADK_SEG_BASELINE_METRICS["Conversation_OnlineMeetings_DTEST_FY23Q1"],
    "dadk_dict_fy23q4": DADK_SEG_BASELINE_METRICS["Dictation_DTEST_L_D_FY23Q4"],
    # hu-HU
    "huhu_conv_fy22q4": HUHU_SEG_BASELINE_METRICS["Conversation_DTEST_FY22Q4"],
    "huhu_conv_om_fy24q2": HUHU_SEG_BASELINE_METRICS["Conversation_OnlineMeetings_DTEST_FY24Q2"],
    "huhu_dict_fy25q2": HUHU_SEG_BASELINE_METRICS["Dictation_DTEST_L_D_FY25Q2"],
    # nb-NO
    "nbno_conv_fy21q3": NBNO_SEG_BASELINE_METRICS["Conversation_DTEST_FY21Q3"],
    "nbno_conv_om_fy23q1": NBNO_SEG_BASELINE_METRICS["Conversation_OnlineMeetings_DTEST_FY23Q1"],
    "nbno_dict_fy23q4": NBNO_SEG_BASELINE_METRICS["Dictation_DTEST_L_D_FY23Q4"],
    # cs-CZ
    "cscz_conv_fy23q2": CSCZ_SEG_BASELINE_METRICS["cscz_conv_fy23q2"],
    "cscz_conv_om_fy24q2": CSCZ_SEG_BASELINE_METRICS["cscz_conv_om_fy24q2"],
    "cscz_dict_fy24q2": CSCZ_SEG_BASELINE_METRICS["cscz_dict_fy24q2"],
}

# Registry of selectable schemas: name -> (groups, baseline_metrics).
# `groups` entries may use either bare strings (key == display label) or
# `(internal_key, display_label)` tuples; `_normalize_groups` handles both.
SCHEMAS: Dict[str, Tuple[List[Tuple[str, List]], Dict[str, float]]] = {
    "default": (INHOUSE_GROUPS, BASELINE_METRICS),
    "enus_seg": (ENUS_SEG_GROUPS, ENUS_SEG_BASELINE_METRICS),
    "nlnl_seg": (NLNL_SEG_GROUPS, NLNL_SEG_BASELINE_METRICS),
    "dadk_seg": (DADK_SEG_GROUPS, DADK_SEG_BASELINE_METRICS),
    "huhu_seg": (HUHU_SEG_GROUPS, HUHU_SEG_BASELINE_METRICS),
    "nbno_seg": (NBNO_SEG_GROUPS, NBNO_SEG_BASELINE_METRICS),
    "cscz_seg": (CSCZ_SEG_GROUPS, CSCZ_SEG_BASELINE_METRICS),
    "all_seg": (ALL_SEG_GROUPS, ALL_SEG_BASELINE_METRICS),
}


def _normalize_groups(groups: List[Tuple[str, List]]) -> List[Tuple[str, List[Tuple[str, str]]]]:
    """Normalize a schema's groups to `(locale, [(internal_key, display_label)])`.

    Legacy schemas store each dataset as a bare string; treat it as both key and
    display label. New-style schemas (e.g. `all_seg`) already use `(key, display)`.
    """
    out: List[Tuple[str, List[Tuple[str, str]]]] = []
    for locale, items in groups:
        pairs: List[Tuple[str, str]] = []
        for item in items:
            if isinstance(item, str):
                pairs.append((item, item))
            else:
                k, d = item
                pairs.append((str(k), str(d)))
        out.append((locale, pairs))
    return out

# Match: val-aux/<corpus>/<key>/mean@1:<float>   where <key> ∈ {dter, dter_n_err, dter_n_ref}
DTER_LINE_RE = re.compile(
    r"val-aux/(?P<dataset>[A-Za-z0-9_.\-]+)/(?P<key>dter|dter_n_err|dter_n_ref)/mean@1[:=]\s*(?P<value>[0-9.eE+-]+)"
)

# Match the long_eval summary line, e.g.:
#   [Conversation_DTEST_FY21Q3] DTER: 23.51% [13668/58149]  EER: 0.00% [0/0]  on 34 recordings
# The bracketed [n_err/n_ref] counts yield the exact micro-DTER per corpus.
DTER_SUMMARY_RE = re.compile(
    r"\[(?P<dataset>[A-Za-z0-9_.\-]+)\]\s+DTER:\s+[0-9.]+%\s+\[(?P<n_err>[0-9]+)/(?P<n_ref>[0-9]+)\]"
)

LOCALE_SUFFIX_RE = re.compile(r"_[a-z]{2}-[A-Z]{2}(?:_.*)?$")


def _canonical_lookup(parsed: Dict[str, float], canonical: str) -> Optional[float]:
    """Look up a parsed value for a canonical dataset name with sensible fallbacks."""
    if canonical in parsed:
        return parsed[canonical]
    # Strip _xx-XX(...) suffix and retry.
    short = LOCALE_SUFFIX_RE.sub("", canonical)
    if short != canonical and short in parsed:
        return parsed[short]
    # Build suffix-stripped, case-insensitive views of the parsed keys so that
    # either side carrying a locale suffix (e.g. `..._hu-HU`) still matches.
    norm: Dict[str, float] = {}
    for k, v in parsed.items():
        norm.setdefault(k.lower(), v)
        ks = LOCALE_SUFFIX_RE.sub("", k)
        if ks != k:
            norm.setdefault(ks.lower(), v)
    for cand in (canonical.lower(), short.lower()):
        if cand in norm:
            return norm[cand]
    return None


def parse_dter_lines(text: str) -> Dict[str, float]:
    """Extract micro-DTER per corpus from verl log text.

    Reads the latest `dter_n_err/mean@1` and `dter_n_ref/mean@1` per corpus and
    returns `n_err / n_ref` (fraction in 0..1). Falls back to the raw `dter/mean@1`
    value only when the n_err / n_ref pair is missing (with a warning) — that
    macro value does NOT match in-house micro-DTER and should be replaced.
    """
    n_err: Dict[str, float] = {}
    n_ref: Dict[str, float] = {}
    macro: Dict[str, float] = {}
    for m in DTER_LINE_RE.finditer(text):
        try:
            v = float(m.group("value"))
        except ValueError:
            continue
        ds = m.group("dataset")
        key = m.group("key")
        if key == "dter_n_err":
            n_err[ds] = v
        elif key == "dter_n_ref":
            n_ref[ds] = v
        else:  # 'dter'
            macro[ds] = v

    # Long-audio eval summary lines: [<corpus>] DTER: x% [n_err/n_ref] ...
    # These carry exact micro-DTER counts and take precedence.
    for m in DTER_SUMMARY_RE.finditer(text):
        ds = m.group("dataset")
        n_err[ds] = float(m.group("n_err"))
        n_ref[ds] = float(m.group("n_ref"))

    out: Dict[str, float] = {}
    for ds in set(n_err) | set(n_ref) | set(macro):
        if ds in n_err and ds in n_ref and n_ref[ds] > 0:
            out[ds] = n_err[ds] / n_ref[ds]
        elif ds in macro:
            print(
                f"[warn] {ds}: only macro `dter/mean@1` available; using it as a fallback. "
                "This is NOT the in-house micro-DTER reference metric.",
                file=sys.stderr,
            )
            out[ds] = macro[ds] if macro[ds] <= 1.5 else macro[ds] / 100.0
    return out


def fetch_ray_logs(node: str, job_id: str) -> str:
    remote = f"bash -l -c 'ray job logs {job_id} 2>&1'"
    cmd = ["brix", "ssh", node, "--", remote]
    proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        print(
            f"[warn] ray job logs failed for {node}/{job_id} (rc={proc.returncode}):\n"
            f"{proc.stderr[-500:]}",
            file=sys.stderr,
        )
    return proc.stdout


def load_metrics(args: argparse.Namespace) -> Dict[str, float]:
    """Collect raw parsed metrics keyed by whatever names appear in the logs."""
    metrics: Dict[str, float] = {}
    if args.metrics:
        data = json.loads(Path(args.metrics).read_text())
        metrics.update({k: float(v) for k, v in data.items()})
    if args.from_text:
        for text_path in args.from_text:
            metrics.update(parse_dter_lines(Path(text_path).read_text()))
    if args.from_ray:
        for node, job_id in args.from_ray:
            metrics.update(parse_dter_lines(fetch_ray_logs(node, job_id)))
    return metrics


def load_source_file(path: str) -> Dict[str, float]:
    """Load one local result into a raw {dataset_or_slug: dter_fraction} dict.

    Auto-detects the source kind:
      * a path/URL ending in ``/`` or that is a local directory, **or** an
        ``az://`` URL, is treated as a directory of ``<slug>/measures.json``
        files (see :func:`load_dir_metrics`) — this is the canonical form for
        the segmented long-audio eval (e.g. ``inhouse_2605_5lang_seg_v2``);
      * a JSON object mapping dataset -> numeric fraction (the ``--metrics``
        format);
      * otherwise a text log containing ``val-aux/<corpus>/dter_n_err|dter_n_ref``
        lines and/or ``[<corpus>] DTER: x% [n_err/n_ref]`` long-eval summary
        lines, parsed via :func:`parse_dter_lines`.
    """
    # Directory / blob: `<root>/<slug>/measures.json` layout.
    if path.startswith("az://") or path.endswith("/") or Path(path).is_dir():
        return load_dir_metrics(path)

    text = Path(path).read_text()
    try:
        data = json.loads(text)
    except (json.JSONDecodeError, ValueError):
        data = None
    if isinstance(data, dict) and data and all(
        isinstance(v, (int, float)) and not isinstance(v, bool) for v in data.values()
    ):
        return {k: float(v) for k, v in data.items()}
    return parse_dter_lines(text)


def load_dir_metrics(root: str) -> Dict[str, float]:
    """Read a tree of per-corpus ``<slug>/measures.json`` into ``{slug: dter}``.

    Supports both local directories and ``az://`` URLs (via ``bbb ls`` /
    ``bbb cat``). ``dter`` is the micro-DTER reported by the long-audio eval
    (``dter_n_err / dter_n_ref``); the ``dter`` field in ``measures.json`` is
    already that ratio, so we use it directly when present.
    """
    out: Dict[str, float] = {}
    if root.startswith("az://"):
        listing = subprocess.run(
            ["bbb", "ls", root.rstrip("/") + "/"],
            capture_output=True, text=True, check=False,
        )
        if listing.returncode != 0:
            print(f"[warn] bbb ls {root} failed:\n{listing.stderr[-500:]}", file=sys.stderr)
            return out
        # bbb ls wraps long lines; flatten and recover URLs.
        raw = listing.stdout.replace("\n", "")
        # Each blob URL starts with az:// and ends just before the next az://.
        entries = [tok for tok in re.split(r"(?=az://)", raw) if tok.startswith("az://")]
        slugs: List[str] = []
        for entry in entries:
            entry = entry.strip()
            # Subdirectories end with `/`; we want their basename as the slug.
            if entry.endswith("/"):
                slug = entry.rstrip("/").rsplit("/", 1)[-1]
                slugs.append(slug)
        for slug in slugs:
            url = f"{root.rstrip('/')}/{slug}/measures.json"
            proc = subprocess.run(
                ["bbb", "cat", url], capture_output=True, text=True, check=False,
            )
            if proc.returncode != 0:
                print(f"[warn] bbb cat {url} failed:\n{proc.stderr[-300:]}", file=sys.stderr)
                continue
            try:
                m = json.loads(proc.stdout)
            except json.JSONDecodeError:
                print(f"[warn] could not parse {url}", file=sys.stderr)
                continue
            dter = _measures_to_dter(m)
            if dter is not None:
                out[slug] = dter
        return out

    # Local directory.
    base = Path(root)
    if not base.is_dir():
        print(f"[warn] {root} is not a directory", file=sys.stderr)
        return out
    for child in sorted(base.iterdir()):
        if not child.is_dir():
            continue
        mfile = child / "measures.json"
        if not mfile.exists():
            continue
        try:
            m = json.loads(mfile.read_text())
        except json.JSONDecodeError:
            print(f"[warn] could not parse {mfile}", file=sys.stderr)
            continue
        dter = _measures_to_dter(m)
        if dter is not None:
            out[child.name] = dter
    return out


def _measures_to_dter(m: Dict) -> Optional[float]:
    """Pull micro-DTER from a ``measures.json`` dict.

    Prefers the exact ``dter_n_err / dter_n_ref`` ratio; falls back to the
    pre-computed ``dter`` field.
    """
    n_err = m.get("dter_n_err")
    n_ref = m.get("dter_n_ref")
    if isinstance(n_err, (int, float)) and isinstance(n_ref, (int, float)) and n_ref > 0:
        return float(n_err) / float(n_ref)
    v = m.get("dter")
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _project_to_canonical(parsed: Dict[str, float]) -> Dict[str, float]:
    """Map parsed metrics to the canonical internal keys used in the report."""
    out: Dict[str, float] = {}
    for _, datasets in _normalize_groups(INHOUSE_GROUPS):
        for key, _display in datasets:
            v = _canonical_lookup(parsed, key)
            if v is not None:
                out[key] = v
    return out


# ---------------------------------------------------------------------------
# xlsx build
# ---------------------------------------------------------------------------
HEADER_ROW = 2
COLUMN_ROW = 3
DATA_START = 4


def _row_layout() -> Tuple[List[Tuple[str, str, str]], int]:
    """Return (rows, overall_avg_row).

    Each row is ``(kind, internal_key, display_label)`` where ``kind`` is
    ``'data'`` or ``'lang_avg'``. For ``lang_avg`` rows ``internal_key`` is the
    locale and ``display_label`` is the ``<locale> avg`` text. The overall avg
    row index is the row immediately following the last layout row.
    """
    rows: List[Tuple[str, str, str]] = []
    for locale, datasets in _normalize_groups(INHOUSE_GROUPS):
        for key, display in datasets:
            rows.append(("data", key, display))
        rows.append(("lang_avg", locale, f"{locale} avg"))
    overall_avg_row = DATA_START + len(rows)
    return rows, overall_avg_row


def build_workbook(columns: List[Tuple[str, Dict[str, float]]], out_path: Path) -> None:
    """``columns`` is a list of (label, canonical metrics dict). First is the baseline.

    Metrics dicts are keyed by each schema's *internal key* (e.g. the slug for
    `all_seg` or the canonical corpus name for the other schemas).
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "inhouse_dter"

    n_models = len(columns)
    model_cols = list(range(2, 2 + n_models))
    werr_cols = list(range(2 + n_models, 2 + n_models + (n_models - 1)))

    bold = Font(bold=True)

    # Header row
    ws.cell(row=HEADER_ROW, column=1, value="Header").font = bold
    ws.cell(row=HEADER_ROW, column=2, value="Baseline").font = bold
    for i, (label, _) in enumerate(columns[1:], start=0):
        ws.cell(row=HEADER_ROW, column=3 + i, value=label).font = bold
    for j, _ in enumerate(model_cols[1:]):
        ws.cell(row=HEADER_ROW, column=werr_cols[j], value="WERR").font = bold

    # Column row
    ws.cell(row=COLUMN_ROW, column=1, value="Column").font = bold
    for i, mc in enumerate(model_cols):
        ws.cell(row=COLUMN_ROW, column=mc, value=chr(ord("A") + i)).font = bold
    for j, wc in enumerate(werr_cols):
        target_letter = chr(ord("A") + 1 + j)
        ws.cell(row=COLUMN_ROW, column=wc, value=f"A->{target_letter}").font = bold

    # Data + lang_avg rows
    layout, overall_avg_row = _row_layout()
    base_letter = get_column_letter(model_cols[0])
    current_lang_start: Optional[int] = None
    lang_avg_rows: List[int] = []

    for offset, (kind, key, display) in enumerate(layout):
        r = DATA_START + offset
        ws.cell(row=r, column=1, value=display)
        if kind == "data":
            if current_lang_start is None:
                current_lang_start = r
            for i, (_, metrics) in enumerate(columns):
                v = metrics.get(key)
                if v is not None:
                    ws.cell(row=r, column=model_cols[i], value=float(v))
            for j, wc in enumerate(werr_cols):
                tgt_letter = get_column_letter(model_cols[1 + j])
                ws.cell(row=r, column=wc, value=f"=1-{tgt_letter}{r}/{base_letter}{r}")
        else:  # lang_avg
            lang_avg_rows.append(r)
            ws.cell(row=r, column=1).font = bold
            assert current_lang_start is not None
            group_keys = [
                k for (kk, k, _d) in layout[offset - (r - current_lang_start):offset]
                if kk == "data"
            ]
            lang_avg_per_col: List[Optional[float]] = []
            for i, (_, metrics) in enumerate(columns):
                vals = [metrics[k] for k in group_keys if k in metrics]
                avg = sum(vals) / len(vals) if vals else None
                lang_avg_per_col.append(avg)
                if avg is not None:
                    ws.cell(row=r, column=model_cols[i], value=float(avg)).font = bold
            for j, wc in enumerate(werr_cols):
                base_v = lang_avg_per_col[0]
                tgt_v = lang_avg_per_col[1 + j]
                if base_v and tgt_v is not None:
                    ws.cell(row=r, column=wc, value=float(1 - tgt_v / base_v)).font = bold
            current_lang_start = None

    # Overall avg row (mean across all data datasets).
    r = overall_avg_row
    ws.cell(row=r, column=1, value="overall avg").font = bold
    all_data_keys = [k for kk, k, _d in layout if kk == "data"]
    overall_per_col: List[Optional[float]] = []
    for i, (_, metrics) in enumerate(columns):
        vals = [metrics[k] for k in all_data_keys if k in metrics]
        avg = sum(vals) / len(vals) if vals else None
        overall_per_col.append(avg)
        if avg is not None:
            ws.cell(row=r, column=model_cols[i], value=float(avg)).font = bold
    for j, wc in enumerate(werr_cols):
        base_v = overall_per_col[0]
        tgt_v = overall_per_col[1 + j]
        if base_v and tgt_v is not None:
            ws.cell(row=r, column=wc, value=float(1 - tgt_v / base_v)).font = bold

    _add_overall_improve_degrade_sheet(wb, columns, all_data_keys, overall_per_col)

    # Formatting.
    pct_fmt = "0.00%"
    for rr in range(DATA_START, overall_avg_row + 1):
        for mc in model_cols + werr_cols:
            ws.cell(row=rr, column=mc).number_format = pct_fmt

    all_cols = [1] + model_cols + werr_cols
    header_rows = [HEADER_ROW, COLUMN_ROW]
    for rr in header_rows:
        for c in all_cols:
            ws.cell(row=rr, column=c).fill = HEADER_FILL
    for rr in lang_avg_rows:
        for c in all_cols:
            ws.cell(row=rr, column=c).fill = LANG_AVG_FILL
    for c in all_cols:
        ws.cell(row=overall_avg_row, column=c).fill = OVERALL_AVG_FILL

    ws.column_dimensions["A"].width = 60
    for mc in model_cols + werr_cols:
        ws.column_dimensions[get_column_letter(mc)].width = 18

    center = Alignment(horizontal="center", vertical="center")
    for rr in range(HEADER_ROW, overall_avg_row + 1):
        for c in model_cols + werr_cols:
            ws.cell(row=rr, column=c).alignment = center

    # WERR 3-color scale.
    for wc in werr_cols:
        col_letter = get_column_letter(wc)
        rng = f"{col_letter}{DATA_START}:{col_letter}{overall_avg_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=-1, start_color="FFF8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFFFF",
            end_type="num", end_value=1, end_color="FF63BE7B",
        )
        ws.conditional_formatting.add(rng, rule)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _add_overall_improve_degrade_sheet(
    wb: Workbook,
    columns: List[Tuple[str, Dict[str, float]]],
    all_data_keys: List[str],
    overall_per_col: List[Optional[float]],
) -> None:
    """Add a workbook-level improve/degrade summary for non-baseline models."""
    if len(columns) <= 1:
        return

    ws = wb.create_sheet("overall_improve_degrade")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    pct_fmt = "0.00%"

    ws.cell(row=1, column=1, value="Overall improve/degrade").font = Font(bold=True, size=14)
    headers = [
        "Rank",
        "Model",
        "Direction",
        "Baseline overall DTER",
        "Model overall DTER",
        "DTER delta",
        "WERR",
        "Datasets",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = bold
        cell.fill = HEADER_FILL
        cell.alignment = center

    baseline_overall = overall_per_col[0] if overall_per_col else None
    summary_rows: List[Tuple[float, str, Optional[float], Optional[float], Optional[float], int]] = []
    for i, (label, metrics) in enumerate(columns[1:], start=1):
        model_overall = overall_per_col[i] if i < len(overall_per_col) else None
        dter_delta = None
        werr = None
        if baseline_overall is not None and model_overall is not None:
            dter_delta = baseline_overall - model_overall
            if baseline_overall:
                werr = 1 - model_overall / baseline_overall
        dataset_count = sum(1 for key in all_data_keys if key in metrics)
        sort_value = werr if werr is not None else float("-inf")
        summary_rows.append((sort_value, label, model_overall, dter_delta, werr, dataset_count))

    summary_rows.sort(key=lambda row: row[0], reverse=True)

    for rank, (_sort_value, label, model_overall, dter_delta, werr, dataset_count) in enumerate(summary_rows, start=1):
        row = rank + 2
        direction = "improve" if (werr is not None and werr >= 0) else "degrade"
        direction_fill = IMPROVE_FILL if direction == "improve" else DEGRADE_FILL
        values = [
            rank,
            label,
            direction,
            baseline_overall,
            model_overall,
            dter_delta,
            werr,
            f"{dataset_count}/{len(all_data_keys)}",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if col in (1, 3, 4, 5, 6, 7, 8):
                cell.alignment = center
            if col in (4, 5, 6, 7):
                cell.number_format = pct_fmt
            if col == 3:
                cell.fill = direction_fill

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{max(2, len(summary_rows) + 2)}"

    if summary_rows:
        for col_letter in ("F", "G"):
            rng = f"{col_letter}3:{col_letter}{len(summary_rows) + 2}"
            rule = ColorScaleRule(
                start_type="num", start_value=-1, start_color="FFF8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                end_type="num", end_value=1, end_color="FF63BE7B",
            )
            ws.conditional_formatting.add(rng, rule)


def read_existing_xlsx(path: Path) -> List[Tuple[str, Dict[str, float]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["inhouse_dter"] if "inhouse_dter" in wb.sheetnames else wb.active

    model_cols: List[int] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=COLUMN_ROW, column=c).value
        if isinstance(v, str) and len(v) == 1 and v.isalpha():
            model_cols.append(c)
        else:
            break

    labels: List[str] = []
    for mc in model_cols:
        v = ws.cell(row=HEADER_ROW, column=mc).value
        labels.append(str(v) if v is not None else f"col{mc}")

    metrics_per_col: List[Dict[str, float]] = [dict() for _ in model_cols]
    layout, _ = _row_layout()
    for offset, (kind, key, _display) in enumerate(layout):
        if kind != "data":
            continue
        r = DATA_START + offset
        for i, mc in enumerate(model_cols):
            v = ws.cell(row=r, column=mc).value
            if isinstance(v, (int, float)):
                metrics_per_col[i][key] = float(v)

    return [(label, m) for label, m in zip(labels, metrics_per_col)]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("label", nargs="?", help="New model label (column header). Optional with --baseline-only.")
    p.add_argument(
        "--baseline-only",
        action="store_true",
        help="Emit only the embedded baseline column (no model/WERR columns). Useful for registering or publishing a schema's baseline numbers.",
    )
    p.add_argument(
        "--schema",
        choices=sorted(SCHEMAS),
        default="default",
        help=(
            "Dataset schema / embedded baseline to compare against. "
            "'default' = 6-dataset en-US+nl-NL canonical; "
            "'enus_seg' = 5 en-US TER corpora from inhouse_2605_enus_seg; "
            "'nlnl_seg' = 3 nl-NL TER corpora from inhouse_2605_nlnl; "
            "'dadk_seg' = 3 da-DK corpora from inhouse_2605_dadk; "
            "'huhu_seg' = 3 hu-HU corpora from inhouse_2605_huhu; "
            "'nbno_seg' = 3 nb-NO corpora from inhouse_2605_nbno; "
            "'cscz_seg' = 3 cs-CZ corpora from inhouse_2605_cscz; "
            "'all_seg' = 18 corpora across 6 locales from inhouse_2605_5lang_seg_v2."
        ),
    )
    p.add_argument("--from-ray", nargs=2, metavar=("NODE", "JOB_ID"), action="append", default=[])
    p.add_argument("--from-text", action="append", default=[])
    p.add_argument("--metrics", help="JSON file: {dataset: dter_fraction}.")
    p.add_argument(
        "--model",
        "-m",
        nargs=2,
        metavar=("LABEL", "PATH"),
        action="append",
        default=[],
        help=(
            "Add a model column from a single local result (repeatable). PATH "
            "is auto-detected as one of: an `az://` URL or local directory "
            "containing `<slug>/measures.json` per corpus (the canonical "
            "segmented long-audio eval layout, e.g. "
            "`az://.../inhouse_2605_5lang_seg_v2/`); a JSON "
            "`{dataset: dter_fraction}` mapping; or a text log containing "
            "'val-aux/<corpus>/dter_n_err|dter_n_ref/mean@1' lines and/or "
            "'[<corpus>] DTER: x% [n_err/n_ref]' long-eval summary lines. Each "
            "--model becomes its own column (B, C, D, ...), in the order "
            "given, after the baseline (and any --extend-xlsx columns). "
            "Combine multiple local results into one workbook in a single command."
        ),
    )
    p.add_argument("--baseline", help="Override baseline metrics JSON.")
    p.add_argument("--baseline-label", default=BASELINE_LABEL)
    p.add_argument(
        "--extend-xlsx",
        help=(
            "Existing xlsx to extend by appending the new model as the next column. "
            "The sheet layout must match the template."
        ),
    )
    p.add_argument("--out", help="Output xlsx path (default: tmp/inhouse_dter_report/<label>.xlsx).")
    return p.parse_args()


def main() -> int:
    args = parse_args()

    # Select dataset schema + embedded baseline. _row_layout / _project_to_canonical
    # read the module-level INHOUSE_GROUPS, so rebind it for the chosen schema.
    global INHOUSE_GROUPS
    schema_groups, schema_baseline = SCHEMAS[args.schema]
    INHOUSE_GROUPS = schema_groups

    # Baseline-only mode: emit just the embedded (or overridden) baseline column.
    if args.baseline_only:
        baseline_metrics = schema_baseline
        if args.baseline:
            baseline_metrics = {k: float(v) for k, v in json.loads(Path(args.baseline).read_text()).items()}
        columns = [(args.baseline_label, baseline_metrics)]
        out_path = Path(args.out or f"tmp/inhouse_dter_report/{args.schema}_baseline.xlsx")
        build_workbook(columns, out_path)
        print(out_path)
        return 0

    if args.label is None and not args.model:
        print(
            "[error] provide a model column via the positional label (+ "
            "--metrics/--from-text/--from-ray) and/or one or more --model LABEL PATH, "
            "unless --baseline-only is set.",
            file=sys.stderr,
        )
        return 2

    def _project_or_die(raw: Dict[str, float], who: str) -> Optional[Dict[str, float]]:
        if not raw:
            print(f"[error] {who}: no metrics collected.", file=sys.stderr)
            return None
        proj = _project_to_canonical(raw)
        if not proj:
            print(
                f"[error] {who}: none of the parsed metrics map to canonical "
                f"in-house dataset names for schema '{args.schema}'.\n"
                f"  parsed keys: {sorted(raw)[:10]}{'...' if len(raw) > 10 else ''}",
                file=sys.stderr,
            )
            return None
        return proj

    # Assemble the new model columns, in order: the positional-label column first
    # (built from --metrics/--from-text/--from-ray), then each --model file.
    new_columns: List[Tuple[str, Dict[str, float]]] = []

    if args.label is not None:
        proj = _project_or_die(load_metrics(args), f"positional label '{args.label}'")
        if proj is None:
            return 2
        new_columns.append((args.label, proj))
    elif args.metrics or args.from_text or args.from_ray:
        print(
            "[error] --metrics/--from-text/--from-ray require the positional label. "
            "Use --model LABEL PATH to attach a label per local result file.",
            file=sys.stderr,
        )
        return 2

    for label, path in args.model:
        proj = _project_or_die(load_source_file(path), f"--model '{label}' ({path})")
        if proj is None:
            return 2
        new_columns.append((label, proj))

    if args.extend_xlsx:
        columns = read_existing_xlsx(Path(args.extend_xlsx))
        if not columns:
            print(f"[warn] could not parse {args.extend_xlsx}; using baseline only.", file=sys.stderr)
    else:
        columns = []

    if not columns:
        baseline_metrics = schema_baseline
        if args.baseline:
            baseline_metrics = {k: float(v) for k, v in json.loads(Path(args.baseline).read_text()).items()}
        columns = [(args.baseline_label, baseline_metrics)]

    columns.extend(new_columns)

    first_label = new_columns[0][0]
    out_path = Path(
        args.out
        or f"tmp/inhouse_dter_report/{first_label.replace('/', '_').replace('@', '_')}.xlsx"
    )
    build_workbook(columns, out_path)
    print(out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
