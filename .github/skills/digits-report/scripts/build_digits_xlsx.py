#!/usr/bin/env python3
"""Build an XLSX comparison report for the two digits validation datasets."""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATASETS = ("enus_digits_random", "enus_digits_repeat")
BASELINE_LABEL = "Qwen3.5-audio"
BASELINE_METRICS = {
    "enus_digits_random": 0.0004,
    "enus_digits_repeat": 0.0092,
}
METRIC_LINE_RE = re.compile(
    r"val-aux/(?P<dataset>enus_digits_(?:random|repeat))/cer/mean@1[:=]\s*(?P<value>[0-9.eE+-]+)"
)
HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
DATASET_FILL = PatternFill("solid", fgColor="FFF2F2F2")


def parse_metrics(text: str) -> dict[str, float]:
    metrics: dict[str, float] = {}
    for match in METRIC_LINE_RE.finditer(text):
        metrics[match["dataset"]] = float(match["value"])
    return metrics


def load_json_metrics(path: Path) -> dict[str, float]:
    raw: dict[str, Any] = json.loads(path.read_text())
    metrics: dict[str, float] = {}
    for dataset, values in raw.items():
        if dataset not in DATASETS:
            continue
        if isinstance(values, dict):
            metrics[dataset] = float(values["cer"])
        else:
            metrics[dataset] = float(values)
    return metrics


def fetch_ray_logs(node: str, job_id: str) -> str:
    remote = f"bash -l -c 'ray job logs {job_id} 2>&1'"
    result = subprocess.run(["brix", "ssh", node, "--", remote], capture_output=True, text=True, check=False)
    if result.returncode:
        print(f"[warn] ray job logs failed for {node}/{job_id}: {result.stderr[-500:]}", file=sys.stderr)
    return result.stdout


def load_metrics(args: argparse.Namespace) -> dict[str, float]:
    metrics: dict[str, float] = {}
    if args.metrics:
        metrics.update(load_json_metrics(Path(args.metrics)))
    for text_path in args.from_text:
        metrics.update(parse_metrics(Path(text_path).read_text()))
    for node, job_id in args.from_ray:
        metrics.update(parse_metrics(fetch_ray_logs(node, job_id)))
    return metrics


def build_workbook(columns: list[tuple[str, dict[str, float]]], out_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CER"
    model_columns = list(range(2, 2 + len(columns)))
    reduction_columns = list(range(2 + len(columns), 1 + len(columns) * 2))
    bold = Font(bold=True)

    sheet.cell(2, 1, "Header").font = bold
    sheet.cell(2, 2, "Baseline").font = bold
    for index, (label, _) in enumerate(columns[1:], start=3):
        sheet.cell(2, index, label).font = bold
    for column in reduction_columns:
        sheet.cell(2, column, "CER reduction").font = bold

    sheet.cell(3, 1, "Column").font = bold
    for index, column in enumerate(model_columns):
        sheet.cell(3, column, get_column_letter(index + 1)).font = bold
    for index, column in enumerate(reduction_columns, start=1):
        sheet.cell(3, column, f"A->{get_column_letter(index + 1)}").font = bold

    row = 4
    for dataset in DATASETS:
        sheet.cell(row, 1, dataset)
        for index, (_, values) in enumerate(columns):
            value = values.get(dataset)
            if value is not None:
                sheet.cell(row, model_columns[index], value)
        for index, column in enumerate(reduction_columns, start=1):
            baseline = get_column_letter(model_columns[0])
            target = get_column_letter(model_columns[index])
            sheet.cell(row, column, f"=1-{target}{row}/{baseline}{row}")
        row += 1

    center = Alignment(horizontal="center", vertical="center")
    for current_row in range(2, row):
        for column in range(2, 1 + len(columns) * 2):
            cell = sheet.cell(current_row, column)
            cell.alignment = center
            if current_row >= 4:
                cell.number_format = "0.00%"
    for current_row in (2, 3):
        for column in range(1, 1 + len(columns) * 2):
            sheet.cell(current_row, column).fill = HEADER_FILL
    for current_row in range(4, row):
        for column in range(1, 1 + len(columns) * 2):
            sheet.cell(current_row, column).fill = DATASET_FILL

    sheet.column_dimensions["A"].width = 34
    for column in range(2, 1 + len(columns) * 2):
        sheet.column_dimensions[get_column_letter(column)].width = 20
    for column in reduction_columns:
        letter = get_column_letter(column)
        sheet.conditional_formatting.add(
            f"{letter}4:{letter}{row - 1}",
            ColorScaleRule(
                start_type="num", start_value=-1, start_color="FFF8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                end_type="num", end_value=1, end_color="FF63BE7B",
            ),
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)


def read_existing_xlsx(path: Path) -> list[tuple[str, dict[str, float]]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["CER"] if "CER" in workbook.sheetnames else workbook["digits"]
    columns: list[tuple[str, dict[str, float]]] = []
    for column in range(2, sheet.max_column + 1):
        if not isinstance(sheet.cell(3, column).value, str) or not sheet.cell(3, column).value.isalpha():
            break
        label = str(sheet.cell(2, column).value)
        values: dict[str, float] = {}
        for dataset_index, dataset in enumerate(DATASETS):
            value = sheet.cell(4 + dataset_index, column).value
            if isinstance(value, (int, float)):
                values[dataset] = float(value)
        columns.append((label, values))
    return columns


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("label", help="New model label (column header).")
    parser.add_argument("--from-ray", nargs=2, metavar=("NODE", "JOB_ID"), action="append", default=[])
    parser.add_argument("--from-text", action="append", default=[])
    parser.add_argument("--metrics", help="JSON metrics (fractions, not percentages).")
    parser.add_argument("--extend-xlsx", help="Existing digits report to extend.")
    parser.add_argument("--out", help="Output path (default: tmp/digits_report/<label>.xlsx).")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    new_metrics = load_metrics(args)
    missing = [dataset for dataset in DATASETS if dataset not in new_metrics]
    if missing:
        print(f"[error] missing metrics for: {', '.join(missing)}", file=sys.stderr)
        return 2
    columns = read_existing_xlsx(Path(args.extend_xlsx)) if args.extend_xlsx else []
    if not columns:
        columns = [(BASELINE_LABEL, BASELINE_METRICS)]
    columns.append((args.label, new_metrics))
    output = Path(args.out or f"tmp/digits_report/{args.label.replace('/', '_').replace('@', '_')}.xlsx")
    build_workbook(columns, output)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())