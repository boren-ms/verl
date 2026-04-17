#!/usr/bin/env python3
"""Collect p_err and p_edge metrics from ASR eval Ray job logs and produce an Excel report.

Usage:
    # From log files:
    python collect_eval_metrics.py \
        --runs "run_label_1:path/to/log1.log" "run_label_2:path/to/log2.log" \
        --output tmp/eval_metrics.xlsx

    # From metric lines directly:
    python collect_eval_metrics.py \
        --metric-lines "run_label|step:0 - val-aux/librispeech/p_err/mean@1:0.0155 - ..." \
        --output tmp/eval_metrics.xlsx
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def parse_step_line(line: str) -> dict[str, dict[str, float]]:
    """Parse a step: log line and extract per-dataset p_err and p_edge.

    Returns: {dataset: {'p_err': float, 'p_edge': float}}
    """
    metrics: dict[str, dict[str, float]] = defaultdict(dict)
    # Pattern: val-aux/<dataset>/p_err/mean@<n>:<value>
    pattern = re.compile(r"val-aux/([^/]+)/p_(err|edge)/mean@\d+[:\s]+([\d.eE+-]+)")
    for m in pattern.finditer(line):
        dataset = m.group(1)
        metric_name = f"p_{m.group(2)}"
        value = float(m.group(3))
        metrics[dataset][metric_name] = value
    return dict(metrics)


def parse_dict_line(line: str) -> dict[str, dict[str, float]]:
    """Parse a Python dict log line and extract per-dataset p_err and p_edge.

    Returns: {dataset: {'p_err': float, 'p_edge': float}}
    """
    metrics: dict[str, dict[str, float]] = defaultdict(dict)
    # Pattern: 'val-aux/<dataset>/p_err/mean@<n>': <value>
    pattern = re.compile(r"'val-aux/([^/]+)/p_(err|edge)/mean@\d+':\s*([\d.eE+-]+)")
    for m in pattern.finditer(line):
        dataset = m.group(1)
        metric_name = f"p_{m.group(2)}"
        value = float(m.group(3))
        metrics[dataset][metric_name] = value
    return dict(metrics)


def extract_metrics_from_log(log_path: str) -> dict[str, dict[str, float]]:
    """Read a log file and extract metrics from the last step: or dict line."""
    path = Path(log_path)
    if not path.exists():
        print(f"Warning: log file not found: {log_path}", file=sys.stderr)
        return {}

    step_lines = []
    dict_lines = []
    with open(path) as f:
        for raw_line in f:
            line = raw_line.strip()
            if "step:" in line and "p_err" in line:
                step_lines.append(line)
            elif "val-aux/" in line and "p_err" in line and "{" in line:
                dict_lines.append(line)

    # Prefer the last step: line, fall back to last dict line
    if step_lines:
        return parse_step_line(step_lines[-1])
    elif dict_lines:
        return parse_dict_line(dict_lines[-1])
    else:
        print(f"Warning: no metric lines found in {log_path}", file=sys.stderr)
        return {}


def extract_metrics_from_metric_line(line: str) -> dict[str, dict[str, float]]:
    """Parse a raw metric line (step: or dict format)."""
    if "step:" in line:
        return parse_step_line(line)
    elif "{" in line:
        return parse_dict_line(line)
    return {}


def build_workbook(
    all_runs: dict[str, dict[str, dict[str, float]]],
    exp_name: str,
    date_str: str,
) -> Workbook:
    """Build an Excel workbook with p_err and p_edge sheets.

    Args:
        all_runs: {run_label: {dataset: {'p_err': float, 'p_edge': float}}}
        exp_name: Experiment name for the header row
        date_str: Date string for the header row
    """
    wb = Workbook()
    run_labels = list(all_runs.keys())

    # Collect all datasets across all runs, preserve order of first appearance
    all_datasets: list[str] = []
    seen: set[str] = set()
    for run_metrics in all_runs.values():
        for ds in run_metrics:
            if ds not in seen:
                all_datasets.append(ds)
                seen.add(ds)
    all_datasets.sort()

    for metric_idx, metric_name in enumerate(["p_err", "p_edge"]):
        if metric_idx == 0:
            ws = wb.active
            ws.title = metric_name
        else:
            ws = wb.create_sheet(title=metric_name)

        bold = Font(bold=True)
        num_fmt = "0.00"

        # Row 1: header - exp name
        ws.append(["exp", exp_name])
        ws.cell(row=1, column=1).font = bold

        # Row 2: header - date
        ws.append(["date", date_str])
        ws.cell(row=2, column=1).font = bold

        # Row 3: column headers
        header_row = ["dataset"] + run_labels
        ws.append(header_row)
        for cell in ws[3]:
            cell.font = bold

        # Data rows
        col_sums = [0.0] * len(run_labels)
        col_counts = [0] * len(run_labels)

        for ds in all_datasets:
            row = [ds]
            for j, run_label in enumerate(run_labels):
                run_data = all_runs[run_label]
                if ds in run_data and metric_name in run_data[ds]:
                    val = run_data[ds][metric_name] * 100  # Convert to percentage
                    row.append(val)
                    col_sums[j] += val
                    col_counts[j] += 1
                else:
                    row.append(None)
            ws.append(row)

        # Average row
        avg_row = ["average"]
        for j in range(len(run_labels)):
            if col_counts[j] > 0:
                avg_row.append(col_sums[j] / col_counts[j])
            else:
                avg_row.append(None)
        ws.append(avg_row)

        # Format: bold average row, number format for data cells
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = bold

        for row in ws.iter_rows(min_row=4, min_col=2, max_col=1 + len(run_labels)):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = num_fmt

        # Auto-width columns
        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    return wb


def main():
    parser = argparse.ArgumentParser(description="Collect eval metrics into Excel report")
    parser.add_argument(
        "--runs",
        nargs="*",
        metavar="LABEL:PATH",
        help="Run label and log file path pairs, e.g. 'eval_openasr_h100:logs/node/eval.log'",
    )
    parser.add_argument(
        "--metric-lines",
        nargs="*",
        metavar="LABEL|LINE",
        help="Run label and raw metric line pairs, e.g. 'run1|step:0 - ...'",
    )
    parser.add_argument(
        "--output",
        default="tmp/eval_metrics.xlsx",
        help="Output Excel path (default: tmp/eval_metrics.xlsx)",
    )
    parser.add_argument(
        "--exp-name",
        default=None,
        help="Experiment name for header (default: derived from output filename)",
    )
    parser.add_argument(
        "--date",
        default=None,
        help="Date string for header (default: today)",
    )
    args = parser.parse_args()

    if not args.runs and not args.metric_lines:
        parser.error("Provide at least one of --runs or --metric-lines")

    all_runs: dict[str, dict[str, dict[str, float]]] = {}

    # Parse from log files
    if args.runs:
        for run_spec in args.runs:
            if ":" not in run_spec:
                parser.error(f"Invalid --runs format: '{run_spec}'. Expected 'label:path'")
            label, path = run_spec.split(":", 1)
            metrics = extract_metrics_from_log(path)
            if metrics:
                all_runs[label] = metrics
            else:
                print(f"Warning: no metrics extracted for run '{label}'", file=sys.stderr)

    # Parse from raw metric lines
    if args.metric_lines:
        for line_spec in args.metric_lines:
            if "|" not in line_spec:
                parser.error(f"Invalid --metric-lines format: '{line_spec}'. Expected 'label|line'")
            label, line = line_spec.split("|", 1)
            metrics = extract_metrics_from_metric_line(line)
            if metrics:
                all_runs[label] = metrics
            else:
                print(f"Warning: no metrics parsed for run '{label}'", file=sys.stderr)

    if not all_runs:
        print("Error: no metrics extracted from any source.", file=sys.stderr)
        sys.exit(1)

    exp_name = args.exp_name or Path(args.output).stem
    date_str = args.date or date.today().isoformat()

    wb = build_workbook(all_runs, exp_name, date_str)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print(f"Runs: {', '.join(all_runs.keys())}")
    for label, metrics in all_runs.items():
        print(f"  {label}: {len(metrics)} datasets — {', '.join(sorted(metrics.keys()))}")


if __name__ == "__main__":
    main()
