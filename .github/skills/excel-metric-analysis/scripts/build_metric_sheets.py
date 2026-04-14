#!/usr/bin/env python3
from __future__ import annotations

import argparse
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class MetricSpec:
    sheet_suffix: str
    column_name: str


@dataclass
class DatasetRecord:
    values: OrderedDict[str, object] | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a workbook with WERs and EERs sheets from an Excel report whose "
            "name column follows <dataset> or <dataset>_<tag>."
        )
    )
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Directory for the generated workbook. Default: <input_dir>/analysis/<input_stem>",
    )
    parser.add_argument(
        "--output-name",
        default="metric_analysis.xlsx",
        help="Generated workbook file name inside --output-dir. Default: metric_analysis.xlsx",
    )
    parser.add_argument("--sheet", help="Read only one worksheet by name.")
    parser.add_argument(
        "--name-column",
        default="name",
        help="Column containing dataset/tag values. Default: name",
    )
    parser.add_argument(
        "--group-column",
        default="group",
        help="Column used to partition rows into output sheets. Default: group",
    )
    parser.add_argument(
        "--wer-column",
        default="wer",
        help="Column written to the WERs sheet. Default: wer",
    )
    parser.add_argument(
        "--eer-column",
        default="eer",
        help="Column written to the EERs sheet. Default: eer",
    )
    parser.add_argument(
        "--datasets",
        nargs="*",
        default=None,
        help="Explicit dataset names to disambiguate underscores.",
    )
    parser.add_argument(
        "--datasets-file",
        type=Path,
        help="Text file with one dataset name per line.",
    )
    parser.add_argument(
        "--tag-for-exact",
        default="base",
        help="Tag used when name exactly matches a dataset. Default: base",
    )
    parser.add_argument(
        "--keep-empty",
        action="store_true",
        help="Keep datasets even when they do not have a value for a metric.",
    )
    args = parser.parse_args()
    if args.output_dir is None:
        args.output_dir = args.input_xlsx.parent / "analysis" / args.input_xlsx.stem
    return args


def read_explicit_datasets(args: argparse.Namespace) -> list[str]:
    datasets: list[str] = []
    if args.datasets:
        datasets.extend(args.datasets)
    if args.datasets_file:
        datasets.extend(
            line.strip()
            for line in args.datasets_file.read_text().splitlines()
            if line.strip()
        )

    ordered: list[str] = []
    seen: set[str] = set()
    for dataset in datasets:
        if dataset not in seen:
            seen.add(dataset)
            ordered.append(dataset)
    return ordered


def load_rows(ws, header_map: dict[str, int]):
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {column: row[index] for column, index in header_map.items()}


def split_with_datasets(name: str, datasets: Iterable[str], exact_tag: str) -> tuple[str, str] | None:
    for dataset in sorted(datasets, key=len, reverse=True):
        if name == dataset:
            return dataset, exact_tag
        prefix = f"{dataset}_"
        if name.startswith(prefix):
            return dataset, name[len(prefix) :]
    return None


def infer_datasets(rows: list[dict[str, object]], name_column: str) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()

    for row in rows:
        raw_name = row.get(name_column)
        if raw_name in (None, ""):
            continue
        name = str(raw_name)
        if "_" not in name and name not in seen:
            seen.add(name)
            ordered.append(name)

    for row in rows:
        raw_name = row.get(name_column)
        if raw_name in (None, ""):
            continue
        name = str(raw_name)
        if name in seen:
            continue
        dataset = name.split("_", 1)[0]
        if dataset not in seen:
            seen.add(dataset)
            ordered.append(dataset)
    return ordered


def parse_name(name: str, datasets: list[str], exact_tag: str) -> tuple[str, str] | None:
    parsed = split_with_datasets(name, datasets, exact_tag)
    if parsed:
        return parsed
    if "_" not in name:
        return name, exact_tag
    dataset, tag = name.split("_", 1)
    if not dataset or not tag:
        return None
    return dataset, tag


def collect_rows(args: argparse.Namespace) -> tuple[list[dict[str, object]], list[str]]:
    wb = openpyxl.load_workbook(args.input_xlsx, data_only=True)
    worksheets = [wb[args.sheet]] if args.sheet else wb.worksheets
    all_rows: list[dict[str, object]] = []
    header_map_reference: dict[str, int] | None = None

    for ws in worksheets:
        header = [cell.value for cell in ws[1]]
        header_map = {str(name): idx for idx, name in enumerate(header) if name is not None}
        if header_map_reference is None:
            header_map_reference = header_map
        required = [args.name_column, args.wer_column, args.eer_column]
        missing = [column for column in required if column not in header_map]
        if missing:
            raise SystemExit(f"{ws.title}: missing required columns: {', '.join(missing)}")
        all_rows.extend(load_rows(ws, header_map))

    explicit_datasets = read_explicit_datasets(args)
    datasets = explicit_datasets or infer_datasets(all_rows, args.name_column)
    if not datasets:
        raise SystemExit("could not determine any datasets from the workbook")
    return all_rows, datasets


def collect_metric_table(
    rows: list[dict[str, object]],
    datasets: list[str],
    name_column: str,
    group_column: str,
    metric_column: str,
    exact_tag: str,
) -> tuple[OrderedDict[str, OrderedDict[str, DatasetRecord]], dict[str, list[str]], list[str]]:
    grouped_tables: OrderedDict[str, OrderedDict[str, DatasetRecord]] = OrderedDict()
    grouped_tags: dict[str, list[str]] = {}
    unmatched: list[str] = []

    for row in rows:
        raw_name = row.get(name_column)
        metric_value = row.get(metric_column)
        if raw_name in (None, "") or metric_value in (None, ""):
            continue

        parsed = parse_name(str(raw_name), datasets, exact_tag)
        if not parsed:
            unmatched.append(str(raw_name))
            continue
        dataset, tag = parsed
        if dataset not in datasets:
            unmatched.append(str(raw_name))
            continue

        group_value = row.get(group_column)
        group_text = "" if group_value in (None, "") else str(group_value)
        table = grouped_tables.setdefault(group_text, OrderedDict())
        group_tags = grouped_tags.setdefault(group_text, [])
        dataset_record = table.setdefault(dataset, DatasetRecord(values=OrderedDict()))

        assert dataset_record.values is not None
        dataset_record.values[tag] = metric_value
        if tag not in group_tags:
            group_tags.append(tag)

    return grouped_tables, grouped_tags, unmatched


def normalize_sheet_title(title: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]", "_", title)[:31] or "Sheet"
    if cleaned not in used:
        used.add(cleaned)
        return cleaned
    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        index += 1


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = max(
        14, min(40, max(len(str(cell.value or "")) for cell in ws["A"]) + 2)
    )
    for col_idx in range(2, ws.max_column + 1):
        col = get_column_letter(col_idx)
        width = max(len(str(ws.cell(1, col_idx).value or "")) + 2, 12)
        ws.column_dimensions[col].width = min(width, 30)


def write_output(
    output_xlsx: Path,
    datasets: list[str],
    args: argparse.Namespace,
    metric_tables: list[tuple[MetricSpec, OrderedDict[str, OrderedDict[str, DatasetRecord]], dict[str, list[str]]]],
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()

    for spec, grouped_tables, grouped_tags in metric_tables:
        for group_name, table in grouped_tables.items():
            sheet_base = f"{group_name or 'ungrouped'}_{spec.sheet_suffix}"
            ws = wb.create_sheet(normalize_sheet_title(sheet_base, used_titles))
            tags = grouped_tags[group_name]
            ws.append(["dataset", *tags])
            dataset_iterable = datasets if args.keep_empty else [dataset for dataset in datasets if dataset in table]

            for dataset in dataset_iterable:
                record = table.get(dataset)
                values = record.values if record and record.values is not None else {}
                ws.append([dataset, *[values.get(tag) for tag in tags]])
            style_sheet(ws)

    wb.save(output_xlsx)


def main() -> None:
    args = parse_args()
    rows, datasets = collect_rows(args)

    metric_specs = [
        MetricSpec(sheet_suffix="WER", column_name=args.wer_column),
        MetricSpec(sheet_suffix="EER", column_name=args.eer_column),
    ]
    metric_tables: list[
        tuple[MetricSpec, OrderedDict[str, OrderedDict[str, DatasetRecord]], dict[str, list[str]]]
    ] = []
    unmatched_by_metric: dict[str, list[str]] = {}

    for spec in metric_specs:
        table, tags, unmatched = collect_metric_table(
            rows=rows,
            datasets=datasets,
            name_column=args.name_column,
            group_column=args.group_column,
            metric_column=spec.column_name,
            exact_tag=args.tag_for_exact,
        )
        metric_tables.append((spec, table, tags))
        if unmatched:
            unmatched_by_metric[spec.column_name] = unmatched

    args.output_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = args.output_dir / args.output_name
    write_output(output_xlsx, datasets, args, metric_tables)

    for metric_name, unmatched in unmatched_by_metric.items():
        preview = ", ".join(unmatched[:10])
        extra = "" if len(unmatched) <= 10 else f" ... (+{len(unmatched) - 10} more)"
        print(f"Ignored {len(unmatched)} unmatched rows for {metric_name}: {preview}{extra}")
    print(output_xlsx)


if __name__ == "__main__":
    main()
