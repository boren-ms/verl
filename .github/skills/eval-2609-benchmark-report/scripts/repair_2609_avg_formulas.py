#!/usr/bin/env python3
"""Repair average formulas and summary aggregates in existing 2609 reports."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from summary_charts import apply_summary_charts

PCT_FMT = "0.00%"
METRIC_TITLES = {"CER", "DTER", "TER", "WER"}


@dataclass
class SheetResult:
    value_columns: list[int]
    overalls: dict[str, dict[int, float]]
    formula_cells: int = 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbooks", nargs="+", type=Path)
    return parser.parse_args()


def _mean(values: list[float]) -> Optional[float]:
    return sum(values) / len(values) if values else None


def _metric_from_sheet(title: str) -> str:
    return "WER" if "openasr" in title.lower() else "DTER"


def _average_formula(column: int, ranges: list[tuple[int, int]]) -> str:
    letter = get_column_letter(column)
    refs = ",".join(f"{letter}{start}:{letter}{end}" for start, end in ranges)
    return f"=AVERAGE({refs})"


def _numeric_average(worksheet, column: int, rows: list[int]) -> Optional[float]:
    values = [
        float(worksheet.cell(row, column).value)
        for row in rows
        if isinstance(worksheet.cell(row, column).value, (int, float))
    ]
    return _mean(values)


def repair_benchmark_sheet(worksheet) -> Optional[SheetResult]:
    if worksheet.cell(3, 1).value != "Column":
        return None
    if not any(
        isinstance(worksheet.cell(row, 1).value, str)
        and worksheet.cell(row, 1).value.lower().endswith("avg")
        for row in range(4, worksheet.max_row + 1)
    ):
        return None

    column_headers = {
        column: worksheet.cell(3, column).value
        for column in range(2, worksheet.max_column + 1)
    }
    value_columns = [
        column
        for column, header in column_headers.items()
        if header is not None and "->" not in str(header)
    ]
    delta_columns = [
        column
        for column, header in column_headers.items()
        if header is not None and "->" in str(header)
    ]
    if not value_columns:
        raise ValueError(f"{worksheet.title}: no value columns found")
    if len(delta_columns) != max(0, len(value_columns) - 1):
        raise ValueError(
            f"{worksheet.title}: expected {len(value_columns) - 1} delta columns, "
            f"found {len(delta_columns)}"
        )

    result = SheetResult(value_columns=value_columns, overalls={})
    current_metric = _metric_from_sheet(worksheet.title)
    group_rows: list[int] = []
    section_rows: list[int] = []
    section_ranges: list[tuple[int, int]] = []

    for row in range(4, worksheet.max_row + 1):
        label = worksheet.cell(row, 1).value
        if not isinstance(label, str):
            continue
        normalized = label.strip()
        if normalized == "Header" and result.overalls:
            break
        if normalized.upper() in METRIC_TITLES:
            current_metric = normalized.upper()
            group_rows = []
            section_rows = []
            section_ranges = []
            continue

        lower = normalized.lower()
        if not lower.endswith("avg"):
            group_rows.append(row)
            section_rows.append(row)
            continue
        if not group_rows and "overall avg" not in lower:
            raise ValueError(f"{worksheet.title}!A{row}: average row has no dataset rows")

        if "overall avg" in lower:
            metric = normalized.split()[0].upper() if normalized.upper().split()[0] in METRIC_TITLES else current_metric
            result.overalls.setdefault(metric, {})
            for column in value_columns:
                average = _numeric_average(worksheet, column, section_rows)
                if average is None:
                    worksheet.cell(row, column).value = None
                    continue
                worksheet.cell(row, column, _average_formula(column, section_ranges)).number_format = PCT_FMT
                result.overalls[metric][column] = average
                result.formula_cells += 1
            for candidate_column, delta_column in zip(value_columns[1:], delta_columns, strict=True):
                if candidate_column in result.overalls[metric] and value_columns[0] in result.overalls[metric]:
                    candidate = get_column_letter(candidate_column)
                    baseline = get_column_letter(value_columns[0])
                    worksheet.cell(row, delta_column, f"=1-{candidate}{row}/{baseline}{row}").number_format = PCT_FMT
                    result.formula_cells += 1
                else:
                    worksheet.cell(row, delta_column).value = None
            group_rows = []
            section_rows = []
            section_ranges = []
            continue

        group_range = (group_rows[0], group_rows[-1])
        section_ranges.append(group_range)
        for column in value_columns:
            if _numeric_average(worksheet, column, group_rows) is None:
                worksheet.cell(row, column).value = None
                continue
            worksheet.cell(row, column, _average_formula(column, [group_range])).number_format = PCT_FMT
            result.formula_cells += 1
        for candidate_column, delta_column in zip(value_columns[1:], delta_columns, strict=True):
            baseline_column = value_columns[0]
            if (
                _numeric_average(worksheet, baseline_column, group_rows) is not None
                and _numeric_average(worksheet, candidate_column, group_rows) is not None
            ):
                candidate = get_column_letter(candidate_column)
                baseline = get_column_letter(baseline_column)
                worksheet.cell(row, delta_column, f"=1-{candidate}{row}/{baseline}{row}").number_format = PCT_FMT
                result.formula_cells += 1
            else:
                worksheet.cell(row, delta_column).value = None
        group_rows = []

    return result


def _result_for(results: dict[str, SheetResult], sheet_name: str, metric: str) -> tuple[SheetResult, dict[int, float]]:
    result = results[sheet_name]
    values = result.overalls.get(metric.upper())
    if values is None and len(result.overalls) == 1:
        values = next(iter(result.overalls.values()))
    if values is None:
        raise ValueError(f"{sheet_name}: no overall values for {metric}")
    return result, values


def _ensure_summary_delta_colors(worksheet, header_row: int, data_start: int, data_end: int) -> None:
    if data_end < data_start:
        return
    worksheet.conditional_formatting._cf_rules.clear()
    for column in range(1, worksheet.max_column + 1):
        if str(worksheet.cell(header_row, column).value).lower() != "delta":
            continue
        letter = get_column_letter(column)
        cell_range = f"{letter}{data_start}:{letter}{data_end}"
        worksheet.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="num", start_value=-0.1, start_color="FFF8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                end_type="num", end_value=0.1, end_color="FF63BE7B",
            ),
        )


def repair_summary(workbook, results: dict[str, SheetResult]) -> None:
    if "summary" not in workbook.sheetnames:
        return
    worksheet = workbook["summary"]

    if worksheet.cell(3, 1).value == "Checkpoint":
        data_rows = [
            row
            for row in range(4, worksheet.max_row + 1)
            if isinstance(worksheet.cell(row, 1).value, str)
            and isinstance(worksheet.cell(row, 2).value, str)
            and isinstance(worksheet.cell(row, 3).value, str)
        ]
        for row in data_rows:
            checkpoint = worksheet.cell(row, 1).value
            benchmark = worksheet.cell(row, 2).value
            metric = worksheet.cell(row, 3).value
            if not all(isinstance(value, str) for value in (checkpoint, benchmark, metric)):
                continue
            sheet_name = f"{checkpoint}_{benchmark}"
            result, values = _result_for(results, sheet_name, metric)
            baseline = values.get(result.value_columns[0])
            candidate = values.get(result.value_columns[1])
            worksheet.cell(row, 4, baseline).number_format = PCT_FMT
            worksheet.cell(row, 5, candidate).number_format = PCT_FMT
            worksheet.cell(row, 6, 1 - candidate / baseline if baseline not in (None, 0) and candidate is not None else None).number_format = PCT_FMT
        _ensure_summary_delta_colors(worksheet, header_row=3, data_start=4, data_end=max(data_rows, default=3))
        return

    if worksheet.cell(1, 1).value != "Benchmark":
        return
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    if "Metric definition" not in headers:
        benchmark_column = headers.index("Benchmark") + 1
        metric_column = headers.index("Metric") + 1 if "Metric" in headers else None
        baseline_column = next(
            column
            for column, header in enumerate(headers, start=1)
            if str(header).lower() == "baseline" or str(header).startswith("Baseline (")
        )
        candidate_column = next(
            column
            for column, header in enumerate(headers, start=1)
            if str(header).lower() == "candidate" or str(header).startswith("Candidate (")
        )
        delta_column = next(
            column
            for column, header in enumerate(headers, start=1)
            if str(header).lower() == "delta"
        )
        data_rows = [
            row
            for row in range(2, worksheet.max_row + 1)
            if isinstance(worksheet.cell(row, benchmark_column).value, str)
        ]
        for row in data_rows:
            benchmark = worksheet.cell(row, benchmark_column).value
            if benchmark not in results:
                continue
            metric = (
                worksheet.cell(row, metric_column).value
                if metric_column is not None
                else _metric_from_sheet(str(benchmark))
            )
            result, values = _result_for(results, str(benchmark), str(metric))
            baseline = values.get(result.value_columns[0])
            candidate = values.get(result.value_columns[1])
            worksheet.cell(row, baseline_column, baseline).number_format = PCT_FMT
            worksheet.cell(row, candidate_column, candidate).number_format = PCT_FMT
            worksheet.cell(
                row,
                delta_column,
                1 - candidate / baseline
                if baseline not in (None, 0) and candidate is not None
                else None,
            ).number_format = PCT_FMT
        _ensure_summary_delta_colors(
            worksheet,
            header_row=1,
            data_start=2,
            data_end=max(data_rows, default=1),
        )
        return
    metric_definition_column = headers.index("Metric definition") + 1
    summary_value_columns = [3] + [
        column
        for column in range(4, metric_definition_column)
        if str(worksheet.cell(1, column).value).lower() != "delta"
    ]
    summary_delta_columns = [
        column
        for column in range(4, metric_definition_column)
        if str(worksheet.cell(1, column).value).lower() == "delta"
    ]
    data_rows = [
        row
        for row in range(2, worksheet.max_row + 1)
        if isinstance(worksheet.cell(row, 1).value, str)
        and isinstance(worksheet.cell(row, 2).value, str)
    ]
    for row in data_rows:
        benchmark = worksheet.cell(row, 1).value
        metric = worksheet.cell(row, 2).value
        if not isinstance(benchmark, str) or not isinstance(metric, str) or benchmark not in results:
            continue
        result, values = _result_for(results, benchmark, metric)
        if len(summary_value_columns) != len(result.value_columns):
            if len(summary_value_columns) > len(result.value_columns):
                raise ValueError(f"summary/{benchmark}: value-column count mismatch")
            sheet_value_columns = [result.value_columns[0], result.value_columns[-1]]
        else:
            sheet_value_columns = result.value_columns
        for summary_column, sheet_column in zip(summary_value_columns, sheet_value_columns, strict=True):
            worksheet.cell(row, summary_column, values.get(sheet_column)).number_format = PCT_FMT
        baseline = values.get(sheet_value_columns[0])
        for summary_column, sheet_column in zip(summary_delta_columns, sheet_value_columns[1:], strict=True):
            candidate = values.get(sheet_column)
            worksheet.cell(
                row,
                summary_column,
                1 - candidate / baseline if baseline not in (None, 0) and candidate is not None else None,
            ).number_format = PCT_FMT
    _ensure_summary_delta_colors(worksheet, header_row=1, data_start=2, data_end=max(data_rows, default=1))


def ensure_valid_style_fills(workbook) -> None:
    max_fill_id = max((style.fillId for style in workbook._cell_styles), default=0)
    repair_fills = {
        3: PatternFill("solid", fgColor="FFFFF2CC"),
        4: PatternFill("solid", fgColor="FFE2EFDA"),
    }
    while len(workbook._fills) <= max_fill_id:
        fill_id = len(workbook._fills)
        if fill_id not in repair_fills:
            raise ValueError(f"cannot repair missing fillId={fill_id}")
        workbook._fills.append(repair_fills[fill_id])


def repair_workbook(path: Path) -> tuple[int, int]:
    workbook = load_workbook(path, data_only=False)
    results: dict[str, SheetResult] = {}
    for worksheet in workbook.worksheets:
        result = repair_benchmark_sheet(worksheet)
        if result is not None:
            results[worksheet.title] = result
    if not results:
        raise ValueError(f"{path}: no benchmark sheets found")
    repair_summary(workbook, results)
    if "summary" in workbook.sheetnames:
        apply_summary_charts(workbook["summary"])
    ensure_valid_style_fills(workbook)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)
    return len(results), sum(result.formula_cells for result in results.values())


def main() -> int:
    repaired_sheets = 0
    formula_cells = 0
    for path in parse_args().workbooks:
        if path.name.startswith("~$"):
            continue
        sheets, formulas = repair_workbook(path)
        repaired_sheets += sheets
        formula_cells += formulas
        print(f"repaired {path}: {sheets} sheets, {formulas} formulas")
    print(f"repaired {repaired_sheets} benchmark sheets with {formula_cells} formulas")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())