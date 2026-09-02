#!/usr/bin/env python3
"""Collect existing 2609 benchmark eval results into a single XLSX workbook.

One sheet per benchmark, MixLang on its own separate sheet. Every sheet compares
a candidate model column (B) against the config reference baseline column (A) and
computes the relative error-rate reduction (delta) exactly like
``inhouse-dter-report``:

    delta = 1 - candidate / baseline        (positive => candidate improved)

The delta column is labelled per metric: TERR for DTER/TER, WERR for WER,
CERR for CER.

Benchmarks and their result sources ("reuse existing result"):

  inhouse_dter  long_eval_inhouse_2609_all_seg30   micro-DTER   measures tree
  digits_enus   eval_digits_enus_2609              CER + WER    ray/text/json
  openasr_ml    eval_openasr_ml_verb_2609          WER (p_err)  ray/text/json
  mixlang       long_eval_mixlang_fy26q2_zh_seg_2609  DTER       measures tree
  digits_tier1  eval_digits_tier1_2609 (optional)  CER + WER    ray/text/json

A "source" string is auto-detected as one of:
  * ``az://.../<root>/`` or a local directory  -> tree of ``<slug>/measures.json``
  * ``*.json``                                 -> {ds: value} or {ds: {metric: value}}
  * ``ray:<node>:<job_id>``                    -> ``ray job logs`` then parsed as text
  * any other existing file                    -> text log with ``val-aux/...`` lines

Only benchmarks whose candidate source is supplied get a sheet. If a benchmark's
baseline source is omitted and an embedded baseline exists, that embedded
baseline is used for column A.

Requires ``openpyxl`` (use ``/home/boren/.virtualenvs/openai/bin/python``).
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from summary_charts import apply_summary_charts

# ---------------------------------------------------------------------------
# Styling (mirrors inhouse-dter-report)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")   # light blue
METRIC_FILL = PatternFill("solid", fgColor="FFDDEBF7")    # lighter blue
GROUP_AVG_FILL = PatternFill("solid", fgColor="FFFFF2CC")  # light yellow
OVERALL_FILL = PatternFill("solid", fgColor="FFE2EFDA")   # light green
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")
PCT_FMT = "0.00%"

# Delta (error-rate reduction) header per metric.
DELTA_LABEL = {"dter": "TERR", "ter": "TERR", "wer": "WERR", "cer": "CERR"}

# ---------------------------------------------------------------------------
# Benchmark schemas: (metric(s), groups). Groups are (group_name, [(key, display)]).
# ``key`` matches the eval ``data_source`` / measures-tree slug.
# ---------------------------------------------------------------------------
INHOUSE_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    ("en-US", [
        ("enus_conv_fy21q1", "en-US_Conversation_DTEST_FY21Q1"),
        ("enus_conv_om_fy25q3", "en-US_Conversation_OnlineMeetings_DTEST_FY25Q3"),
        ("enus_dict_office_fy24q3", "en-US_Dictation_Commonset_OfficeOffline_FY24Q3"),
    ]),
    ("nl-NL", [
        ("nlnl_conv_fy23q2", "nl-NL_Conversation_DTEST_FY23Q2"),
        ("nlnl_conv_om_fy23q1", "nl-NL_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("nlnl_dict_fy23q4", "nl-NL_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("da-DK", [
        ("dadk_conv_fy21q3", "da-DK_Conversation_DTEST_FY21Q3"),
        ("dadk_conv_om_fy23q1", "da-DK_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("dadk_dict_fy23q4", "da-DK_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("hu-HU", [
        ("huhu_conv_fy22q4", "hu-HU_Conversation_DTEST_FY22Q4"),
        ("huhu_conv_om_fy24q2", "hu-HU_Conversation_OnlineMeetings_DTEST_FY24Q2"),
        ("huhu_dict_fy25q2", "hu-HU_Dictation_DTEST_L_D_FY25Q2"),
    ]),
    ("nb-NO", [
        ("nbno_conv_fy21q3", "nb-NO_Conversation_DTEST_FY21Q3"),
        ("nbno_conv_om_fy23q1", "nb-NO_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("nbno_dict_fy23q4", "nb-NO_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("cs-CZ", [
        ("cscz_conv_fy23q2", "cs-CZ_Conversation_DTEST_FY23Q2"),
        ("cscz_conv_om_fy24q2", "cs-CZ_Conversation_OnlineMeetings_DTEST_FY24Q2"),
        ("cscz_dict_fy24q2", "cs-CZ_Dictation_DTEST_L_D_FY24Q2"),
    ]),
]

# Embedded 2609v1,LID baseline, micro-DTER per slug.
INHOUSE_BASELINE: dict[str, float] = {
    "enus_conv_fy21q1": 0.1875, "enus_conv_om_fy25q3": 0.1404, "enus_dict_office_fy24q3": 0.0995,
    "nlnl_conv_fy23q2": 0.2275, "nlnl_conv_om_fy23q1": 0.2330, "nlnl_dict_fy23q4": 0.1086,
    "dadk_conv_fy21q3": 0.2242, "dadk_conv_om_fy23q1": 0.2110, "dadk_dict_fy23q4": 0.1689,
    "huhu_conv_fy22q4": 0.1834, "huhu_conv_om_fy24q2": 0.1662, "huhu_dict_fy25q2": 0.1852,
    "nbno_conv_fy21q3": 0.2204, "nbno_conv_om_fy23q1": 0.1612, "nbno_dict_fy23q4": 0.1635,
    "cscz_conv_fy23q2": 0.2421, "cscz_conv_om_fy24q2": 0.1424, "cscz_dict_fy24q2": 0.1335,
}

DIGITS_ENUS_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    ("en-US", [
        ("enus_digits_random", "enus_digits_random"),
        ("enus_digits_repeat", "enus_digits_repeat"),
    ]),
]

OPENASR_ML_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    ("de", [("de_fleurs", "de_fleurs"), ("de_mcv", "de_mcv")]),
    ("es", [("es_fleurs", "es_fleurs"), ("es_mcv", "es_mcv"), ("es_mls", "es_mls")]),
    ("fr", [("fr_fleurs", "fr_fleurs"), ("fr_mcv", "fr_mcv"), ("fr_mls", "fr_mls")]),
    ("it", [("it_fleurs", "it_fleurs"), ("it_mcv", "it_mcv"), ("it_mls", "it_mls")]),
    ("pt", [("pt_fleurs", "pt_fleurs"), ("pt_mls", "pt_mls")]),
]

OPENASR_ML_BASELINE: dict[str, float] = {
    "de_fleurs": 0.0260, "de_mcv": 0.0202,
    "es_fleurs": 0.0272, "es_mcv": 0.0234, "es_mls": 0.0290,
    "fr_fleurs": 0.0336, "fr_mcv": 0.0436, "fr_mls": 0.0272,
    "it_fleurs": 0.0152, "it_mcv": 0.0182, "it_mls": 0.0473,
    "pt_fleurs": 0.0297, "pt_mls": 0.0392,
}

MIXLANG_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    ("zh-cn", [("mixlang_fy26q2", "Dictation_SimuMixedLang_DTEST_FY26Q2")]),
]

MIXLANG_BASELINE: dict[str, float] = {"mixlang_fy26q2": 0.2124}

_TIER1_LOCALES = [
    ("de-DE", "dede"), ("en-US", "enus"), ("es-ES", "eses"), ("fr-FR", "frfr"),
    ("it-IT", "itit"), ("ja-JP", "jajp"), ("ko-KR", "kokr"), ("pt-BR", "ptbr"),
    ("zh-CN", "zhcn"),
]
DIGITS_TIER1_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    (loc, [(f"{p}_digits_random", f"{p}_digits_random"),
           (f"{p}_digits_repeat", f"{p}_digits_repeat")])
    for loc, p in _TIER1_LOCALES
]


# ---------------------------------------------------------------------------
# Source collection
# ---------------------------------------------------------------------------
def fetch_ray_logs(node: str, job_id: str) -> str:
    remote = f"bash -l -c 'ray job logs {job_id} 2>&1'"
    proc = subprocess.run(["brix", "ssh", node, "--", remote],
                          capture_output=True, text=True, check=False)
    if proc.returncode:
        print(f"[warn] ray job logs failed for {node}/{job_id}: {proc.stderr[-400:]}", file=sys.stderr)
    return proc.stdout


def _measures_metric(m: Dict, metric: str) -> Optional[float]:
    if metric in ("dter", "ter"):
        v = m.get("dter_p_err")
        if isinstance(v, (int, float)):
            return float(v)
        ne, nr = m.get("dter_n_err"), m.get("dter_n_ref")
        if isinstance(ne, (int, float)) and isinstance(nr, (int, float)) and nr > 0:
            return float(ne) / float(nr)
        v = m.get("dter") if metric == "dter" else m.get("ter")
        return float(v) if isinstance(v, (int, float)) else None
    v = m.get("p_err") if metric == "wer" else m.get(metric)
    return float(v) if isinstance(v, (int, float)) else None


def read_measures_tree(root: str, metrics: List[str]) -> Dict[str, Dict[str, float]]:
    """Read ``<root>/<slug>/measures.json`` into ``{slug: {metric: value}}``.

    Supports local directories and ``az://`` URLs (via ``bbb ls`` / ``bbb cat``).
    """
    out: Dict[str, Dict[str, float]] = {}

    def ingest(slug: str, text: str) -> None:
        try:
            m = json.loads(text)
        except json.JSONDecodeError:
            print(f"[warn] could not parse measures.json for {slug}", file=sys.stderr)
            return
        vals = {mt: _measures_metric(m, mt) for mt in metrics}
        vals = {k: v for k, v in vals.items() if v is not None}
        if any(mt in ("dter", "ter") for mt in metrics):
            ne, nr = m.get("dter_n_err"), m.get("dter_n_ref")
            if isinstance(ne, (int, float)) and isinstance(nr, (int, float)) and nr > 0:
                vals["dter_n_err"] = float(ne)
                vals["dter_n_ref"] = float(nr)
        if vals:
            out[slug] = vals

    if root.startswith("az://"):
        listing = subprocess.run(["bbb", "ls", root.rstrip("/") + "/"],
                                 capture_output=True, text=True, check=False)
        if listing.returncode:
            print(f"[warn] bbb ls {root} failed:\n{listing.stderr[-400:]}", file=sys.stderr)
            return out
        raw = listing.stdout.replace("\n", "")
        entries = [tok.strip() for tok in re.split(r"(?=az://)", raw) if tok.startswith("az://")]
        for entry in entries:
            if not entry.endswith("/"):
                continue
            slug = entry.rstrip("/").rsplit("/", 1)[-1]
            url = f"{root.rstrip('/')}/{slug}/measures.json"
            proc = subprocess.run(["bbb", "cat", url], capture_output=True, text=True, check=False)
            if proc.returncode:
                continue
            ingest(slug, proc.stdout)
        return out

    base = Path(root)
    if not base.is_dir():
        print(f"[warn] {root} is not a directory", file=sys.stderr)
        return out
    for child in sorted(base.iterdir()):
        mfile = child / "measures.json"
        if child.is_dir() and mfile.exists():
            ingest(child.name, mfile.read_text())
    return out


_VAL_METRIC_RE = re.compile(
    r"val-(?:aux|core)/(?P<ds>[A-Za-z0-9_\-]+)/(?P<metric>cer|p_err|dter_p_err|dter_n_err|dter_n_ref)/mean@1[:=]\s*(?P<value>[0-9.eE+\-]+)"
)
_DTER_SUMMARY_RE = re.compile(
    r"\[(?P<ds>[A-Za-z0-9_\-]+)\]\s*DTER:\s*[0-9.]+%\s*\[(?P<ne>\d+)\s*/\s*(?P<nr>\d+)\]"
)


def parse_text(
    text: str,
    metrics: List[str],
) -> Dict[str, Dict[str, float]]:
    """Parse ``val-aux/...`` metric lines into per-dataset metrics."""
    raw: Dict[str, Dict[str, float]] = {}
    for mt in _VAL_METRIC_RE.finditer(text):
        raw.setdefault(mt["ds"], {})[mt["metric"]] = float(mt["value"])
    for mt in _DTER_SUMMARY_RE.finditer(text):
        ne, nr = float(mt["ne"]), float(mt["nr"])
        if nr > 0:
            raw.setdefault(mt["ds"], {})["dter"] = ne / nr

    out: Dict[str, Dict[str, float]] = {}
    for ds, mv in raw.items():
        vals: Dict[str, float] = {}
        for metric in metrics:
            if metric in ("dter", "ter"):
                if "dter_p_err" in mv:
                    vals[metric] = mv["dter_p_err"]
                elif "dter_n_err" in mv and mv.get("dter_n_ref", 0) > 0:
                    vals[metric] = mv["dter_n_err"] / mv["dter_n_ref"]
            elif metric == "wer":
                if "p_err" in mv:
                    vals[metric] = mv["p_err"]
            elif metric in mv:
                vals[metric] = mv[metric]
        if vals:
            out[ds] = vals
    return out


def load_json_text(text: str, metrics: List[str]) -> Dict[str, Dict[str, float]]:
    raw = json.loads(text)
    out: Dict[str, Dict[str, float]] = {}
    for ds, val in raw.items():
        if isinstance(val, dict):
            vals = {mt: float(val[mt]) for mt in metrics if mt in val}
        else:  # single scalar => first requested metric
            vals = {metrics[0]: float(val)}
        if vals:
            out[ds] = vals
    return out


def load_json_source(path: Path, metrics: List[str]) -> Dict[str, Dict[str, float]]:
    return load_json_text(path.read_text(), metrics)


def collect(
    source: Optional[str],
    metrics: List[str],
) -> Dict[str, Dict[str, float]]:
    if not source:
        return {}
    if source.startswith("ray:"):
        parts = source.split(":", 2)
        if len(parts) != 3:
            print(f"[warn] bad ray source '{source}', expected ray:<node>:<job_id>", file=sys.stderr)
            return {}
        return parse_text(fetch_ray_logs(parts[1], parts[2]), metrics)
    if source.startswith("az://") and source.endswith((".json", ".log", ".txt")):
        proc = subprocess.run(["bbb", "cat", source], capture_output=True, text=True, check=False)
        if proc.returncode:
            print(f"[warn] bbb cat {source} failed:\n{proc.stderr[-400:]}", file=sys.stderr)
            return {}
        if source.endswith(".json"):
            return load_json_text(proc.stdout, metrics)
        return parse_text(proc.stdout, metrics)
    if source.startswith("az://"):
        return read_measures_tree(source, metrics)
    p = Path(source)
    if p.is_dir():
        return read_measures_tree(source, metrics)
    if p.is_file():
        if source.endswith(".json"):
            return load_json_source(p, metrics)
        return parse_text(p.read_text(), metrics)
    print(f"[warn] source not found: {source}", file=sys.stderr)
    return {}


# ---------------------------------------------------------------------------
# Sheet building
# ---------------------------------------------------------------------------
def _mean(values: List[float]) -> Optional[float]:
    vals = [v for v in values if v is not None]
    return sum(vals) / len(vals) if vals else None


def build_benchmark_sheet(
    wb: Workbook,
    title: str,
    metrics: List[str],
    groups: List[Tuple[str, List[Tuple[str, str]]]],
    baseline: Dict[str, Dict[str, float]],
    candidate: Dict[str, Dict[str, float]],
    baseline_label: str,
    candidate_label: str,
) -> Dict[str, float]:
    """Render one benchmark sheet. Returns {metric: overall candidate value}."""
    ws = wb.create_sheet(title[:31])
    ws.cell(2, 1, "Header").font = BOLD
    ws.cell(2, 2, baseline_label).font = BOLD
    ws.cell(2, 3, candidate_label).font = BOLD
    delta_hdr = "/".join(DELTA_LABEL.get(m, "delta") for m in metrics)
    ws.cell(2, 4, delta_hdr).font = BOLD
    ws.cell(3, 1, "Column").font = BOLD
    for col, lab in ((2, "A"), (3, "B"), (4, "A->B")):
        ws.cell(3, col, lab).font = BOLD
    for cell in (ws.cell(2, 1), ws.cell(2, 2), ws.cell(2, 3), ws.cell(2, 4),
                 ws.cell(3, 1), ws.cell(3, 2), ws.cell(3, 3), ws.cell(3, 4)):
        cell.fill = HEADER_FILL

    row = 4
    delta_ranges: List[Tuple[int, int]] = []
    overall_candidate: Dict[str, float] = {}

    for metric in metrics:
        delta_label = DELTA_LABEL.get(metric, "delta")
        # Metric section title (only annotate when multiple metrics stacked).
        if len(metrics) > 1:
            c = ws.cell(row, 1, metric.upper())
            c.font = BOLD
            c.fill = METRIC_FILL
            for col in (2, 3, 4):
                ws.cell(row, col).fill = METRIC_FILL
            row += 1
        delta_start = row
        all_base: List[float] = []
        all_cand: List[float] = []
        data_ranges: List[Tuple[int, int]] = []
        for group_name, datasets in groups:
            g_base: List[float] = []
            g_cand: List[float] = []
            group_start = row
            for key, display in datasets:
                ws.cell(row, 1, display)
                b = baseline.get(key, {}).get(metric)
                v = candidate.get(key, {}).get(metric)
                if b is not None:
                    ws.cell(row, 2, b).number_format = PCT_FMT
                    g_base.append(b)
                    all_base.append(b)
                if v is not None:
                    ws.cell(row, 3, v).number_format = PCT_FMT
                    g_cand.append(v)
                    all_cand.append(v)
                if b is not None and v is not None and b != 0:
                    dc = ws.cell(row, 4, f"=1-C{row}/B{row}")
                    dc.number_format = PCT_FMT
                row += 1
            group_end = row - 1
            data_ranges.append((group_start, group_end))
            # per-group average row
            ws.cell(row, 1, f"{group_name} avg").font = BOLD
            if g_base:
                ws.cell(row, 2, f"=AVERAGE(B{group_start}:B{group_end})").number_format = PCT_FMT
            if g_cand:
                ws.cell(row, 3, f"=AVERAGE(C{group_start}:C{group_end})").number_format = PCT_FMT
            if g_base and g_cand:
                ws.cell(row, 4, f"=1-C{row}/B{row}").number_format = PCT_FMT
            for col in (1, 2, 3, 4):
                ws.cell(row, col).fill = GROUP_AVG_FILL
            row += 1
        # overall average row
        ob = _mean(all_base)
        oc = _mean(all_cand)
        label = "overall avg" if len(metrics) == 1 else f"{metric.upper()} overall avg"
        ws.cell(row, 1, label).font = BOLD
        if ob is not None:
            ranges = ",".join(f"B{start}:B{end}" for start, end in data_ranges)
            ws.cell(row, 2, f"=AVERAGE({ranges})").number_format = PCT_FMT
        if oc is not None:
            ranges = ",".join(f"C{start}:C{end}" for start, end in data_ranges)
            ws.cell(row, 3, f"=AVERAGE({ranges})").number_format = PCT_FMT
            overall_candidate[metric] = oc
        if ob not in (None, 0) and oc is not None:
            ws.cell(row, 4, f"=1-C{row}/B{row}").number_format = PCT_FMT
        for col in (1, 2, 3, 4):
            ws.cell(row, col).fill = OVERALL_FILL
        row += 1
        delta_ranges.append((delta_start, row - 1))

    # Column widths + centering + delta color scale.
    ws.column_dimensions["A"].width = 46
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14
    for r in range(2, row):
        for col in (2, 3, 4):
            ws.cell(r, col).alignment = CENTER
    scale = ColorScaleRule(
        start_type="num", start_value=-0.1, start_color="FFF8696B",
        mid_type="num", mid_value=0, mid_color="FFFFFFFF",
        end_type="num", end_value=0.1, end_color="FF63BE7B",
    )
    for start, end in delta_ranges:
        ws.conditional_formatting.add(f"D{start}:D{end}", scale)
    return overall_candidate


def build_summary_sheet(wb: Workbook, rows: List[Dict[str, object]],
                        baseline_label: str, candidate_label: str) -> None:
    ws = wb.create_sheet("summary")
    headers = ["Benchmark", "Metric", f"Baseline ({baseline_label})",
               f"Candidate ({candidate_label})", "delta", "Metric definition",
               "Config", "Reference model", "Candidate model",
               "Baseline result", "Candidate result",
               "Artifacts sidecar (local)", "Artifacts sidecar (remote)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.alignment = CENTER
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r["benchmark"])
        ws.cell(i, 2, r["metric"])
        if r.get("baseline") is not None:
            ws.cell(i, 3, r["baseline"]).number_format = PCT_FMT
        if r.get("candidate") is not None:
            ws.cell(i, 4, r["candidate"]).number_format = PCT_FMT
        if r.get("delta") is not None:
            ws.cell(i, 5, r["delta"]).number_format = PCT_FMT
        ws.cell(i, 6, r.get("metric_definition", ""))
        ws.cell(i, 7, r.get("config", ""))
        ws.cell(i, 8, r.get("reference_model", ""))
        ws.cell(i, 9, r.get("candidate_model", ""))
        ws.cell(i, 10, r.get("baseline_source", ""))
        ws.cell(i, 11, r.get("candidate_source", ""))
        ws.cell(i, 12, r.get("artifacts_sidecar_local", ""))
        ws.cell(i, 13, r.get("artifacts_sidecar_remote", ""))
    widths = [16, 8, 22, 24, 10, 42, 52, 56, 56, 52, 52, 58, 72]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=2, max_row=1 + len(rows), min_col=2, max_col=5):
        for cell in row:
            cell.alignment = CENTER
    if rows:
        ws.conditional_formatting.add(
            f"E2:E{1 + len(rows)}",
            ColorScaleRule(
                start_type="num", start_value=-0.1, start_color="FFF8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                end_type="num", end_value=0.1, end_color="FF63BE7B",
            ),
        )
    apply_summary_charts(ws)


# ---------------------------------------------------------------------------
# Benchmark registry
# ---------------------------------------------------------------------------
BENCHMARKS = {
    "inhouse_dter": {
        "title": "inhouse_dter",
        "metrics": ["dter"],
        "groups": INHOUSE_GROUPS,
        "config": "recipe/phimm/config/eval_v2609/long_eval_inhouse_all_seg30.yaml",
        "embedded_baseline": {k: {"dter": v} for k, v in INHOUSE_BASELINE.items()},
        "default_baseline": None,
        "baseline_label": "2609v1,LID",
        "metric_definition": "micro-DTER = sum edits / sum reference tokens",
    },
    "digits_enus": {
        "title": "digits_enus",
        "metrics": ["cer", "wer"],
        "groups": DIGITS_ENUS_GROUPS,
        "config": "recipe/phimm/config/eval_v2609/eval_digits_enus.yaml",
        "embedded_baseline": {},
        "default_baseline": None,
        "metric_definition": "Digit CER and WER from digits_measure.eval_score",
    },
    "openasr_ml": {
        "title": "openasr_ml",
        "metrics": ["wer"],
        "groups": OPENASR_ML_GROUPS,
        "config": "recipe/phimm/config/eval_v2609/eval_openasr_ml_verb.yaml",
        "embedded_baseline": {k: {"wer": v} for k, v in OPENASR_ML_BASELINE.items()},
        "default_baseline": None,
        "baseline_label": "2609v1",
        "metric_definition": "WER / p_err per dataset; arithmetic language and overall averages",
    },
    "mixlang": {
        "title": "mixlang",
        "metrics": ["dter"],
        "groups": MIXLANG_GROUPS,
        "config": "recipe/phimm/config/eval_v2609/long_eval_mixlang_fy26q2_zh_seg.yaml",
        "embedded_baseline": {k: {"dter": v} for k, v in MIXLANG_BASELINE.items()},
        "default_baseline": None,
        "baseline_label": "2609v1",
        "metric_definition": "zh-CN DTER/TER = sum edits / sum reference tokens",
    },
    "digits_tier1": {
        "title": "digits_tier1",
        "metrics": ["cer", "wer"],
        "groups": DIGITS_TIER1_GROUPS,
        "config": "recipe/phimm/config/eval_v2609/eval_digits_tier1.yaml",
        "embedded_baseline": {},
        "default_baseline": None,
        "metric_definition": "Digit CER and WER from digits_measure.eval_score",
    },
}

# Sheet order: summary first, then benchmarks, with mixlang on its own sheet.
SHEET_ORDER = ["inhouse_dter", "digits_enus", "openasr_ml", "mixlang", "digits_tier1"]


def default_report_path(label: str, candidate_model_path: str) -> Path:
    safe_label = re.sub(r"[^A-Za-z0-9._-]+", "_", label).strip("._-") or "candidate"
    model_match = re.search(
        r"/([^/]+)/global_step_\d+(?:/|$)",
        candidate_model_path.rstrip("/") + "/",
    )
    if model_match:
        model_name = model_match.group(1)
    else:
        model_name = re.sub(r"(?:[_-]?step\d+)$", "", safe_label, flags=re.IGNORECASE)
        model_name = model_name.rstrip("._-") or safe_label
    safe_model_name = re.sub(r"[^A-Za-z0-9._-]+", "_", model_name).strip("._-")
    return Path("tmp/eval_2609_reports") / safe_model_name / f"{safe_label}.xlsx"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--label", default="candidate", help="Candidate column (B) label.")
    p.add_argument("--baseline-label", default=None,
                   help="Override the embedded per-benchmark baseline labels.")
    p.add_argument("--reference-model-path", default="", help="Reference HF model path for provenance.")
    p.add_argument("--candidate-model-path", default="", help="Candidate HF model path for provenance.")
    p.add_argument("--artifacts-sidecar-local", default="",
                   help="Local raw-artifact provenance sidecar path.")
    p.add_argument("--artifacts-sidecar-remote", default="",
                   help="Durable remote raw-artifact provenance sidecar path.")
    p.add_argument("--out", default=None, help="Output xlsx path.")
    for name in BENCHMARKS:
        p.add_argument(f"--{name.replace('_', '-')}", dest=name, default=None,
                       help=f"Candidate result source for {name}.")
        p.add_argument(f"--{name.replace('_', '-')}-baseline", dest=f"{name}_baseline",
                       default=None, help=f"Reference result source for {name}.")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet; we add our own in order.

    summary_rows: List[Dict[str, object]] = []
    built = 0
    for name in SHEET_ORDER:
        cand_src = getattr(args, name)
        if not cand_src:
            continue
        spec = BENCHMARKS[name]
        metrics = spec["metrics"]
        candidate = collect(cand_src, metrics)
        if not candidate:
            print(f"[warn] no candidate metrics parsed for {name} from {cand_src}", file=sys.stderr)
            continue
        base_src = getattr(args, f"{name}_baseline") or spec.get("default_baseline")
        baseline = collect(base_src, metrics) if base_src else {}
        if not baseline and spec["embedded_baseline"]:
            baseline = dict(spec["embedded_baseline"])
            if base_src:
                print(f"[warn] {name}: baseline source {base_src} yielded nothing; "
                      f"using embedded baseline", file=sys.stderr)
        if not baseline:
            print(f"[warn] no baseline for {name}; column A will be empty. "
                  f"Pass --{name.replace('_', '-')}-baseline <source>.", file=sys.stderr)

        overall = build_benchmark_sheet(
            wb, spec["title"], metrics, spec["groups"], baseline, candidate,
            args.baseline_label or spec.get("baseline_label", "reference"), args.label,
        )
        built += 1
        for metric in metrics:
            all_base = [baseline[k][metric] for _, ds in spec["groups"] for k, _ in ds
                        if k in baseline and metric in baseline[k]]
            b_overall = _mean(all_base)
            c_overall = overall.get(metric)
            delta = (1 - c_overall / b_overall) if (b_overall not in (None, 0) and c_overall is not None) else None
            summary_rows.append({
                "benchmark": name, "metric": metric.upper(),
                "baseline": b_overall, "candidate": c_overall, "delta": delta,
                "metric_definition": spec["metric_definition"],
                "config": spec["config"],
                "reference_model": args.reference_model_path,
                "candidate_model": args.candidate_model_path,
                "baseline_source": base_src or "embedded baseline",
                "candidate_source": cand_src,
                "artifacts_sidecar_local": args.artifacts_sidecar_local,
                "artifacts_sidecar_remote": args.artifacts_sidecar_remote,
            })

    if built == 0:
        print("[error] no benchmark sources supplied; nothing to build.", file=sys.stderr)
        return 2

    build_summary_sheet(wb, summary_rows, args.baseline_label or "2609v1", args.label)
    wb.move_sheet("summary", -len(wb.sheetnames) + 1)  # summary first

    out_path = Path(args.out) if args.out else default_report_path(
        args.label,
        args.candidate_model_path,
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(out_path)
    print(f"Wrote {out_path} with sheets: {', '.join(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
