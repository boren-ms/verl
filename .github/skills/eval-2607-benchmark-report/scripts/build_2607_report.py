#!/usr/bin/env python3
# ruff: noqa: UP006
"""Collect existing 2607 benchmark eval results into a single XLSX workbook.

Each benchmark sheet compares a candidate model against its reference baseline
and computes relative error-rate reduction as ``1 - candidate / baseline``.
Sources may be Azure/local measures trees, JSON, text logs, or Ray job logs.
Requires ``openpyxl`` from the repository environment (for example, ``uv run``).
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from summary_charts import apply_summary_charts

HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
METRIC_FILL = PatternFill("solid", fgColor="FFDDEBF7")
GROUP_AVG_FILL = PatternFill("solid", fgColor="FFFFF2CC")
OVERALL_FILL = PatternFill("solid", fgColor="FFE2EFDA")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")
PCT_FMT = "0.00%"
DELTA_LABEL = {"dter": "TERR", "ter": "TERR", "wer": "WERR", "cer": "CERR"}

INHOUSE_GROUPS: List[Tuple[str, List[Tuple[str, str]]]] = [
    ("en-US", [
        ("enus_conv_fy21q1", "en-US_Conversation_DTEST_FY21Q1"),
        ("enus_conv_om_fy25q3", "en-US_Conversation_OnlineMeetings_DTEST_FY25Q3"),
        ("enus_dict_office_fy24q3", "en-US_Dictation_Commonset_OfficeOffline_FY24Q3"),
    ]),
    ("nl-NL", [
        ("nlnl_conv_fy23q2", "nl-NL_Conversation_DTEST_FY23Q2"),
        ("nlnl_conv_om_fy23q1", "nl-NL_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("nlnl_dict_fy23q4", "nl-NL_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("da-DK", [
        ("dadk_conv_fy21q3", "da-DK_Conversation_DTEST_FY21Q3"),
        ("dadk_conv_om_fy23q1", "da-DK_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("dadk_dict_fy23q4", "da-DK_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("hu-HU", [
        ("huhu_conv_fy22q4", "hu-HU_Conversation_DTEST_FY22Q4"),
        ("huhu_conv_om_fy24q2", "hu-HU_Conversation_OnlineMeetings_DTEST_FY24Q2"),
        ("huhu_dict_fy25q2", "hu-HU_Dictation_DTEST_L_D_FY25Q2"),
    ]),
    ("nb-NO", [
        ("nbno_conv_fy21q3", "nb-NO_Conversation_DTEST_FY21Q3"),
        ("nbno_conv_om_fy23q1", "nb-NO_Conversation_OnlineMeetings_DTEST_FY23Q1"),
        ("nbno_dict_fy23q4", "nb-NO_Dictation_DTEST_L_D_FY23Q4"),
    ]),
    ("cs-CZ", [
        ("cscz_conv_fy23q2", "cs-CZ_Conversation_DTEST_FY23Q2"),
        ("cscz_conv_om_fy24q2", "cs-CZ_Conversation_OnlineMeetings_DTEST_FY24Q2"),
        ("cscz_dict_fy24q2", "cs-CZ_Dictation_DTEST_L_D_FY24Q2"),
    ]),
]

INHOUSE_BASELINE: dict[str, float] = {
    "enus_conv_fy21q1": 0.1875, "enus_conv_om_fy25q3": 0.1404, "enus_dict_office_fy24q3": 0.0995,
    "nlnl_conv_fy23q2": 0.2275, "nlnl_conv_om_fy23q1": 0.2330, "nlnl_dict_fy23q4": 0.1086,
    "dadk_conv_fy21q3": 0.2242, "dadk_conv_om_fy23q1": 0.2110, "dadk_dict_fy23q4": 0.1689,
    "huhu_conv_fy22q4": 0.1834, "huhu_conv_om_fy24q2": 0.1662, "huhu_dict_fy25q2": 0.1852,
    "nbno_conv_fy21q3": 0.2204, "nbno_conv_om_fy23q1": 0.1612, "nbno_dict_fy23q4": 0.1635,
    "cscz_conv_fy23q2": 0.2421, "cscz_conv_om_fy24q2": 0.1424, "cscz_dict_fy24q2": 0.1335,
}

DIGITS_ENUS_GROUPS = [("en-US", [
    ("enus_digits_random", "enus_digits_random"),
    ("enus_digits_repeat", "enus_digits_repeat"),
])]
OPENASR_ML_GROUPS = [
    ("de", [("de_fleurs", "de_fleurs"), ("de_mcv", "de_mcv")]),
    ("es", [("es_fleurs", "es_fleurs"), ("es_mcv", "es_mcv"), ("es_mls", "es_mls")]),
    ("fr", [("fr_fleurs", "fr_fleurs"), ("fr_mcv", "fr_mcv"), ("fr_mls", "fr_mls")]),
    ("it", [("it_fleurs", "it_fleurs"), ("it_mcv", "it_mcv"), ("it_mls", "it_mls")]),
    ("pt", [("pt_fleurs", "pt_fleurs"), ("pt_mls", "pt_mls")]),
]
OPENASR_ML_BASELINE = {
    "de_fleurs": 0.0260, "de_mcv": 0.0202,
    "es_fleurs": 0.0272, "es_mcv": 0.0234, "es_mls": 0.0290,
    "fr_fleurs": 0.0336, "fr_mcv": 0.0436, "fr_mls": 0.0272,
    "it_fleurs": 0.0152, "it_mcv": 0.0182, "it_mls": 0.0473,
    "pt_fleurs": 0.0297, "pt_mls": 0.0392,
}
MIXLANG_GROUPS = [("zh-cn", [("mixlang_fy26q2", "Dictation_SimuMixedLang_DTEST_FY26Q2")])]
MIXLANG_BASELINE = {"mixlang_fy26q2": 0.2124}
_TIER1_LOCALES = [
    ("de-DE", "dede"), ("en-US", "enus"), ("es-ES", "eses"), ("fr-FR", "frfr"),
    ("it-IT", "itit"), ("ja-JP", "jajp"), ("ko-KR", "kokr"), ("pt-BR", "ptbr"),
    ("zh-CN", "zhcn"),
]
DIGITS_TIER1_GROUPS = [
    (locale, [(f"{prefix}_digits_random", f"{prefix}_digits_random"),
              (f"{prefix}_digits_repeat", f"{prefix}_digits_repeat")])
    for locale, prefix in _TIER1_LOCALES
]


def fetch_ray_logs(node: str, job_id: str) -> str:
    remote = f"bash -l -c 'ray job logs {job_id} 2>&1'"
    proc = subprocess.run(["brix", "ssh", node, "--", remote], capture_output=True, text=True, check=False)
    if proc.returncode:
        print(f"[warn] ray job logs failed for {node}/{job_id}: {proc.stderr[-400:]}", file=sys.stderr)
    return proc.stdout


def _measures_metric(measures: Dict, metric: str) -> Optional[float]:
    if metric in ("dter", "ter"):
        value = measures.get("dter_p_err")
        if isinstance(value, (int, float)):
            return float(value)
        n_err, n_ref = measures.get("dter_n_err"), measures.get("dter_n_ref")
        if isinstance(n_err, (int, float)) and isinstance(n_ref, (int, float)) and n_ref > 0:
            return float(n_err) / float(n_ref)
        value = measures.get("dter") if metric == "dter" else measures.get("ter")
        return float(value) if isinstance(value, (int, float)) else None
    value = measures.get("p_err") if metric == "wer" else measures.get(metric)
    return float(value) if isinstance(value, (int, float)) else None


def read_measures_tree(root: str, metrics: List[str]) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}

    def ingest(slug: str, text: str) -> None:
        try:
            measures = json.loads(text)
        except json.JSONDecodeError:
            print(f"[warn] could not parse measures.json for {slug}", file=sys.stderr)
            return
        values = {metric: _measures_metric(measures, metric) for metric in metrics}
        values = {key: value for key, value in values.items() if value is not None}
        if any(metric in ("dter", "ter") for metric in metrics):
            n_err, n_ref = measures.get("dter_n_err"), measures.get("dter_n_ref")
            if isinstance(n_err, (int, float)) and isinstance(n_ref, (int, float)) and n_ref > 0:
                values["dter_n_err"] = float(n_err)
                values["dter_n_ref"] = float(n_ref)
        if values:
            out[slug] = values

    if root.startswith("az://"):
        listing = subprocess.run(["bbb", "ls", root.rstrip("/") + "/"], capture_output=True, text=True, check=False)
        if listing.returncode:
            print(f"[warn] bbb ls {root} failed:\n{listing.stderr[-400:]}", file=sys.stderr)
            return out
        raw = listing.stdout.replace("\n", "")
        entries = [token.strip() for token in re.split(r"(?=az://)", raw) if token.startswith("az://")]
        for entry in entries:
            if not entry.endswith("/"):
                continue
            slug = entry.rstrip("/").rsplit("/", 1)[-1]
            proc = subprocess.run(
                ["bbb", "cat", f"{root.rstrip('/')}/{slug}/measures.json"],
                capture_output=True,
                text=True,
                check=False,
            )
            if not proc.returncode:
                ingest(slug, proc.stdout)
        return out

    base = Path(root)
    if not base.is_dir():
        print(f"[warn] {root} is not a directory", file=sys.stderr)
        return out
    for child in sorted(base.iterdir()):
        measures_file = child / "measures.json"
        if child.is_dir() and measures_file.exists():
            ingest(child.name, measures_file.read_text())
    return out


_VAL_METRIC_RE = re.compile(
    r"val-(?:aux|core)/(?P<ds>[A-Za-z0-9_\-]+)/(?P<metric>cer|p_err|dter_p_err|dter_n_err|dter_n_ref)/mean@1[:=]\s*(?P<value>[0-9.eE+\-]+)"
)
_DTER_SUMMARY_RE = re.compile(
    r"\[(?P<ds>[A-Za-z0-9_\-]+)\]\s*DTER:\s*[0-9.]+%\s*\[(?P<ne>\d+)\s*/\s*(?P<nr>\d+)\]"
)


def parse_text(text: str, metrics: List[str]) -> Dict[str, Dict[str, float]]:
    raw: Dict[str, Dict[str, float]] = {}
    for match in _VAL_METRIC_RE.finditer(text):
        raw.setdefault(match["ds"], {})[match["metric"]] = float(match["value"])
    for match in _DTER_SUMMARY_RE.finditer(text):
        n_err, n_ref = float(match["ne"]), float(match["nr"])
        if n_ref > 0:
            raw.setdefault(match["ds"], {})["dter"] = n_err / n_ref

    out: Dict[str, Dict[str, float]] = {}
    for dataset, metric_values in raw.items():
        values: Dict[str, float] = {}
        for metric in metrics:
            if metric in ("dter", "ter"):
                if "dter_p_err" in metric_values:
                    values[metric] = metric_values["dter_p_err"]
                elif "dter_n_err" in metric_values and metric_values.get("dter_n_ref", 0) > 0:
                    values[metric] = metric_values["dter_n_err"] / metric_values["dter_n_ref"]
            elif metric == "wer" and "p_err" in metric_values:
                values[metric] = metric_values["p_err"]
            elif metric in metric_values:
                values[metric] = metric_values[metric]
        if values:
            out[dataset] = values
    return out


def load_json_text(text: str, metrics: List[str]) -> Dict[str, Dict[str, float]]:
    raw = json.loads(text)
    out: Dict[str, Dict[str, float]] = {}
    for dataset, value in raw.items():
        if isinstance(value, dict):
            values = {metric: float(value[metric]) for metric in metrics if metric in value}
        else:
            values = {metrics[0]: float(value)}
        if values:
            out[dataset] = values
    return out


def collect(source: Optional[str], metrics: List[str]) -> Dict[str, Dict[str, float]]:
    if not source:
        return {}
    if source.startswith("ray:"):
        parts = source.split(":", 2)
        if len(parts) != 3:
            print(f"[warn] bad ray source '{source}', expected ray:<node>:<job_id>", file=sys.stderr)
            return {}
        return parse_text(fetch_ray_logs(parts[1], parts[2]), metrics)
    if source.startswith("az://") and source.endswith((".json", ".log", ".txt")):
        proc = subprocess.run(["bbb", "cat", source], capture_output=True, text=True, check=False)
        if proc.returncode:
            print(f"[warn] bbb cat {source} failed:\n{proc.stderr[-400:]}", file=sys.stderr)
            return {}
        return load_json_text(proc.stdout, metrics) if source.endswith(".json") else parse_text(proc.stdout, metrics)
    if source.startswith("az://") or Path(source).is_dir():
        return read_measures_tree(source, metrics)
    path = Path(source)
    if path.is_file():
        return (
            load_json_text(path.read_text(), metrics)
            if source.endswith(".json")
            else parse_text(path.read_text(), metrics)
        )
    print(f"[warn] source not found: {source}", file=sys.stderr)
    return {}


def _mean(values: List[float]) -> Optional[float]:
    return sum(values) / len(values) if values else None


def build_benchmark_sheet(
    workbook: Workbook,
    title: str,
    metrics: List[str],
    groups: List[Tuple[str, List[Tuple[str, str]]]],
    baseline: Dict[str, Dict[str, float]],
    candidate: Dict[str, Dict[str, float]],
    baseline_label: str,
    candidate_label: str,
) -> Dict[str, float]:
    worksheet = workbook.create_sheet(title[:31])
    worksheet.cell(2, 1, "Header").font = BOLD
    worksheet.cell(2, 2, baseline_label).font = BOLD
    worksheet.cell(2, 3, candidate_label).font = BOLD
    worksheet.cell(2, 4, "/".join(DELTA_LABEL.get(metric, "delta") for metric in metrics)).font = BOLD
    worksheet.cell(3, 1, "Column").font = BOLD
    for column, label in ((2, "A"), (3, "B"), (4, "A->B")):
        worksheet.cell(3, column, label).font = BOLD
    for row in (2, 3):
        for column in range(1, 5):
            worksheet.cell(row, column).fill = HEADER_FILL

    row = 4
    delta_ranges: List[Tuple[int, int]] = []
    overall_candidate: Dict[str, float] = {}
    for metric in metrics:
        if len(metrics) > 1:
            worksheet.cell(row, 1, metric.upper()).font = BOLD
            for column in range(1, 5):
                worksheet.cell(row, column).fill = METRIC_FILL
            row += 1
        delta_start = row
        all_baseline: List[float] = []
        all_candidate: List[float] = []
        data_ranges: List[Tuple[int, int]] = []
        for group_name, datasets in groups:
            group_baseline: List[float] = []
            group_candidate: List[float] = []
            group_start = row
            for key, display in datasets:
                worksheet.cell(row, 1, display)
                baseline_value = baseline.get(key, {}).get(metric)
                candidate_value = candidate.get(key, {}).get(metric)
                if baseline_value is not None:
                    worksheet.cell(row, 2, baseline_value).number_format = PCT_FMT
                    group_baseline.append(baseline_value)
                    all_baseline.append(baseline_value)
                if candidate_value is not None:
                    worksheet.cell(row, 3, candidate_value).number_format = PCT_FMT
                    group_candidate.append(candidate_value)
                    all_candidate.append(candidate_value)
                if baseline_value not in (None, 0) and candidate_value is not None:
                    worksheet.cell(row, 4, f"=1-C{row}/B{row}").number_format = PCT_FMT
                row += 1
            group_end = row - 1
            data_ranges.append((group_start, group_end))
            worksheet.cell(row, 1, f"{group_name} avg").font = BOLD
            if group_baseline:
                worksheet.cell(row, 2, f"=AVERAGE(B{group_start}:B{group_end})").number_format = PCT_FMT
            if group_candidate:
                worksheet.cell(row, 3, f"=AVERAGE(C{group_start}:C{group_end})").number_format = PCT_FMT
            if group_baseline and group_candidate:
                worksheet.cell(row, 4, f"=1-C{row}/B{row}").number_format = PCT_FMT
            for column in range(1, 5):
                worksheet.cell(row, column).fill = GROUP_AVG_FILL
            row += 1

        baseline_overall = _mean(all_baseline)
        candidate_overall = _mean(all_candidate)
        label = "overall avg" if len(metrics) == 1 else f"{metric.upper()} overall avg"
        worksheet.cell(row, 1, label).font = BOLD
        if baseline_overall is not None:
            ranges = ",".join(f"B{start}:B{end}" for start, end in data_ranges)
            worksheet.cell(row, 2, f"=AVERAGE({ranges})").number_format = PCT_FMT
        if candidate_overall is not None:
            ranges = ",".join(f"C{start}:C{end}" for start, end in data_ranges)
            worksheet.cell(row, 3, f"=AVERAGE({ranges})").number_format = PCT_FMT
            overall_candidate[metric] = candidate_overall
        if baseline_overall not in (None, 0) and candidate_overall is not None:
            worksheet.cell(row, 4, f"=1-C{row}/B{row}").number_format = PCT_FMT
        for column in range(1, 5):
            worksheet.cell(row, column).fill = OVERALL_FILL
        row += 1
        delta_ranges.append((delta_start, row - 1))

    worksheet.column_dimensions["A"].width = 46
    for column in ("B", "C", "D"):
        worksheet.column_dimensions[column].width = 14
    for current_row in range(2, row):
        for column in (2, 3, 4):
            worksheet.cell(current_row, column).alignment = CENTER
    scale = ColorScaleRule(
        start_type="num", start_value=-0.1, start_color="FFF8696B",
        mid_type="num", mid_value=0, mid_color="FFFFFFFF",
        end_type="num", end_value=0.1, end_color="FF63BE7B",
    )
    for start, end in delta_ranges:
        worksheet.conditional_formatting.add(f"D{start}:D{end}", scale)
    return overall_candidate


def build_summary_sheet(
    workbook: Workbook,
    rows: List[Dict[str, object]],
    baseline_label: str,
    candidate_label: str,
) -> None:
    worksheet = workbook.create_sheet("summary")
    headers = [
        "Benchmark", "Metric", f"Baseline ({baseline_label})", f"Candidate ({candidate_label})", "delta",
        "Metric definition", "Config", "Reference model", "Candidate model", "Baseline result", "Candidate result",
        "Artifacts sidecar (local)", "Artifacts sidecar (remote)",
    ]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(1, column, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    for row_number, values in enumerate(rows, start=2):
        worksheet.cell(row_number, 1, values["benchmark"])
        worksheet.cell(row_number, 2, values["metric"])
        for column, key in ((3, "baseline"), (4, "candidate"), (5, "delta")):
            if values.get(key) is not None:
                worksheet.cell(row_number, column, values[key]).number_format = PCT_FMT
        for column, key in enumerate((
            "metric_definition", "config", "reference_model", "candidate_model", "baseline_source", "candidate_source",
            "artifacts_sidecar_local", "artifacts_sidecar_remote",
        ), start=6):
            worksheet.cell(row_number, column, values.get(key, ""))
    widths = [16, 8, 22, 24, 10, 42, 52, 56, 56, 52, 52, 58, 72]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    for cells in worksheet.iter_rows(min_row=2, max_row=1 + len(rows), min_col=2, max_col=5):
        for cell in cells:
            cell.alignment = CENTER
    if rows:
        worksheet.conditional_formatting.add(
            f"E2:E{1 + len(rows)}",
            ColorScaleRule(
                start_type="num", start_value=-0.1, start_color="FFF8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                end_type="num", end_value=0.1, end_color="FF63BE7B",
            ),
        )
    apply_summary_charts(worksheet)


BENCHMARKS = {
    "inhouse_dter": {
        "title": "inhouse_dter", "metrics": ["dter"], "groups": INHOUSE_GROUPS,
        "config": "recipe/phimm/config/eval/long_eval_inhouse_2607v1_all_seg30.yaml",
        "embedded_baseline": {key: {"dter": value} for key, value in INHOUSE_BASELINE.items()},
        "baseline_label": "2607v1,LID",
        "metric_definition": "micro-DTER = sum edits / sum reference tokens",
    },
    "digits_enus": {
        "title": "digits_enus", "metrics": ["cer", "wer"], "groups": DIGITS_ENUS_GROUPS,
        "config": "recipe/phimm/config/eval/eval_digits_enus_2607v1.yaml",
        "embedded_baseline": {}, "baseline_label": "2607v1",
        "metric_definition": "Digit CER and WER from digits_measure.eval_score",
    },
    "openasr_ml": {
        "title": "openasr_ml", "metrics": ["wer"], "groups": OPENASR_ML_GROUPS,
        "config": "recipe/phimm/config/eval/eval_openasr_ml_verb_2607v1.yaml",
        "embedded_baseline": {key: {"wer": value} for key, value in OPENASR_ML_BASELINE.items()},
        "baseline_label": "2607v1",
        "metric_definition": "WER / p_err per dataset; arithmetic language and overall averages",
    },
    "mixlang": {
        "title": "mixlang", "metrics": ["dter"], "groups": MIXLANG_GROUPS,
        "config": "recipe/phimm/config/eval/long_eval_mixlang_fy26q2_zh_seg_2607v1.yaml",
        "embedded_baseline": {key: {"dter": value} for key, value in MIXLANG_BASELINE.items()},
        "baseline_label": "2607v1",
        "metric_definition": "zh-CN DTER/TER = sum edits / sum reference tokens",
    },
    "digits_tier1": {
        "title": "digits_tier1", "metrics": ["cer", "wer"], "groups": DIGITS_TIER1_GROUPS,
        "config": "recipe/phimm/config/eval/eval_digits_tier1_2607v1.yaml",
        "embedded_baseline": {}, "baseline_label": "2607v1",
        "metric_definition": "Digit CER and WER from digits_measure.eval_score",
    },
}
SHEET_ORDER = ["inhouse_dter", "digits_enus", "openasr_ml", "mixlang", "digits_tier1"]


def default_report_path(label: str, candidate_model_path: str) -> Path:
    safe_label = re.sub(r"[^A-Za-z0-9._-]+", "_", label).strip("._-") or "candidate"
    model_match = re.search(r"/([^/]+)/global_step_\d+(?:/|$)", candidate_model_path.rstrip("/") + "/")
    model_name = (
        model_match.group(1)
        if model_match
        else re.sub(r"(?:[_-]?step\d+)$", "", safe_label, flags=re.IGNORECASE)
    )
    safe_model_name = re.sub(r"[^A-Za-z0-9._-]+", "_", model_name.rstrip("._-") or safe_label).strip("._-")
    return Path("tmp/eval_2607_reports") / safe_model_name / f"{safe_label}.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--label", default="candidate", help="Candidate column (B) label.")
    parser.add_argument("--baseline-label", default=None, help="Override embedded per-benchmark baseline labels.")
    parser.add_argument("--reference-model-path", default="", help="Reference HF model path for provenance.")
    parser.add_argument("--candidate-model-path", default="", help="Candidate HF model path for provenance.")
    parser.add_argument("--artifacts-sidecar-local", default="", help="Local raw-artifact provenance sidecar path.")
    parser.add_argument(
        "--artifacts-sidecar-remote",
        default="",
        help="Durable remote raw-artifact provenance sidecar path.",
    )
    parser.add_argument("--out", default=None, help="Output xlsx path.")
    for name in BENCHMARKS:
        option = name.replace("_", "-")
        parser.add_argument(f"--{option}", dest=name, default=None, help=f"Candidate result source for {name}.")
        parser.add_argument(f"--{option}-baseline", dest=f"{name}_baseline", default=None,
                            help=f"Reference result source for {name}.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary_rows: List[Dict[str, object]] = []
    built = 0
    for name in SHEET_ORDER:
        candidate_source = getattr(args, name)
        if not candidate_source:
            continue
        spec = BENCHMARKS[name]
        metrics = spec["metrics"]
        candidate = collect(candidate_source, metrics)
        if not candidate:
            print(f"[warn] no candidate metrics parsed for {name} from {candidate_source}", file=sys.stderr)
            continue
        baseline_source = getattr(args, f"{name}_baseline")
        baseline = collect(baseline_source, metrics) if baseline_source else {}
        if not baseline and spec["embedded_baseline"]:
            baseline = dict(spec["embedded_baseline"])
            if baseline_source:
                print(f"[warn] {name}: baseline source yielded nothing; using embedded baseline", file=sys.stderr)
        if not baseline:
            print(f"[warn] no baseline for {name}; pass --{name.replace('_', '-')}-baseline <source>.", file=sys.stderr)

        overall = build_benchmark_sheet(
            workbook, spec["title"], metrics, spec["groups"], baseline, candidate,
            args.baseline_label or spec["baseline_label"], args.label,
        )
        built += 1
        for metric in metrics:
            all_baseline = [
                baseline[key][metric]
                for _, datasets in spec["groups"]
                for key, _ in datasets
                if key in baseline and metric in baseline[key]
            ]
            baseline_overall = _mean(all_baseline)
            candidate_overall = overall.get(metric)
            delta = (
                1 - candidate_overall / baseline_overall
                if baseline_overall not in (None, 0) and candidate_overall is not None
                else None
            )
            summary_rows.append({
                "benchmark": name, "metric": metric.upper(), "baseline": baseline_overall,
                "candidate": candidate_overall, "delta": delta,
                "metric_definition": spec["metric_definition"], "config": spec["config"],
                "reference_model": args.reference_model_path, "candidate_model": args.candidate_model_path,
                "baseline_source": baseline_source or "embedded baseline", "candidate_source": candidate_source,
                "artifacts_sidecar_local": args.artifacts_sidecar_local,
                "artifacts_sidecar_remote": args.artifacts_sidecar_remote,
            })

    if built == 0:
        print("[error] no benchmark sources supplied; nothing to build.", file=sys.stderr)
        return 2
    build_summary_sheet(workbook, summary_rows, args.baseline_label or "2607v1", args.label)
    workbook.move_sheet("summary", -len(workbook.sheetnames) + 1)
    output_path = Path(args.out) if args.out else default_report_path(args.label, args.candidate_model_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)
    print(f"Wrote {output_path} with sheets: {', '.join(workbook.sheetnames)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
