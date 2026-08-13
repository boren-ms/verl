#!/usr/bin/env python3
"""Merge 2607 checkpoint reports for one model into a single XLSX workbook."""

from __future__ import annotations

import argparse
import copy
import hashlib
import re
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from summary_charts import apply_summary_charts

HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")
PCT_FMT = "0.00%"
SUMMARY_COLUMNS = [
    "Checkpoint",
    "Benchmark",
    "Metric",
    "Baseline",
    "Candidate",
    "Delta",
    "Metric definition",
    "Config",
    "Reference model",
    "Candidate model",
    "Baseline result",
    "Candidate result",
    "Artifacts sidecar (local)",
    "Artifacts sidecar (remote)",
    "Source workbook",
]


def parse_report(value: str) -> Tuple[str, Path]:
    if "=" not in value:
        raise argparse.ArgumentTypeError("expected LABEL=PATH")
    label, raw_path = value.split("=", 1)
    label = label.strip()
    path = Path(raw_path).expanduser()
    if not label:
        raise argparse.ArgumentTypeError("checkpoint label must not be empty")
    if not path.is_file():
        raise argparse.ArgumentTypeError(f"report does not exist: {path}")
    return label, path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--model-label", required=True, help="Shared model/run label.")
    parser.add_argument(
        "--report",
        action="append",
        required=True,
        type=parse_report,
        metavar="CHECKPOINT=PATH",
        help="Checkpoint label and report path; repeat in desired sheet order.",
    )
    parser.add_argument("--out", required=True, help="Merged workbook path.")
    return parser.parse_args()


def safe_sheet_name(checkpoint: str, benchmark: str, used: set[str]) -> str:
    raw = re.sub(r"[\\/*?:\[\]]", "_", f"{checkpoint}_{benchmark}").strip(" '")
    candidate = raw[:31] or "report"
    if candidate not in used:
        used.add(candidate)
        return candidate
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:6]
    candidate = f"{raw[:24]}_{digest}"[:31]
    suffix = 2
    while candidate in used:
        candidate = f"{raw[:27]}_{suffix}"[:31]
        suffix += 1
    used.add(candidate)
    return candidate


def copy_sheet(source, target) -> None:
    for row in source.iter_rows():
        for source_cell in row:
            target_cell = target[source_cell.coordinate]
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.alignment = copy.copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy.copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy.copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy.copy(source_cell.comment)

    for key, dimension in source.column_dimensions.items():
        target_dimension = target.column_dimensions[key]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.bestFit = dimension.bestFit
        target_dimension.outlineLevel = dimension.outlineLevel
    for key, dimension in source.row_dimensions.items():
        target_dimension = target.row_dimensions[key]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel
    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))
    for conditional_format in source.conditional_formatting:
        for rule in conditional_format.rules:
            target.conditional_formatting.add(str(conditional_format.sqref), copy.copy(rule))

    target.freeze_panes = source.freeze_panes
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.auto_filter.ref = source.auto_filter.ref


def summary_rows(workbook, checkpoint: str, source_path: Path) -> List[List[object]]:
    if "summary" not in workbook.sheetnames:
        raise ValueError(f"{source_path}: missing summary sheet")
    worksheet = workbook["summary"]
    headers = [cell.value for cell in worksheet[1]]
    required = SUMMARY_COLUMNS[1:-1]
    indexes: Dict[str, int] = {}
    for header in required:
        if header == "Baseline":
            matches = [index for index, value in enumerate(headers, start=1)
                       if isinstance(value, str) and value.startswith("Baseline (")]
        elif header == "Candidate":
            matches = [index for index, value in enumerate(headers, start=1)
                       if isinstance(value, str) and value.startswith("Candidate (")]
        elif header == "Delta":
            matches = [index for index, value in enumerate(headers, start=1)
                       if isinstance(value, str) and value.lower() == "delta"]
        else:
            matches = [index for index, value in enumerate(headers, start=1) if value == header]
        if len(matches) == 1:
            indexes[header] = matches[0]
    missing = [header for header in required if header not in indexes]
    if missing:
        raise ValueError(f"{source_path}: summary missing columns: {', '.join(missing)}")
    rows: List[List[object]] = []
    for row_number in range(2, worksheet.max_row + 1):
        if worksheet.cell(row_number, indexes["Benchmark"]).value is None:
            continue
        rows.append(
            [checkpoint]
            + [worksheet.cell(row_number, indexes[header]).value for header in required]
            + [str(source_path.resolve())]
        )
    return rows


def report_schema(rows: Iterable[Sequence[object]]) -> Tuple[Tuple[object, ...], ...]:
    return tuple(
        sorted(
            (
                row[1],  # benchmark
                row[2],  # metric
                row[3],  # baseline value
                row[7],  # config
                row[8],  # reference model
                row[10],  # baseline result source
            )
            for row in rows
        )
    )


def build_summary(workbook: Workbook, model_label: str, rows: List[List[object]]) -> None:
    worksheet = workbook.create_sheet("summary")
    worksheet.cell(1, 1, "Model")
    worksheet.cell(1, 2, model_label)
    worksheet.cell(1, 1).font = BOLD
    worksheet.cell(1, 1).fill = HEADER_FILL
    worksheet.cell(1, 2).font = BOLD
    worksheet.cell(1, 2).fill = HEADER_FILL
    for column, header in enumerate(SUMMARY_COLUMNS, start=1):
        cell = worksheet.cell(3, column, header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    for row_number, values in enumerate(rows, start=4):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_number, column, value)
        for column in range(4, 7):
            worksheet.cell(row_number, column).number_format = PCT_FMT
            worksheet.cell(row_number, column).alignment = CENTER
    widths = [18, 16, 10, 14, 14, 12, 42, 52, 56, 56, 52, 52, 58, 72, 52]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = f"A3:{get_column_letter(len(SUMMARY_COLUMNS))}{3 + len(rows)}"
    if rows:
        worksheet.conditional_formatting.add(
            f"F4:F{3 + len(rows)}",
            ColorScaleRule(
                start_type="num", start_value=-0.1, start_color="FFF8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                end_type="num", end_value=0.1, end_color="FF63BE7B",
            ),
        )
    apply_summary_charts(worksheet)


def main() -> int:
    args = parse_args()
    reports: List[Tuple[str, Path]] = args.report
    labels = [label for label, _ in reports]
    if len(labels) != len(set(labels)):
        raise SystemExit("[error] checkpoint labels must be unique")

    output = Workbook()
    output.remove(output.active)
    used_names = {"summary"}
    merged_rows: List[List[object]] = []
    expected_schema: Tuple[Tuple[object, ...], ...] | None = None

    for checkpoint, path in reports:
        source = load_workbook(path, data_only=False)
        rows = summary_rows(source, checkpoint, path)
        schema = report_schema(rows)
        if expected_schema is None:
            expected_schema = schema
        elif schema != expected_schema:
            raise SystemExit(
                f"[error] {path}: benchmark/config/baseline schema differs from the first report"
            )
        merged_rows.extend(rows)
        for sheet_name in source.sheetnames:
            if sheet_name == "summary":
                continue
            target_name = safe_sheet_name(checkpoint, sheet_name, used_names)
            copy_sheet(source[sheet_name], output.create_sheet(target_name))

    build_summary(output, args.model_label, merged_rows)
    output.move_sheet("summary", -len(output.sheetnames) + 1)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    output.calculation.calcMode = "auto"
    output.calculation.fullCalcOnLoad = True
    output.calculation.forceFullCalc = True
    output.save(out_path)
    print(f"Wrote {out_path} with sheets: {', '.join(output.sheetnames)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())