"""Add baseline-versus-candidate checkpoint charts to 2607 summary sheets."""

from __future__ import annotations

from collections import OrderedDict

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter

CHART_DATA_MARKER = "__chart_data__"
CHECKPOINT_COLORS = [
    "4F81BD",
    "C0504D",
    "9BBB59",
    "8064A2",
    "FFC000",
    "5B9BD5",
    "C55A11",
    "8064A2",
    "2F75B5",
    "548235",
]
DATASET_COLORS = ["4F81BD", "C0504D", "9BBB59", "8064A2", "F79646"]


def _show_axis(axis, *, position: str) -> None:
    axis.delete = False
    axis.axPos = position
    axis.tickLblPos = "nextTo"
    axis.majorTickMark = "out"
    axis.spPr = GraphicalProperties()
    axis.spPr.ln.solidFill = "595959"
    axis.spPr.ln.width = 12700


def _show_zero_crossing_category_axis(axis) -> None:
    axis.delete = False
    axis.axPos = "b"
    axis.title = None
    axis.tickLblPos = "low"
    axis.majorTickMark = "none"
    axis.minorTickMark = "none"
    axis.spPr = GraphicalProperties()
    axis.spPr.ln.solidFill = "595959"
    axis.spPr.ln.width = 12700


def _single_summary_data(worksheet) -> list[tuple[str, list[tuple[str, float, float]]]]:
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    baseline_column = next(
        column for column, header in enumerate(headers, start=1) if str(header).startswith("Baseline")
    )
    candidate_column = next(
        column for column, header in enumerate(headers, start=1) if str(header).startswith("Candidate")
    )
    candidate_header = str(worksheet.cell(1, candidate_column).value or "Candidate")
    checkpoint = candidate_header.removeprefix("Candidate (").removesuffix(")")
    return [
        (
            str(worksheet.cell(row, 1).value),
            [(checkpoint, worksheet.cell(row, baseline_column).value, worksheet.cell(row, candidate_column).value)],
        )
        for row in range(2, worksheet.max_row + 1)
        if isinstance(worksheet.cell(row, baseline_column).value, (int, float))
        and isinstance(worksheet.cell(row, candidate_column).value, (int, float))
    ]


def _merged_summary_data(worksheet) -> list[tuple[str, list[tuple[str, float, float]]]]:
    groups: OrderedDict[str, list[tuple[str, float, float]]] = OrderedDict()
    for row in range(4, worksheet.max_row + 1):
        checkpoint = worksheet.cell(row, 1).value
        benchmark = worksheet.cell(row, 2).value
        metric = worksheet.cell(row, 3).value
        baseline = worksheet.cell(row, 4).value
        candidate = worksheet.cell(row, 5).value
        if not all(isinstance(value, str) for value in (checkpoint, benchmark, metric)):
            continue
        if not isinstance(baseline, (int, float)) or not isinstance(candidate, (int, float)):
            continue
        groups.setdefault(benchmark, []).append((checkpoint, baseline, candidate))
    return list(groups.items())


def _legacy_summary_data(worksheet) -> list[tuple[str, list[tuple[str, float, float]]]]:
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    baseline_column = next(
        column for column, header in enumerate(headers, start=1) if str(header).startswith("Baseline")
    )
    delta_columns = [
        column for column, header in enumerate(headers, start=1) if str(header).lower() == "delta"
    ]
    candidate_columns = [
        column - 1 for column in delta_columns if column > baseline_column + 1
    ]
    output = []
    for row in range(2, worksheet.max_row + 1):
        baseline = worksheet.cell(row, baseline_column).value
        if not isinstance(baseline, (int, float)):
            continue
        points = [
            (str(worksheet.cell(1, column).value), baseline, worksheet.cell(row, column).value)
            for column in candidate_columns
            if isinstance(worksheet.cell(row, column).value, (int, float))
        ]
        if points:
            output.append((str(worksheet.cell(row, 1).value), points))
    return output


def _chart_data(worksheet) -> tuple[int, list[tuple[str, list[tuple[str, float, float]]]]]:
    marker_column = next(
        (column for column in range(1, worksheet.max_column + 1) if worksheet.cell(1, column).value == CHART_DATA_MARKER),
        None,
    )
    if marker_column is not None:
        return marker_column, []
    if worksheet.cell(3, 1).value == "Checkpoint":
        return worksheet.max_column + 10, _merged_summary_data(worksheet)
    if worksheet.cell(1, 1).value == "Benchmark":
        delta_count = sum(
            str(worksheet.cell(1, column).value).lower() == "delta"
            for column in range(1, worksheet.max_column + 1)
        )
        parser = _single_summary_data if delta_count == 1 else _legacy_summary_data
        return worksheet.max_column + 10, parser(worksheet)
    return worksheet.max_column + 10, []


def apply_summary_charts(worksheet) -> int:
    marker_column, data = _chart_data(worksheet)
    if not data:
        marker_column = next(
            (column for column in range(1, worksheet.max_column + 1) if worksheet.cell(1, column).value == CHART_DATA_MARKER),
            marker_column,
        )
        for row in range(1, worksheet.max_row + 1):
            for column in range(marker_column, worksheet.max_column + 1):
                worksheet.cell(row, column).value = None
        if worksheet.cell(3, 1).value == "Checkpoint":
            data = _merged_summary_data(worksheet)
        elif worksheet.cell(1, 1).value == "Benchmark":
            delta_count = sum(
                str(worksheet.cell(1, column).value).lower() == "delta"
                for column in range(1, worksheet.max_column + 1)
            )
            parser = _single_summary_data if delta_count == 1 else _legacy_summary_data
            data = parser(worksheet)
    if not data:
        return 0

    worksheet._charts.clear()
    worksheet.cell(1, marker_column, CHART_DATA_MARKER)
    if worksheet.cell(3, 1).value == "Checkpoint":
        visible_end_row = max(
            row
            for row in range(4, worksheet.max_row + 1)
            if isinstance(worksheet.cell(row, 1).value, str)
            and isinstance(worksheet.cell(row, 2).value, str)
            and isinstance(worksheet.cell(row, 3).value, str)
        )
    else:
        visible_end_row = max(
            row
            for row in range(2, worksheet.max_row + 1)
            if isinstance(worksheet.cell(row, 1).value, str)
            and any(
                isinstance(worksheet.cell(row, column).value, (int, float))
                for column in range(2, marker_column)
            )
        )
    chart_start_row = visible_end_row + 3
    helper_row = 2
    checkpoints = list(OrderedDict.fromkeys(checkpoint for _, points in data for checkpoint, _, _ in points))
    worksheet.cell(helper_row, marker_column, "Dataset")
    for offset, checkpoint in enumerate(checkpoints, start=1):
        worksheet.cell(helper_row, marker_column + offset, checkpoint)
    delta_matrix: list[list[float | None]] = []
    for row_offset, (dataset, points) in enumerate(data, start=1):
        worksheet.cell(helper_row + row_offset, marker_column, dataset)
        point_map = {checkpoint: (baseline, candidate) for checkpoint, baseline, candidate in points}
        deltas = []
        for column_offset, checkpoint in enumerate(checkpoints, start=1):
            pair = point_map.get(checkpoint)
            delta = None if pair is None or pair[0] == 0 else 1 - pair[1] / pair[0]
            worksheet.cell(helper_row + row_offset, marker_column + column_offset, delta).number_format = "0.00%"
            deltas.append(delta)
        delta_matrix.append(deltas)
    for column in range(marker_column, marker_column + len(checkpoints) + 1):
        worksheet.column_dimensions[get_column_letter(column)].hidden = True

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.overlap = -40
    chart.gapWidth = 260
    chart.title = None
    chart.height = 10
    chart.width = 20
    chart.y_axis.title = None
    chart.y_axis.numFmt = "0.00%"
    chart.y_axis.majorGridlines = None
    deltas = [delta for row in delta_matrix for delta in row if delta is not None]
    minimum = min(deltas)
    maximum = max(deltas)
    span = maximum - minimum
    padding = span * 0.15 if span > 0 else max(abs(minimum), 0.01) * 0.15
    chart.y_axis.scaling.min = min(0, minimum - padding)
    chart.y_axis.scaling.max = max(0, maximum + padding)
    chart.y_axis.crossesAt = 0
    chart.x_axis.title = None
    _show_zero_crossing_category_axis(chart.x_axis)
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties()
    chart.x_axis.majorGridlines.spPr.ln.solidFill = "D9D9D9"
    chart.x_axis.majorGridlines.spPr.ln.width = 6350
    _show_axis(chart.y_axis, position="l")
    chart.legend = None
    chart.visible_cells_only = False
    chart.add_data(
        Reference(
            worksheet,
            min_col=marker_column + 1,
            max_col=marker_column + len(checkpoints),
            min_row=helper_row,
            max_row=helper_row + len(data),
        ),
        titles_from_data=True,
    )
    categories = Reference(
        worksheet,
        min_col=marker_column,
        min_row=helper_row + 1,
        max_row=helper_row + len(data),
    )
    chart.set_categories(categories)
    for series_index, series in enumerate(chart.series):
        series.invertIfNegative = False
        series_color = CHECKPOINT_COLORS[series_index % len(CHECKPOINT_COLORS)]
        series.graphicalProperties.solidFill = series_color
        series.graphicalProperties.line.solidFill = series_color
        series.dPt = []
        for point_index, row in enumerate(delta_matrix):
            delta = row[series_index]
            if delta is None:
                continue
            color = (
                DATASET_COLORS[point_index % len(DATASET_COLORS)]
                if len(checkpoints) == 1
                else series_color
            )
            point = DataPoint(idx=point_index)
            point.invertIfNegative = False
            point.graphicalProperties.solidFill = color
            point.graphicalProperties.line.solidFill = color
            point.graphicalProperties.line.width = 12700
            series.dPt.append(point)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.numFmt = "0.00%"
    chart.dLbls.position = "outEnd"
    chart.legend = Legend(legendPos="t")

    worksheet.add_chart(chart, f"A{chart_start_row}")
    return 1