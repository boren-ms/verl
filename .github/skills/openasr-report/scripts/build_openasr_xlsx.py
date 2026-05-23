#!/usr/bin/env python3
"""Build an OpenASR + OpenASR_ML xlsx report matching ~/code/MoE/results/template.xlsx.

Layout (single sheet `openasr`):
  Row 2  : Header | Baseline | <model-label-1> | <model-label-2> | ... | WERR(s)
  Row 3  : Column | A        | B               | C               | ... | A->B, A->C, ...
  Row 4-11 : OpenASR datasets (ami..voxpopuli)
  Row 12 : avg (formula =AVERAGE(<col>4:<col>11))
  Row 13 : Column row repeated for the ML section
  Row 14-31 : ML datasets with per-language avg rows mixed in
  Row 32 : ml avg (formula =AVERAGE(<col>14:<col>31))
  Column(s) after the last model column: WERR = 1 - <model>/Baseline for each model column.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Row-highlight colors.
HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")  # light blue
LANG_AVG_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # light yellow
OVERALL_AVG_FILL = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")  # light green

# ---------------------------------------------------------------------------
# Fixed schemas
# ---------------------------------------------------------------------------

OPENASR_DATASETS: List[str] = [
    "ami",
    "earnings22",
    "gigaspeech",
    "ls_clean",
    "ls_other",
    "spgispeech",
    "tedlium",
    "voxpopuli",
]

# Ordered: list of (language_code, [datasets in that language]).
OPENASR_ML_GROUPS: List[Tuple[str, List[str]]] = [
    ("de", ["de_fleurs", "de_mcv"]),
    ("es", ["es_fleurs", "es_mcv", "es_mls"]),
    ("fr", ["fr_fleurs", "fr_mcv", "fr_mls"]),
    ("it", ["it_fleurs", "it_mcv", "it_mls"]),
    ("pt", ["pt_fleurs", "pt_mls"]),
]

BASELINE_LABEL = "Qwen3.5-audio"
# Baseline = fast-llm-2605-qwen3-5-9b-s2-st-example-r2 @ step 90000
# Captured 2026-05-22 from Ray jobs raysubmit_AkJUtDv4nPuJ3Gta (verl-n1-i0, openasr)
# and raysubmit_F9u4EHAY5CQsvBjY (verl-n1-i4, openasr_ml).
BASELINE_METRICS: Dict[str, float] = {
    # OpenASR
    "ami": 0.0814,
    "earnings22": 0.0788,
    "gigaspeech": 0.0876,
    "ls_clean": 0.0154,
    "ls_other": 0.0287,
    "spgispeech": 0.0269,
    "tedlium": 0.0240,
    "voxpopuli": 0.0531,
    # OpenASR_ML
    "de_fleurs": 0.0238,
    "de_mcv": 0.0193,
    "es_fleurs": 0.0289,
    "es_mcv": 0.0224,
    "es_mls": 0.0280,
    "fr_fleurs": 0.0315,
    "fr_mcv": 0.0430,
    "fr_mls": 0.0276,
    "it_fleurs": 0.0123,
    "it_mcv": 0.0183,
    "it_mls": 0.0512,
    "pt_fleurs": 0.0308,
    "pt_mls": 0.0342,
}

WER_LINE_RE = re.compile(
    r"val-aux/(?P<dataset>[a-zA-Z0-9_]+)/p_err/mean@1[:=]\s*(?P<value>[0-9.eE+-]+)"
)


def parse_wer_lines(text: str) -> Dict[str, float]:
    """Extract the latest ``val-aux/<ds>/p_err/mean@1`` value per dataset."""
    out: Dict[str, float] = {}
    for m in WER_LINE_RE.finditer(text):
        try:
            out[m.group("dataset")] = float(m.group("value"))
        except ValueError:
            continue
    return out


def fetch_ray_logs(node: str, job_id: str) -> str:
    # brix ssh joins trailing args by space, so wrap the remote command in a single
    # pre-quoted string to preserve `bash -l -c "..."` argument boundaries.
    remote = f"bash -l -c 'ray job logs {job_id} 2>&1'"
    cmd = ["brix", "ssh", node, "--", remote]
    proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        print(
            f"[warn] ray job logs failed for {node}/{job_id} (rc={proc.returncode}):\n"
            f"{proc.stderr[-500:]}",
            file=sys.stderr,
        )
    return proc.stdout


def load_metrics(args: argparse.Namespace) -> Dict[str, float]:
    metrics: Dict[str, float] = {}
    if args.metrics:
        data = json.loads(Path(args.metrics).read_text())
        metrics.update({k: float(v) for k, v in data.items()})
    if args.from_text:
        for text_path in args.from_text:
            metrics.update(parse_wer_lines(Path(text_path).read_text()))
    if args.from_ray:
        for node, job_id in args.from_ray:
            metrics.update(parse_wer_lines(fetch_ray_logs(node, job_id)))
    return metrics


OPENASR_START = 4  # row of first OpenASR dataset
OPENASR_END = OPENASR_START + len(OPENASR_DATASETS) - 1  # 11
OPENASR_AVG = OPENASR_END + 1  # 12

ML_HEADER = OPENASR_AVG + 1  # 13
ML_START = ML_HEADER + 1  # 14


def _ml_layout() -> Tuple[List[Tuple[str, str]], int]:
    """Return ((row_kind, label), ...) starting at row ML_START and the ml-avg row.

    row_kind is 'data' for dataset rows or 'lang_avg' for per-language avg rows.
    """
    layout: List[Tuple[str, str]] = []
    for lang, datasets in OPENASR_ML_GROUPS:
        for ds in datasets:
            layout.append(("data", ds))
        layout.append(("lang_avg", f"{lang} avg"))
    return layout, ML_START + len(layout)  # ml avg row index


def build_workbook(columns: List[Tuple[str, Dict[str, float]]], out_path: Path) -> None:
    """``columns`` is a list of (label, metrics_dict). The first is the baseline (column B)."""

    wb = Workbook()
    ws = wb.active
    ws.title = "openasr"

    n_models = len(columns)  # baseline + new model(s)
    model_cols = list(range(2, 2 + n_models))  # spreadsheet columns 2..(1+n_models)
    werr_cols = list(range(2 + n_models, 2 + 2 * n_models - 1 + 1))
    # WERR is computed relative to baseline (column B), so we make one WERR column per
    # non-baseline model column.
    werr_cols = list(range(2 + n_models, 2 + n_models + (n_models - 1)))

    bold = Font(bold=True)

    # Header row 2
    ws.cell(row=2, column=1, value="Header").font = bold
    ws.cell(row=2, column=2, value="Baseline").font = bold
    for i, (label, _) in enumerate(columns[1:], start=0):
        ws.cell(row=2, column=3 + i, value=label).font = bold
    for j, mc in enumerate(model_cols[1:]):
        ws.cell(row=2, column=werr_cols[j], value="WERR").font = bold

    # Column row 3
    ws.cell(row=3, column=1, value="Column").font = bold
    for i, mc in enumerate(model_cols):
        ws.cell(row=3, column=mc, value=chr(ord("A") + i)).font = bold
    for j, wc in enumerate(werr_cols):
        target_letter = chr(ord("A") + 1 + j)  # B, C, D...
        ws.cell(row=3, column=wc, value=f"A->{target_letter}").font = bold

    # OpenASR section (rows 4..11)
    for r, ds in enumerate(OPENASR_DATASETS, start=OPENASR_START):
        ws.cell(row=r, column=1, value=ds)
        for i, (_, metrics) in enumerate(columns):
            v = metrics.get(ds)
            if v is not None:
                ws.cell(row=r, column=model_cols[i], value=float(v))
        for j, wc in enumerate(werr_cols):
            base_letter = get_column_letter(model_cols[0])
            tgt_letter = get_column_letter(model_cols[1 + j])
            ws.cell(row=r, column=wc, value=f"=1-{tgt_letter}{r}/{base_letter}{r}")

    # OpenASR avg row 12 (computed numeric values so cells are always filled)
    r = OPENASR_AVG
    ws.cell(row=r, column=1, value="avg").font = bold
    avg_per_col: List[Optional[float]] = []
    for i, (_, metrics) in enumerate(columns):
        vals = [metrics[ds] for ds in OPENASR_DATASETS if ds in metrics]
        avg = sum(vals) / len(vals) if vals else None
        avg_per_col.append(avg)
        if avg is not None:
            ws.cell(row=r, column=model_cols[i], value=float(avg)).font = bold
    for j, wc in enumerate(werr_cols):
        base_v = avg_per_col[0]
        tgt_v = avg_per_col[1 + j]
        if base_v and tgt_v is not None:
            ws.cell(row=r, column=wc, value=float(1 - tgt_v / base_v)).font = bold

    # ML header row 13
    ws.cell(row=ML_HEADER, column=1, value="Column").font = bold
    for i, mc in enumerate(model_cols):
        ws.cell(row=ML_HEADER, column=mc, value=chr(ord("A") + i)).font = bold
    for j, wc in enumerate(werr_cols):
        target_letter = chr(ord("A") + 1 + j)
        ws.cell(row=ML_HEADER, column=wc, value=f"A->{target_letter}").font = bold

    # ML datasets / per-language avg rows
    layout, ml_avg_row = _ml_layout()
    current_lang_start: Optional[int] = None
    for offset, (kind, label) in enumerate(layout):
        r = ML_START + offset
        ws.cell(row=r, column=1, value=label)
        if kind == "data":
            if current_lang_start is None:
                current_lang_start = r
            for i, (_, metrics) in enumerate(columns):
                v = metrics.get(label)
                if v is not None:
                    ws.cell(row=r, column=model_cols[i], value=float(v))
            for j, wc in enumerate(werr_cols):
                base_letter = get_column_letter(model_cols[0])
                tgt_letter = get_column_letter(model_cols[1 + j])
                ws.cell(row=r, column=wc, value=f"=1-{tgt_letter}{r}/{base_letter}{r}")
        else:  # lang_avg
            ws.cell(row=r, column=1).font = bold
            assert current_lang_start is not None
            # Datasets contributing to this language group.
            group_datasets = [
                lbl for k, lbl in layout[offset - (r - current_lang_start):offset] if k == "data"
            ]
            lang_avg_per_col: List[Optional[float]] = []
            for i, (_, metrics) in enumerate(columns):
                vals = [metrics[ds] for ds in group_datasets if ds in metrics]
                avg = sum(vals) / len(vals) if vals else None
                lang_avg_per_col.append(avg)
                if avg is not None:
                    ws.cell(row=r, column=model_cols[i], value=float(avg)).font = bold
            for j, wc in enumerate(werr_cols):
                base_v = lang_avg_per_col[0]
                tgt_v = lang_avg_per_col[1 + j]
                if base_v and tgt_v is not None:
                    ws.cell(row=r, column=wc, value=float(1 - tgt_v / base_v)).font = bold
            current_lang_start = None

    # ml avg row: mean across all ML datasets (excludes per-language avg rows).
    r = ml_avg_row
    ws.cell(row=r, column=1, value="ml avg").font = bold
    all_ml = [lbl for kind, lbl in layout if kind == "data"]
    ml_avg_per_col: List[Optional[float]] = []
    for i, (_, metrics) in enumerate(columns):
        vals = [metrics[ds] for ds in all_ml if ds in metrics]
        avg = sum(vals) / len(vals) if vals else None
        ml_avg_per_col.append(avg)
        if avg is not None:
            ws.cell(row=r, column=model_cols[i], value=float(avg)).font = bold
    for j, wc in enumerate(werr_cols):
        base_v = ml_avg_per_col[0]
        tgt_v = ml_avg_per_col[1 + j]
        if base_v and tgt_v is not None:
            ws.cell(row=r, column=wc, value=float(1 - tgt_v / base_v)).font = bold

    # Cell formatting: percentages for numeric data, column widths.
    pct_fmt = "0.00%"
    for r in list(range(OPENASR_START, OPENASR_AVG + 1)) + list(range(ML_START, ml_avg_row + 1)):
        for mc in model_cols + werr_cols:
            ws.cell(row=r, column=mc).number_format = pct_fmt

    # Row highlights.
    all_cols = [1] + model_cols + werr_cols
    header_rows = [2, 3, ML_HEADER]
    lang_avg_rows = [
        ML_START + offset for offset, (kind, _) in enumerate(layout) if kind == "lang_avg"
    ]
    overall_avg_rows = [OPENASR_AVG, ml_avg_row]
    for r in header_rows:
        for c in all_cols:
            ws.cell(row=r, column=c).fill = HEADER_FILL
    for r in lang_avg_rows:
        for c in all_cols:
            ws.cell(row=r, column=c).fill = LANG_AVG_FILL
    for r in overall_avg_rows:
        for c in all_cols:
            ws.cell(row=r, column=c).fill = OVERALL_AVG_FILL

    ws.column_dimensions["A"].width = 18
    for mc in model_cols + werr_cols:
        ws.column_dimensions[get_column_letter(mc)].width = 18

    # Center-align every cell except column A across the whole used range.
    center = Alignment(horizontal="center", vertical="center")
    for r in range(2, ml_avg_row + 1):
        for c in model_cols + werr_cols:
            ws.cell(row=r, column=c).alignment = center

    # Conditional 3-color scale on WERR columns: red for negative (regression),
    # white at 0, green for positive (WER reduction). Midpoint fixed at 0.
    for wc in werr_cols:
        col_letter = get_column_letter(wc)
        rng = f"{col_letter}{OPENASR_START}:{col_letter}{ml_avg_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=-1, start_color="FFF8696B",  # red
            mid_type="num", mid_value=0, mid_color="FFFFFFFF",          # white
            end_type="num", end_value=1, end_color="FF63BE7B",          # green
        )
        ws.conditional_formatting.add(rng, rule)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("label", help="New model label (column header).")
    p.add_argument("--from-ray", nargs=2, metavar=("NODE", "JOB_ID"), action="append", default=[])
    p.add_argument("--from-text", action="append", default=[])
    p.add_argument("--metrics", help="JSON file: {dataset: wer_fraction}.")
    p.add_argument("--baseline", help="Override baseline metrics JSON.")
    p.add_argument("--baseline-label", default=BASELINE_LABEL)
    p.add_argument(
        "--extend-xlsx",
        help=(
            "Existing xlsx to extend by appending the new model as the next column. "
            "The sheet layout must match the template."
        ),
    )
    p.add_argument("--out", help="Output xlsx path (default: tmp/openasr_report/<label>.xlsx).")
    return p.parse_args()


def read_existing_xlsx(path: Path) -> List[Tuple[str, Dict[str, float]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["openasr"] if "openasr" in wb.sheetnames else wb.active

    # Discover model columns: row 3 cells with single letter values starting at B.
    model_cols: List[int] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if isinstance(v, str) and len(v) == 1 and v.isalpha():
            model_cols.append(c)
        else:
            break

    labels: List[str] = []
    for mc in model_cols:
        v = ws.cell(row=2, column=mc).value
        labels.append(str(v) if v is not None else f"col{mc}")
    # First column is "Baseline" header → relabel with --baseline-label or keep as-is.

    metrics_per_col: List[Dict[str, float]] = [dict() for _ in model_cols]
    layout, ml_avg_row = _ml_layout()
    rows: List[Tuple[int, str]] = []
    for r, ds in enumerate(OPENASR_DATASETS, start=OPENASR_START):
        rows.append((r, ds))
    for offset, (kind, label) in enumerate(layout):
        if kind == "data":
            rows.append((ML_START + offset, label))
    for r, ds in rows:
        for i, mc in enumerate(model_cols):
            v = ws.cell(row=r, column=mc).value
            if isinstance(v, (int, float)):
                metrics_per_col[i][ds] = float(v)

    return [(label, m) for label, m in zip(labels, metrics_per_col)]


def main() -> int:
    args = parse_args()

    new_metrics = load_metrics(args)
    if not new_metrics:
        print("[error] no metrics collected; pass --from-ray/--from-text/--metrics.", file=sys.stderr)
        return 2

    if args.extend_xlsx:
        columns = read_existing_xlsx(Path(args.extend_xlsx))
        if not columns:
            print(f"[warn] could not parse {args.extend_xlsx}; using baseline only.", file=sys.stderr)
    else:
        columns = []

    if not columns:
        baseline_metrics = BASELINE_METRICS
        if args.baseline:
            baseline_metrics = {k: float(v) for k, v in json.loads(Path(args.baseline).read_text()).items()}
        columns = [(args.baseline_label, baseline_metrics)]

    columns.append((args.label, new_metrics))

    out_path = Path(
        args.out
        or f"tmp/openasr_report/{args.label.replace('/', '_').replace('@', '_')}.xlsx"
    )
    build_workbook(columns, out_path)
    print(out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
