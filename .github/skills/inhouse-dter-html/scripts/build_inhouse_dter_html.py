#!/usr/bin/env python3
"""Build a self-contained Chart.js HTML report from an in-house DTER xlsx."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


HEADER_ROW = 2
DATA_START = 4


TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>In-house DTER Report</title>
  <style>
:root { --blue:#1a6fb5; --red:#dc2626; --green:#16a34a; --orange:#d97706; --gray:#475569; --bg:#f8fafc; --border:#e2e8f0; }
body { font-family: -apple-system, "Segoe UI", "Helvetica Neue", Arial, sans-serif;
       background: var(--bg); padding: 28px; color:#0f172a; }
.container { max-width: 1280px; margin: 0 auto; }
header { padding:18px 24px; background:white; border-left:5px solid var(--blue);
         border-radius:6px; margin-bottom:24px; box-shadow:0 1px 3px rgba(0,0,0,.06); }
.tldr { background:#f0f9ff; border-left:4px solid var(--blue); padding:16px 22px;
        margin-top:14px; line-height:1.85; border-radius:4px; font-size:14px; }
section { background:white; padding:18px 22px; border-radius:6px; margin-bottom:22px;
          box-shadow:0 1px 3px rgba(0,0,0,.06); }
section h2 { color:var(--blue); border-bottom:2px solid var(--blue); padding-bottom:6px;
             margin:0 0 12px; font-size:17px; }
.chart-box { position:relative; height:360px; margin:10px 0; }
.chart-box.tall { height:460px; }
.chart-grid { display:grid; grid-template-columns:1fr 1fr; gap:18px; }
.chart-grid .chart-box { height:240px; }
table { width:100%; border-collapse:collapse; font-size:12px; margin-top:8px; }
th, td { padding:5px 8px; border-bottom:1px solid var(--border); text-align:right; white-space:nowrap; }
th:first-child, td:first-child { text-align:left; }
th { background:#dbeafe; color:#1e3a8a; }
tr.lang-avg { background:#fef9c3; font-weight:600; }
tr.overall  { background:#dcfce7; font-weight:700; }
.neg { color:var(--red); } .pos { color:var(--green); }
.hl-red { color:var(--red); font-weight:700; }
.hl-green { color:var(--green); font-weight:700; }
.hl-orange { color:var(--orange); font-weight:700; }
.legend-tag { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:700; }
.tag-improve { background:#dcfce7; color:#15803d; }
.tag-degrade { background:#fee2e2; color:#b91c1c; }
.meta { color:#334155; font-size:13px; line-height:1.7; }
ul.takeaways { margin:0; padding-left:20px; line-height:1.8; }
code { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; font-size: 12px; }
@media (max-width: 960px) {
  .chart-grid { grid-template-columns:1fr; }
  body { padding: 16px; }
}
  </style>
</head>
<body>
  <div class="container">
    <header>
      <h1 id="report-title"></h1>
      <div class="meta" id="report-meta"></div>
      <div class="tldr" id="report-tldr"></div>
    </header>

    <section>
      <h2>Section 1 - Per-locale WERR averages</h2>
      <div class="chart-box tall"><canvas id="locale-dter-chart"></canvas></div>
    </section>

    <section>
      <h2>Section 2 - Dataset WERR Delta Vs Baseline (Pre-Locale Charts)</h2>
      <div class="chart-box tall"><canvas id="dataset-werr-chart"></canvas></div>
    </section>

    <section>
      <h2>Section 3 - Overall improve / degrade ranking table</h2>
      <table id="ranking-table"></table>
    </section>

    <section>
      <h2>Section 4 - Full per-dataset tables</h2>
      <div id="dter-table-wrap"></div>
      <div id="werr-table-wrap"></div>
    </section>

    <section>
      <h2>Section 5 - Takeaways</h2>
      <ul class="takeaways" id="takeaways"></ul>
    </section>
  </div>

  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
  <script>
const DATA = __DATA__;
const LANG_COLORS = {
  'en-US':'#16a34a', 'nl-NL':'#dc2626', 'da-DK':'#d97706',
  'hu-HU':'#7c3aed', 'nb-NO':'#0ea5e9', 'cs-CZ':'#be185d',
  'fr-FR':'#0891b2', 'de-DE':'#a16207', 'es-ES':'#9333ea',
  'it-IT':'#65a30d', 'ja-JP':'#db2777', 'ko-KR':'#0d9488'
};
const fmtPct    = v => v == null ? '' : (v*100).toFixed(2) + '%';
const fmtSigned = v => v == null ? '' : (v >= 0 ? '+' : '') + (v*100).toFixed(2) + '%';

function escapeHtml(value) {
  return String(value)
    .replaceAll('&', '&amp;')
    .replaceAll('<', '&lt;')
    .replaceAll('>', '&gt;')
    .replaceAll('"', '&quot;')
    .replaceAll("'", '&#39;');
}

function colorForLang(lang) {
  return LANG_COLORS[lang] || '#475569';
}

function colorCycle(i) {
  const colors = ['#1a6fb5', '#dc2626', '#16a34a', '#d97706', '#7c3aed', '#0891b2', '#9333ea'];
  return colors[i % colors.length];
}

function shortDatasetName(name, maxLen = 26) {
  if (!name || typeof name !== 'string') return '';
  const parts = name.split('_');
  let shortName = parts.length > 1 ? parts.slice(1).join('_') : name;
  shortName = shortName
    .replaceAll('Conversation', 'Conv')
    .replaceAll('OnlineMeetings', 'OnlineMtg')
    .replaceAll('Dictation', 'Dict')
    .replaceAll('Commonset', 'Common')
    .replaceAll('OfficeOffline', 'OfficeOff');
  shortName = shortName
    .replaceAll('_DTEST_', '_')
    .replaceAll('_FY', '_')
    .replaceAll('_Q', 'Q')
    .replaceAll('_L_D_', '_')
    .replaceAll('_', '-');
  if (shortName.length > maxLen) {
    shortName = shortName.slice(0, Math.max(4, maxLen - 3)) + '...';
  }
  return shortName;
}

function displayDatasetName(row) {
  if (!row || !row.name) return '';
  if (row.kind !== 'data') return row.name;
  return shortDatasetName(row.name, 30);
}

function seriesRange(values, opts = {}) {
  const { clamp01 = false, includeZero = false } = opts;
  let lo = Infinity;
  let hi = -Infinity;
  values.forEach((v) => {
    if (v == null) return;
    if (v < lo) lo = v;
    if (v > hi) hi = v;
  });
  if (!Number.isFinite(lo) || !Number.isFinite(hi)) {
    return clamp01 ? { min: 0, max: 1 } : { min: -0.05, max: 0.05 };
  }
  if (includeZero) {
    lo = Math.min(lo, 0);
    hi = Math.max(hi, 0);
  }

  const span = hi - lo;
  const pad = Math.max(0.002, span * 0.15, Math.abs(hi) * 0.02, Math.abs(lo) * 0.02);
  let min = lo - pad;
  let max = hi + pad;

  if (max <= min) {
    min -= 0.01;
    max += 0.01;
  }

  if (clamp01) {
    min = Math.max(0, min);
    max = Math.min(1, max);
  }

  return {
    min,
    max,
  };
}

const valueLabelPlugin = {
  id: 'valueLabelPlugin',
  afterDatasetsDraw(chart, args, options) {
    const { ctx, chartArea } = chart;
    ctx.save();
    ctx.font = '11px -apple-system, Segoe UI, Helvetica Neue, Arial, sans-serif';
    ctx.fillStyle = '#334155';
    ctx.textAlign = 'center';
    chart.data.datasets.forEach((dataset, dsIdx) => {
      const meta = chart.getDatasetMeta(dsIdx);
      if (meta.hidden) return;
      meta.data.forEach((bar, i) => {
        const val = dataset.data[i];
        if (val == null) return;
        const isPositive = val >= 0;
        ctx.textBaseline = isPositive ? 'bottom' : 'top';
        let y = isPositive ? (bar.y - 4) : (bar.y + 4);
        y = Math.max(chartArea.top + 2, Math.min(chartArea.bottom - 2, y));
        ctx.fillText(fmtSigned(val), bar.x, y);
      });
    });
    ctx.restore();
  }
};

const twoLayerTickPlugin = {
  id: 'twoLayerTickPlugin',
  afterDraw(chart, args, options) {
    const datasets = options.datasets || [];
    if (!datasets.length) return;
    const { ctx, chartArea, scales } = chart;
    const x = scales.x;
    if (!x) return;
    ctx.save();
    ctx.textAlign = 'center';

    // Top layer: dataset names colored by locale.
    ctx.font = '11px -apple-system, Segoe UI, Helvetica Neue, Arial, sans-serif';
    ctx.textBaseline = 'top';
    datasets.forEach((row, idx) => {
      ctx.fillStyle = colorForLang(row.lang);
      const xp = x.getPixelForValue(idx);
      const yBase = chartArea.bottom + 10 + ((idx % 2) * 8);
      ctx.save();
      ctx.translate(xp, yBase);
      ctx.rotate(-0.8);
      ctx.textAlign = 'right';
      ctx.fillText(shortDatasetName(row.name, 16), 0, 0);
      ctx.restore();
    });

    // Bottom layer: locale labels once per contiguous locale group.
    ctx.font = '700 12px -apple-system, Segoe UI, Helvetica Neue, Arial, sans-serif';
    ctx.textBaseline = 'top';
    let i = 0;
    while (i < datasets.length) {
      const lang = datasets[i].lang;
      let j = i;
      while (j + 1 < datasets.length && datasets[j + 1].lang === lang) {
        j += 1;
      }
      const xStart = x.getPixelForValue(i);
      const xEnd = x.getPixelForValue(j);
      const center = (xStart + xEnd) / 2;
      ctx.fillStyle = colorForLang(lang);
      ctx.fillText(lang, center, chartArea.bottom + 82);
      i = j + 1;
    }
    ctx.restore();
  }
};

function bestModelIndex() {
  let idx = 0;
  let best = -Infinity;
  DATA.overall.werr.forEach((v, i) => {
    if (v != null && v > best) {
      best = v;
      idx = i;
    }
  });
  return idx;
}

function buildHeader() {
  const bestIdx = bestModelIndex();
  const bestModel = DATA.models[bestIdx];
  const bestOverall = DATA.overall.werr[bestIdx];
  let bestLocale = null;
  let worstLocale = null;
  DATA.lang_order.forEach((lang) => {
    const w = DATA.lang_avgs[lang].werr[bestIdx];
    if (w == null) return;
    if (!bestLocale || w > bestLocale.werr) bestLocale = { lang, werr: w };
    if (!worstLocale || w < worstLocale.werr) worstLocale = { lang, werr: w };
  });

  document.getElementById('report-title').textContent = DATA.title;
  document.getElementById('report-meta').innerHTML = [
    `Source xlsx: <code>${escapeHtml(DATA.source)}</code>`,
    `Baseline label: <span class="hl-orange">${escapeHtml(DATA.baseline_label)}</span>`,
    `Dataset count: <strong>${DATA.dataset_count}</strong>`,
    `Locale count: <strong>${DATA.lang_order.length}</strong>`
  ].join(' · ');
  document.getElementById('report-tldr').innerHTML = [
    `Best model is <span class="hl-orange">${escapeHtml(bestModel)}</span> with overall WERR <span class="hl-green">${fmtSigned(bestOverall)}</span>.`,
    `${bestLocale ? `Best locale behavior at this checkpoint is <span class="hl-green">${escapeHtml(bestLocale.lang)}</span> (<span class="hl-green">${fmtSigned(bestLocale.werr)}</span>).` : ''}`,
    `${worstLocale ? `Weakest locale behavior is <span class="hl-red">${escapeHtml(worstLocale.lang)}</span> (<span class="hl-red">${fmtSigned(worstLocale.werr)}</span>).` : ''}`
  ].filter(Boolean).join('<br>');
}

function buildLocaleWerrChart() {
  const labels = DATA.lang_order;
  const chartDatasets = [];
  DATA.models.forEach((model, idx) => {
    chartDatasets.push({
      label: model,
      data: labels.map((lang) => DATA.lang_avgs[lang].werr[idx]),
      backgroundColor: labels.map((lang) => colorForLang(lang)),
    });
  });
  const allVals = chartDatasets.flatMap((d) => d.data);
  const range = seriesRange(allVals, { includeZero: true });

  new Chart(document.getElementById('locale-dter-chart'), {
    type: 'bar',
    data: { labels, datasets: chartDatasets },
    options: {
      maintainAspectRatio: false,
      plugins: {
        legend: { position: 'top' },
      },
      scales: {
        x: {
          ticks: { 
            color: (context) => colorForLang(labels[context.index])
          }
        },
        y: {
          min: range.min,
          max: range.max,
          ticks: { callback: (v) => (v >= 0 ? '+' : '') + (v * 100).toFixed(2) + '%' }
        }
      }
    },
    plugins: [valueLabelPlugin]
  });
}

function buildDatasetWerrChart() {
  const bestIdx = bestModelIndex();
  const labels = DATA.datasets.map((r) => shortDatasetName(r.name, 16));
  const vals = DATA.datasets.map((r) => r.werr[bestIdx]);
  const barColors = DATA.datasets.map((r) => colorForLang(r.lang));
  const range = seriesRange(vals, { includeZero: true });

  new Chart(document.getElementById('dataset-werr-chart'), {
    type: 'bar',
    data: {
      labels,
      datasets: [{
        label: `WERR delta vs baseline (${DATA.models[bestIdx]})`,
        data: vals,
        backgroundColor: barColors,
      }]
    },
    options: {
      maintainAspectRatio: false,
      layout: { padding: { bottom: 150 } },
      plugins: {
        legend: { display: true },
        twoLayerTickPlugin: { datasets: DATA.datasets },
      },
      scales: {
        x: { ticks: { display: false } },
        y: {
          min: range.min,
          max: range.max,
          ticks: { callback: (v) => (v * 100).toFixed(2) + '%' }
        }
      }
    },
    plugins: [valueLabelPlugin, twoLayerTickPlugin]
  });
}

function buildRankingTable() {
  const table = document.getElementById('ranking-table');
  const headers = ['Rank', 'Model', 'Direction', 'Baseline overall DTER', 'Model overall DTER', 'DTER delta', 'WERR', 'Datasets'];
  const head = `<thead><tr>${headers.map((h) => `<th>${escapeHtml(h)}</th>`).join('')}</tr></thead>`;
  const body = DATA.ranking.map((row) => {
    const dir = (row.direction || '').toLowerCase();
    const tagClass = dir === 'improve' ? 'tag-improve' : 'tag-degrade';
    return `<tr>
      <td>${row.rank ?? ''}</td>
      <td>${escapeHtml(row.model ?? '')}</td>
      <td><span class="legend-tag ${tagClass}">${escapeHtml(row.direction ?? '')}</span></td>
      <td>${fmtPct(row.baseline_overall_dter)}</td>
      <td>${fmtPct(row.model_overall_dter)}</td>
      <td class="${row.dter_delta >= 0 ? 'pos' : 'neg'}">${fmtSigned(row.dter_delta)}</td>
      <td class="${row.werr >= 0 ? 'pos' : 'neg'}">${fmtSigned(row.werr)}</td>
      <td>${escapeHtml(row.datasets ?? '')}</td>
    </tr>`;
  }).join('');
  table.innerHTML = head + `<tbody>${body}</tbody>`;
}

function rowClass(kind) {
  if (kind === 'overall') return 'overall';
  if (kind === 'lang_avg') return 'lang-avg';
  return '';
}

function buildDataTables() {
  const dterHeaders = ['Dataset', DATA.baseline_label].concat(DATA.models);
  const dterHead = `<thead><tr>${dterHeaders.map((h) => `<th>${escapeHtml(h)}</th>`).join('')}</tr></thead>`;
  const dterBody = DATA.ordered_rows.map((row) => (
    `<tr class="${rowClass(row.kind)}"><td>${escapeHtml(displayDatasetName(row))}</td>${row.dter.map((v) => `<td>${fmtPct(v)}</td>`).join('')}</tr>`
  )).join('');
  document.getElementById('dter-table-wrap').innerHTML = `<div>Full DTER table</div><table>${dterHead}<tbody>${dterBody}</tbody></table>`;

  const werrHeaders = ['Dataset'].concat(DATA.models);
  const werrHead = `<thead><tr>${werrHeaders.map((h) => `<th>${escapeHtml(h)}</th>`).join('')}</tr></thead>`;
  const werrBody = DATA.ordered_rows.map((row) => (
    `<tr class="${rowClass(row.kind)}"><td>${escapeHtml(displayDatasetName(row))}</td>${row.werr.map((v) => `<td class="${v == null ? '' : (v >= 0 ? 'pos' : 'neg')}">${fmtSigned(v)}</td>`).join('')}</tr>`
  )).join('');
  document.getElementById('werr-table-wrap').innerHTML = `<div style="margin-top:16px;">Full WERR table</div><table>${werrHead}<tbody>${werrBody}</tbody></table>`;
}

function buildTakeaways() {
  const items = [];
  const overall = DATA.models.map((m, i) => ({ model: m, werr: DATA.overall.werr[i], dter: DATA.overall.dter[i + 1] }));
  const best = overall.reduce((acc, x) => (!acc || x.werr > acc.werr ? x : acc), null);
  const collapse = overall.find((x) => x.werr < 0);
  const worstLocales = DATA.lang_order.map((lang) => ({ lang, werr: DATA.lang_avgs[lang].werr[bestModelIndex()] }))
    .filter((x) => x.werr != null)
    .sort((a, b) => a.werr - b.werr);

  if (best) {
    items.push(`Best checkpoint is <span class="hl-orange">${escapeHtml(best.model)}</span> at overall DTER <span class="hl-green">${fmtPct(best.dter)}</span> and WERR <span class="hl-green">${fmtSigned(best.werr)}</span>.`);
  }
  if (worstLocales.length) {
    const worst = worstLocales[0];
    const second = worstLocales[1] || worst;
    items.push(`Largest degradation risk locales at the best checkpoint are <span class="hl-red">${escapeHtml(worst.lang)}</span> (${fmtSigned(worst.werr)}) and <span class="hl-red">${escapeHtml(second.lang)}</span> (${fmtSigned(second.werr)}).`);
  }
  if (collapse) {
    items.push(`Overall collapse starts at <span class="hl-orange">${escapeHtml(collapse.model)}</span> where WERR becomes <span class="hl-red">${fmtSigned(collapse.werr)}</span>.`);
  } else {
    items.push(`No overall collapse is observed in this sheet; all compared models remain at or above <span class="hl-green">0.00%</span> WERR.`);
  }
  if (overall.length > 1) {
    const last = overall[overall.length - 1];
    const prev = overall[overall.length - 2];
    const delta = (last.werr ?? 0) - (prev.werr ?? 0);
    items.push(`Late-stage movement from <span class="hl-orange">${escapeHtml(prev.model)}</span> to <span class="hl-orange">${escapeHtml(last.model)}</span> is <span class="${delta >= 0 ? 'hl-green' : 'hl-red'}">${fmtSigned(delta)}</span>.`);
  }
  items.push(`Suggested next step: keep <span class="hl-orange">${escapeHtml(best ? best.model : DATA.models[0] || 'best checkpoint')}</span> as default and focus error analysis on the most negative locale WERR bars in Section 2.`);

  document.getElementById('takeaways').innerHTML = items.slice(0, 5).map((x) => `<li>${x}</li>`).join('');
}

buildHeader();
buildLocaleWerrChart();
buildDatasetWerrChart();
buildRankingTable();
buildDataTables();
buildTakeaways();
  </script>
</body>
</html>
"""


def _maybe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _compute_werr(baseline: Optional[float], model: Optional[float], cell_value: Any) -> Optional[float]:
    if isinstance(cell_value, (int, float)):
        return float(cell_value)
    if baseline is None or model is None:
        return None
    if baseline == 0:
        return None
    return 1 - float(model) / float(baseline)


def _row_kind(name: str) -> str:
    if name == "overall avg":
        return "overall"
    if name.endswith(" avg"):
        return "lang_avg"
    return "data"


def _row_lang(name: str, kind: str) -> str:
    if kind == "lang_avg":
        return name[:-4]
    if kind == "overall":
        return "overall"
    return name.split("_", 1)[0]


def _load_report(xlsx_path: Path, baseline_label: str, title: str) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb["inhouse_dter"]
    rank_ws = wb["overall_improve_degrade"]

    model_labels: List[str] = []
    col = 3
    while True:
        value = ws.cell(row=HEADER_ROW, column=col).value
        if value in (None, "WERR"):
            break
        model_labels.append(str(value))
        col += 1

    n_value_cols = 1 + len(model_labels)
    value_cols = list(range(2, 2 + n_value_cols))
    werr_cols = list(range(2 + n_value_cols, 2 + n_value_cols + len(model_labels)))

    rows: List[Dict[str, Any]] = []
    datasets: List[Dict[str, Any]] = []
    lang_avgs: Dict[str, Dict[str, Any]] = {}
    lang_order: List[str] = []
    datasets_by_lang: Dict[str, List[Dict[str, Any]]] = {}
    overall: Optional[Dict[str, Any]] = None

    row_idx = DATA_START
    while True:
        raw_name = ws.cell(row=row_idx, column=1).value
        if raw_name is None:
            break
        name = str(raw_name)
        kind = _row_kind(name)
        lang = _row_lang(name, kind)
        dter: List[Optional[float]] = []
        for vc in value_cols:
            value = ws.cell(row=row_idx, column=vc).value
            dter.append(float(value) if isinstance(value, (int, float)) else None)

        baseline = dter[0]
        werr: List[Optional[float]] = []
        for i, wc in enumerate(werr_cols, start=1):
            werr_cell = ws.cell(row=row_idx, column=wc).value
            werr.append(_compute_werr(baseline, dter[i], werr_cell))

        row = {"name": name, "lang": lang, "kind": kind, "dter": dter, "werr": werr}
        rows.append(row)

        if kind == "data":
            datasets.append(row)
            datasets_by_lang.setdefault(lang, []).append(row)
            if lang not in lang_order:
                lang_order.append(lang)
        elif kind == "lang_avg":
            lang_avgs[lang] = row
            if lang not in lang_order:
                lang_order.append(lang)
        else:
            overall = row

        row_idx += 1

    if overall is None:
        raise ValueError("Missing overall avg row in inhouse_dter sheet")

    ranking: List[Dict[str, Any]] = []
    rank_row = 3
    while True:
        rank = rank_ws.cell(row=rank_row, column=1).value
        if rank is None:
            break
        ranking.append(
            {
                "rank": rank,
                "model": rank_ws.cell(row=rank_row, column=2).value,
                "direction": rank_ws.cell(row=rank_row, column=3).value,
                "baseline_overall_dter": _maybe_float(rank_ws.cell(row=rank_row, column=4).value),
                "model_overall_dter": _maybe_float(rank_ws.cell(row=rank_row, column=5).value),
                "dter_delta": _maybe_float(rank_ws.cell(row=rank_row, column=6).value),
                "werr": _maybe_float(rank_ws.cell(row=rank_row, column=7).value),
                "datasets": rank_ws.cell(row=rank_row, column=8).value,
            }
        )
        rank_row += 1

    ordered_rows: List[Dict[str, Any]] = []
    ordered_datasets: List[Dict[str, Any]] = []
    for lang in lang_order:
        lang_rows = datasets_by_lang.get(lang, [])
        ordered_rows.extend(lang_rows)
        ordered_datasets.extend(lang_rows)
        if lang in lang_avgs:
            ordered_rows.append(lang_avgs[lang])
    ordered_rows.append(overall)

    return {
        "title": title,
        "source": str(xlsx_path),
        "baseline_label": baseline_label,
        "steps": [baseline_label] + model_labels,
        "models": model_labels,
        "datasets": ordered_datasets,
        "lang_avgs": lang_avgs,
        "overall": overall,
        "ranking": ranking,
        "ordered_rows": ordered_rows,
        "lang_order": lang_order,
        "dataset_count": len(datasets),
    }


def build_html(data: Dict[str, Any]) -> str:
    return TEMPLATE.replace("__DATA__", json.dumps(data, ensure_ascii=True))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="Path to in-house DTER xlsx")
    parser.add_argument("--out", help="Output html path")
    parser.add_argument("--baseline-label", default="Qwen3.5-audio", help="Label shown for baseline")
    parser.add_argument("--title", help="Report title")
    return parser.parse_args()


def _summary(data: Dict[str, Any]) -> str:
    overall_werr = data["overall"]["werr"]
    valid = [(i, v) for i, v in enumerate(overall_werr) if v is not None]
    if not valid:
        return "Summary best=none worst=none collapse=none"

    best_i, best_v = max(valid, key=lambda x: x[1])
    collapse_i = next((i for i, v in enumerate(overall_werr) if v is not None and v < 0), None)
    collapse_label = data["models"][collapse_i] if collapse_i is not None else "none"

    worst_locale = "none"
    worst_locale_v: Optional[float] = None
    for lang in data["lang_order"]:
        row = data["lang_avgs"].get(lang)
        if not row:
            continue
        v = row["werr"][best_i]
        if v is None:
            continue
        if worst_locale_v is None or v < worst_locale_v:
            worst_locale_v = v
            worst_locale = lang

    return (
        "Summary "
        f"best={data['models'][best_i]}:{best_v:.4f} "
        f"worst_locale={worst_locale}:{(worst_locale_v if worst_locale_v is not None else 0.0):.4f} "
        f"collapse={collapse_label}"
    )


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).resolve()
    out_path = Path(args.out).resolve() if args.out else xlsx_path.with_suffix(".html")
    title = args.title or xlsx_path.stem

    data = _load_report(xlsx_path, baseline_label=args.baseline_label, title=title)
    html = build_html(data)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")

    written = out_path.read_text(encoding="utf-8")
    if "__DATA__" in written:
        raise RuntimeError("__DATA__ placeholder remained in output")

    print(f"Wrote {out_path} {out_path.stat().st_size}")
    print(_summary(data))


if __name__ == "__main__":
    main()