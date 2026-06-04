#!/usr/bin/env python3
"""Build an in-house DTER xlsx report (en-US + nl-NL corpora).

Layout (single sheet `inhouse_dter`):
  Row 2  : Header | Baseline | <model-label-1> | ... | WERR(s)
  Row 3  : Column | A        | B               | ... | A->B, ...
  Row 4-6 : en-US datasets
  Row 7   : en-US avg
  Row 8-10: nl-NL datasets
  Row 11  : nl-NL avg
  Row 12  : overall avg
  WERR columns: 1 - <model>/Baseline per non-baseline model column.

Metrics are micro-DTER (sum_edits / sum_tokens), recovered per corpus from verl
`val-aux/<corpus>/dter_n_err/mean@1` and `val-aux/<corpus>/dter_n_ref/mean@1`.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")  # light blue
LANG_AVG_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # light yellow
OVERALL_AVG_FILL = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")  # light green

# ---------------------------------------------------------------------------
# Fixed schema: (locale, [canonical dataset names in report])
# ---------------------------------------------------------------------------
INHOUSE_GROUPS: List[Tuple[str, List[str]]] = [
    ("en-US", [
        "Conversation_DTEST_FY21Q1_en-US",
        "Conversation_OnlineMeetings_DTEST_FY25Q3_en-US_DTEST_OfflineDataCollection",
        "Dictation_Commonset_OfficeOffline_FY24Q3_en-US_DTEST_OfflineDataCollection",
    ]),
    ("nl-NL", [
        "Conversation_DTEST_FY23Q2_nl-NL_DTEST",
        "Conversation_OnlineMeetings_DTEST_FY23Q1_nl-NL_DTEST",
        "Dictation_DTEST_L_D_FY23Q4_nl-NL_DTEST",
    ]),
]

BASELINE_LABEL = "Qwen3.5-audio"
# Baseline = fast-llm-2605-qwen3-5-9b-s2-st-example-r2 @ step 90000.
# Values supplied in the original report (percent → fractions).
BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q1_en-US": 0.1863,
    "Conversation_OnlineMeetings_DTEST_FY25Q3_en-US_DTEST_OfflineDataCollection": 0.1374,
    "Dictation_Commonset_OfficeOffline_FY24Q3_en-US_DTEST_OfflineDataCollection": 0.1010,
    "Conversation_DTEST_FY23Q2_nl-NL_DTEST": 0.2476,
    "Conversation_OnlineMeetings_DTEST_FY23Q1_nl-NL_DTEST": 0.2422,
    "Dictation_DTEST_L_D_FY23Q4_nl-NL_DTEST": 0.1570,
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_enus_seg (segmented long-audio eval).
# 5 en-US TER corpora (the two CustomerSpeechDomainSet_* Entity sets are excluded).
# data_source keys are the short corpus names emitted by recipe.phimm long-audio eval.
# ---------------------------------------------------------------------------
ENUS_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("en-US", [
        "Conversation_DTEST_FY21Q1",
        "Conversation_OnlineMeetings_DTEST_FY25Q3",
        "Dictation_Commonset_OfficeOffline_FY24Q3",
        "OnlineMeetings_CS_Product_FY22_FullMeeting",
        "OnlineMeetings_CS_Shiproom_FY22",
    ]),
]

# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_enus_seg), micro-DTER = n_err / n_ref.
ENUS_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q1": 7766 / 41826,                  # 0.18567
    "Conversation_OnlineMeetings_DTEST_FY25Q3": 6395 / 47149,   # 0.13563
    "Dictation_Commonset_OfficeOffline_FY24Q3": 3888 / 38383,   # 0.10129
    "OnlineMeetings_CS_Product_FY22_FullMeeting": 8165 / 35020,  # 0.23315
    "OnlineMeetings_CS_Shiproom_FY22": 10158 / 39042,           # 0.26018
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_nlnl (segmented long-audio eval).
# 3 nl-NL TER corpora (the two Conversation_DomainSet_*_Entity_* sets are excluded).
# data_source keys are the short corpus names emitted by recipe.phimm long-audio eval.
# ---------------------------------------------------------------------------
NLNL_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("nl-NL", [
        "Conversation_DTEST_FY23Q2",
        "Conversation_OnlineMeetings_DTEST_FY23Q1",
        "Dictation_DTEST_L_D_FY23Q4",
    ]),
]

# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_nlnl), micro-DTER = n_err / n_ref.
NLNL_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY23Q2": 11359 / 45958,                 # 0.24716
    "Conversation_OnlineMeetings_DTEST_FY23Q1": 11148 / 46574,  # 0.23936
    "Dictation_DTEST_L_D_FY23Q4": 6391 / 40642,                 # 0.15725
}

# Registry of selectable schemas: name -> (groups, baseline_metrics).
# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_dadk (segmented long-audio eval).
# 3 da-DK TER corpora. data_source keys are the short corpus names.
# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_dadk), micro-DTER = n_err / n_ref.
# ---------------------------------------------------------------------------
DADK_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("da-DK", [
        "Conversation_DTEST_FY21Q3",
        "Conversation_OnlineMeetings_DTEST_FY23Q1",
        "Dictation_DTEST_L_D_FY23Q4",
    ]),
]

DADK_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q3": 13725 / 58149,                 # 0.23603
    "Conversation_OnlineMeetings_DTEST_FY23Q1": 11911 / 48951,  # 0.24332
    "Dictation_DTEST_L_D_FY23Q4": 10003 / 44482,                # 0.22488
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_huhu (segmented long-audio eval).
# 3 hu-HU TER corpora. data_source keys are the short corpus names.
# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_huhu), micro-DTER = n_err / n_ref.
# ---------------------------------------------------------------------------
HUHU_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("hu-HU", [
        "Conversation_DTEST_FY22Q4",
        "Conversation_OnlineMeetings_DTEST_FY24Q2",
        "Dictation_DTEST_L_D_FY25Q2",
    ]),
]

HUHU_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY22Q4": 7655 / 33569,                  # 0.22804
    "Conversation_OnlineMeetings_DTEST_FY24Q2": 7676 / 35005,   # 0.21928
    "Dictation_DTEST_L_D_FY25Q2": 7557 / 31098,                 # 0.24301
}

# ---------------------------------------------------------------------------
# Alternate schema: inhouse_2605_nbno (segmented long-audio eval).
# 3 nb-NO TER corpora. data_source keys are the short corpus names.
# Baseline = Qwen3.5-audio (eval_qwen/inhouse_2605_nbno), micro-DTER = n_err / n_ref.
# ---------------------------------------------------------------------------
NBNO_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("nb-NO", [
        "Conversation_DTEST_FY21Q3",
        "Conversation_OnlineMeetings_DTEST_FY23Q1",
        "Dictation_DTEST_L_D_FY23Q4",
    ]),
]

NBNO_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q3": 10345 / 47273,                 # 0.21884
    "Conversation_OnlineMeetings_DTEST_FY23Q1": 7803 / 37996,   # 0.20536
    "Dictation_DTEST_L_D_FY23Q4": 8584 / 40601,                 # 0.21142
}

# ---------------------------------------------------------------------------
# Combined schema: the three new locales (da-DK, hu-HU, nb-NO) in one report.
# Canonical names carry a locale suffix to avoid short-name collisions between
# da-DK and nb-NO (which share corpus short names). Baseline = Qwen3.5-audio.
# When sourcing from a single-locale eval job, pass --metrics with these
# suffixed keys (or build per-locale reports with the dadk/huhu/nbno schemas).
# ---------------------------------------------------------------------------
NEWLOCS_SEG_GROUPS: List[Tuple[str, List[str]]] = [
    ("da-DK", [
        "Conversation_DTEST_FY21Q3_da-DK",
        "Conversation_OnlineMeetings_DTEST_FY23Q1_da-DK",
        "Dictation_DTEST_L_D_FY23Q4_da-DK",
    ]),
    ("hu-HU", [
        "Conversation_DTEST_FY22Q4_hu-HU",
        "Conversation_OnlineMeetings_DTEST_FY24Q2_hu-HU",
        "Dictation_DTEST_L_D_FY25Q2_hu-HU",
    ]),
    ("nb-NO", [
        "Conversation_DTEST_FY21Q3_nb-NO",
        "Conversation_OnlineMeetings_DTEST_FY23Q1_nb-NO",
        "Dictation_DTEST_L_D_FY23Q4_nb-NO",
    ]),
]

NEWLOCS_SEG_BASELINE_METRICS: Dict[str, float] = {
    "Conversation_DTEST_FY21Q3_da-DK": 13725 / 58149,
    "Conversation_OnlineMeetings_DTEST_FY23Q1_da-DK": 11911 / 48951,
    "Dictation_DTEST_L_D_FY23Q4_da-DK": 10003 / 44482,
    "Conversation_DTEST_FY22Q4_hu-HU": 7655 / 33569,
    "Conversation_OnlineMeetings_DTEST_FY24Q2_hu-HU": 7676 / 35005,
    "Dictation_DTEST_L_D_FY25Q2_hu-HU": 7557 / 31098,
    "Conversation_DTEST_FY21Q3_nb-NO": 10345 / 47273,
    "Conversation_OnlineMeetings_DTEST_FY23Q1_nb-NO": 7803 / 37996,
    "Dictation_DTEST_L_D_FY23Q4_nb-NO": 8584 / 40601,
}

# Registry of selectable schemas: name -> (groups, baseline_metrics).
SCHEMAS: Dict[str, Tuple[List[Tuple[str, List[str]]], Dict[str, float]]] = {
    "default": (INHOUSE_GROUPS, BASELINE_METRICS),
    "enus_seg": (ENUS_SEG_GROUPS, ENUS_SEG_BASELINE_METRICS),
    "nlnl_seg": (NLNL_SEG_GROUPS, NLNL_SEG_BASELINE_METRICS),
    "dadk_seg": (DADK_SEG_GROUPS, DADK_SEG_BASELINE_METRICS),
    "huhu_seg": (HUHU_SEG_GROUPS, HUHU_SEG_BASELINE_METRICS),
    "nbno_seg": (NBNO_SEG_GROUPS, NBNO_SEG_BASELINE_METRICS),
    "newlocs_seg": (NEWLOCS_SEG_GROUPS, NEWLOCS_SEG_BASELINE_METRICS),
}

# Match: val-aux/<corpus>/<key>/mean@1:<float>   where <key> ∈ {dter, dter_n_err, dter_n_ref}
DTER_LINE_RE = re.compile(
    r"val-aux/(?P<dataset>[A-Za-z0-9_.\-]+)/(?P<key>dter|dter_n_err|dter_n_ref)/mean@1[:=]\s*(?P<value>[0-9.eE+-]+)"
)

LOCALE_SUFFIX_RE = re.compile(r"_[a-z]{2}-[A-Z]{2}(?:_.*)?$")


def _canonical_lookup(parsed: Dict[str, float], canonical: str) -> Optional[float]:
    """Look up a parsed value for a canonical dataset name with sensible fallbacks."""
    if canonical in parsed:
        return parsed[canonical]
    # Strip _xx-XX(...) suffix and retry.
    short = LOCALE_SUFFIX_RE.sub("", canonical)
    if short != canonical and short in parsed:
        return parsed[short]
    # Case-insensitive search.
    ci = {k.lower(): v for k, v in parsed.items()}
    for cand in (canonical.lower(), short.lower()):
        if cand in ci:
            return ci[cand]
    return None


def parse_dter_lines(text: str) -> Dict[str, float]:
    """Extract micro-DTER per corpus from verl log text.

    Reads the latest `dter_n_err/mean@1` and `dter_n_ref/mean@1` per corpus and
    returns `n_err / n_ref` (fraction in 0..1). Falls back to the raw `dter/mean@1`
    value only when the n_err / n_ref pair is missing (with a warning) — that
    macro value does NOT match in-house micro-DTER and should be replaced.
    """
    n_err: Dict[str, float] = {}
    n_ref: Dict[str, float] = {}
    macro: Dict[str, float] = {}
    for m in DTER_LINE_RE.finditer(text):
        try:
            v = float(m.group("value"))
        except ValueError:
            continue
        ds = m.group("dataset")
        key = m.group("key")
        if key == "dter_n_err":
            n_err[ds] = v
        elif key == "dter_n_ref":
            n_ref[ds] = v
        else:  # 'dter'
            macro[ds] = v

    out: Dict[str, float] = {}
    for ds in set(n_err) | set(n_ref) | set(macro):
        if ds in n_err and ds in n_ref and n_ref[ds] > 0:
            out[ds] = n_err[ds] / n_ref[ds]
        elif ds in macro:
            print(
                f"[warn] {ds}: only macro `dter/mean@1` available; using it as a fallback. "
                "This is NOT the in-house micro-DTER reference metric.",
                file=sys.stderr,
            )
            out[ds] = macro[ds] if macro[ds] <= 1.5 else macro[ds] / 100.0
    return out


def fetch_ray_logs(node: str, job_id: str) -> str:
    remote = f"bash -l -c 'ray job logs {job_id} 2>&1'"
    cmd = ["brix", "ssh", node, "--", remote]
    proc = subprocess.run(cmd, capture_output=True, text=True, check=False)
    if proc.returncode != 0:
        print(
            f"[warn] ray job logs failed for {node}/{job_id} (rc={proc.returncode}):\n"
            f"{proc.stderr[-500:]}",
            file=sys.stderr,
        )
    return proc.stdout


def load_metrics(args: argparse.Namespace) -> Dict[str, float]:
    """Collect raw parsed metrics keyed by whatever names appear in the logs."""
    metrics: Dict[str, float] = {}
    if args.metrics:
        data = json.loads(Path(args.metrics).read_text())
        metrics.update({k: float(v) for k, v in data.items()})
    if args.from_text:
        for text_path in args.from_text:
            metrics.update(parse_dter_lines(Path(text_path).read_text()))
    if args.from_ray:
        for node, job_id in args.from_ray:
            metrics.update(parse_dter_lines(fetch_ray_logs(node, job_id)))
    return metrics


def _project_to_canonical(parsed: Dict[str, float]) -> Dict[str, float]:
    """Map parsed metrics to the canonical dataset names used in the report."""
    out: Dict[str, float] = {}
    for _, datasets in INHOUSE_GROUPS:
        for ds in datasets:
            v = _canonical_lookup(parsed, ds)
            if v is not None:
                out[ds] = v
    return out


# ---------------------------------------------------------------------------
# xlsx build
# ---------------------------------------------------------------------------
HEADER_ROW = 2
COLUMN_ROW = 3
DATA_START = 4


def _row_layout() -> Tuple[List[Tuple[str, str]], int]:
    """Return (rows, overall_avg_row).

    Each row is ('data'|'lang_avg', label). The overall avg row index is the
    row index immediately following the last layout row.
    """
    rows: List[Tuple[str, str]] = []
    for locale, datasets in INHOUSE_GROUPS:
        for ds in datasets:
            rows.append(("data", ds))
        rows.append(("lang_avg", f"{locale} avg"))
    overall_avg_row = DATA_START + len(rows)
    return rows, overall_avg_row


def build_workbook(columns: List[Tuple[str, Dict[str, float]]], out_path: Path) -> None:
    """``columns`` is a list of (label, canonical metrics dict). First is the baseline."""

    wb = Workbook()
    ws = wb.active
    ws.title = "inhouse_dter"

    n_models = len(columns)
    model_cols = list(range(2, 2 + n_models))
    werr_cols = list(range(2 + n_models, 2 + n_models + (n_models - 1)))

    bold = Font(bold=True)

    # Header row
    ws.cell(row=HEADER_ROW, column=1, value="Header").font = bold
    ws.cell(row=HEADER_ROW, column=2, value="Baseline").font = bold
    for i, (label, _) in enumerate(columns[1:], start=0):
        ws.cell(row=HEADER_ROW, column=3 + i, value=label).font = bold
    for j, _ in enumerate(model_cols[1:]):
        ws.cell(row=HEADER_ROW, column=werr_cols[j], value="WERR").font = bold

    # Column row
    ws.cell(row=COLUMN_ROW, column=1, value="Column").font = bold
    for i, mc in enumerate(model_cols):
        ws.cell(row=COLUMN_ROW, column=mc, value=chr(ord("A") + i)).font = bold
    for j, wc in enumerate(werr_cols):
        target_letter = chr(ord("A") + 1 + j)
        ws.cell(row=COLUMN_ROW, column=wc, value=f"A->{target_letter}").font = bold

    # Data + lang_avg rows
    layout, overall_avg_row = _row_layout()
    base_letter = get_column_letter(model_cols[0])
    current_lang_start: Optional[int] = None
    lang_avg_rows: List[int] = []

    for offset, (kind, label) in enumerate(layout):
        r = DATA_START + offset
        ws.cell(row=r, column=1, value=label)
        if kind == "data":
            if current_lang_start is None:
                current_lang_start = r
            for i, (_, metrics) in enumerate(columns):
                v = metrics.get(label)
                if v is not None:
                    ws.cell(row=r, column=model_cols[i], value=float(v))
            for j, wc in enumerate(werr_cols):
                tgt_letter = get_column_letter(model_cols[1 + j])
                ws.cell(row=r, column=wc, value=f"=1-{tgt_letter}{r}/{base_letter}{r}")
        else:  # lang_avg
            lang_avg_rows.append(r)
            ws.cell(row=r, column=1).font = bold
            assert current_lang_start is not None
            group_datasets = [
                lbl for k, lbl in layout[offset - (r - current_lang_start):offset] if k == "data"
            ]
            lang_avg_per_col: List[Optional[float]] = []
            for i, (_, metrics) in enumerate(columns):
                vals = [metrics[ds] for ds in group_datasets if ds in metrics]
                avg = sum(vals) / len(vals) if vals else None
                lang_avg_per_col.append(avg)
                if avg is not None:
                    ws.cell(row=r, column=model_cols[i], value=float(avg)).font = bold
            for j, wc in enumerate(werr_cols):
                base_v = lang_avg_per_col[0]
                tgt_v = lang_avg_per_col[1 + j]
                if base_v and tgt_v is not None:
                    ws.cell(row=r, column=wc, value=float(1 - tgt_v / base_v)).font = bold
            current_lang_start = None

    # Overall avg row (mean across all data datasets).
    r = overall_avg_row
    ws.cell(row=r, column=1, value="overall avg").font = bold
    all_data = [lbl for kind, lbl in layout if kind == "data"]
    overall_per_col: List[Optional[float]] = []
    for i, (_, metrics) in enumerate(columns):
        vals = [metrics[ds] for ds in all_data if ds in metrics]
        avg = sum(vals) / len(vals) if vals else None
        overall_per_col.append(avg)
        if avg is not None:
            ws.cell(row=r, column=model_cols[i], value=float(avg)).font = bold
    for j, wc in enumerate(werr_cols):
        base_v = overall_per_col[0]
        tgt_v = overall_per_col[1 + j]
        if base_v and tgt_v is not None:
            ws.cell(row=r, column=wc, value=float(1 - tgt_v / base_v)).font = bold

    # Formatting.
    pct_fmt = "0.00%"
    for rr in range(DATA_START, overall_avg_row + 1):
        for mc in model_cols + werr_cols:
            ws.cell(row=rr, column=mc).number_format = pct_fmt

    all_cols = [1] + model_cols + werr_cols
    header_rows = [HEADER_ROW, COLUMN_ROW]
    for rr in header_rows:
        for c in all_cols:
            ws.cell(row=rr, column=c).fill = HEADER_FILL
    for rr in lang_avg_rows:
        for c in all_cols:
            ws.cell(row=rr, column=c).fill = LANG_AVG_FILL
    for c in all_cols:
        ws.cell(row=overall_avg_row, column=c).fill = OVERALL_AVG_FILL

    ws.column_dimensions["A"].width = 60
    for mc in model_cols + werr_cols:
        ws.column_dimensions[get_column_letter(mc)].width = 18

    center = Alignment(horizontal="center", vertical="center")
    for rr in range(HEADER_ROW, overall_avg_row + 1):
        for c in model_cols + werr_cols:
            ws.cell(row=rr, column=c).alignment = center

    # WERR 3-color scale.
    for wc in werr_cols:
        col_letter = get_column_letter(wc)
        rng = f"{col_letter}{DATA_START}:{col_letter}{overall_avg_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=-1, start_color="FFF8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFFFF",
            end_type="num", end_value=1, end_color="FF63BE7B",
        )
        ws.conditional_formatting.add(rng, rule)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def read_existing_xlsx(path: Path) -> List[Tuple[str, Dict[str, float]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["inhouse_dter"] if "inhouse_dter" in wb.sheetnames else wb.active

    model_cols: List[int] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=COLUMN_ROW, column=c).value
        if isinstance(v, str) and len(v) == 1 and v.isalpha():
            model_cols.append(c)
        else:
            break

    labels: List[str] = []
    for mc in model_cols:
        v = ws.cell(row=HEADER_ROW, column=mc).value
        labels.append(str(v) if v is not None else f"col{mc}")

    metrics_per_col: List[Dict[str, float]] = [dict() for _ in model_cols]
    layout, _ = _row_layout()
    for offset, (kind, label) in enumerate(layout):
        if kind != "data":
            continue
        r = DATA_START + offset
        for i, mc in enumerate(model_cols):
            v = ws.cell(row=r, column=mc).value
            if isinstance(v, (int, float)):
                metrics_per_col[i][label] = float(v)

    return [(label, m) for label, m in zip(labels, metrics_per_col)]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("label", nargs="?", help="New model label (column header). Optional with --baseline-only.")
    p.add_argument(
        "--baseline-only",
        action="store_true",
        help="Emit only the embedded baseline column (no model/WERR columns). Useful for registering or publishing a schema's baseline numbers.",
    )
    p.add_argument(
        "--schema",
        choices=sorted(SCHEMAS),
        default="default",
        help=(
            "Dataset schema / embedded baseline to compare against. "
            "'default' = 6-dataset en-US+nl-NL canonical; "
            "'enus_seg' = 5 en-US TER corpora from inhouse_2605_enus_seg; "
            "'nlnl_seg' = 3 nl-NL TER corpora from inhouse_2605_nlnl; "
            "'dadk_seg' = 3 da-DK corpora from inhouse_2605_dadk; "
            "'huhu_seg' = 3 hu-HU corpora from inhouse_2605_huhu; "
            "'nbno_seg' = 3 nb-NO corpora from inhouse_2605_nbno; "
            "'newlocs_seg' = combined da-DK + hu-HU + nb-NO (locale-suffixed keys)."
        ),
    )
    p.add_argument("--from-ray", nargs=2, metavar=("NODE", "JOB_ID"), action="append", default=[])
    p.add_argument("--from-text", action="append", default=[])
    p.add_argument("--metrics", help="JSON file: {dataset: dter_fraction}.")
    p.add_argument("--baseline", help="Override baseline metrics JSON.")
    p.add_argument("--baseline-label", default=BASELINE_LABEL)
    p.add_argument(
        "--extend-xlsx",
        help=(
            "Existing xlsx to extend by appending the new model as the next column. "
            "The sheet layout must match the template."
        ),
    )
    p.add_argument("--out", help="Output xlsx path (default: tmp/inhouse_dter_report/<label>.xlsx).")
    return p.parse_args()


def main() -> int:
    args = parse_args()

    # Select dataset schema + embedded baseline. _row_layout / _project_to_canonical
    # read the module-level INHOUSE_GROUPS, so rebind it for the chosen schema.
    global INHOUSE_GROUPS
    schema_groups, schema_baseline = SCHEMAS[args.schema]
    INHOUSE_GROUPS = schema_groups

    # Baseline-only mode: emit just the embedded (or overridden) baseline column.
    if args.baseline_only:
        baseline_metrics = schema_baseline
        if args.baseline:
            baseline_metrics = {k: float(v) for k, v in json.loads(Path(args.baseline).read_text()).items()}
        columns = [(args.baseline_label, baseline_metrics)]
        out_path = Path(args.out or f"tmp/inhouse_dter_report/{args.schema}_baseline.xlsx")
        build_workbook(columns, out_path)
        print(out_path)
        return 0

    if args.label is None:
        print("[error] label is required unless --baseline-only is set.", file=sys.stderr)
        return 2

    raw_metrics = load_metrics(args)
    if not raw_metrics:
        print("[error] no metrics collected; pass --from-ray/--from-text/--metrics.", file=sys.stderr)
        return 2
    new_metrics = _project_to_canonical(raw_metrics)
    if not new_metrics:
        print(
            "[error] none of the parsed metrics map to canonical in-house dataset names.\n"
            f"  parsed keys: {sorted(raw_metrics)[:10]}{'...' if len(raw_metrics) > 10 else ''}",
            file=sys.stderr,
        )
        return 2

    if args.extend_xlsx:
        columns = read_existing_xlsx(Path(args.extend_xlsx))
        if not columns:
            print(f"[warn] could not parse {args.extend_xlsx}; using baseline only.", file=sys.stderr)
    else:
        columns = []

    if not columns:
        baseline_metrics = schema_baseline
        if args.baseline:
            baseline_metrics = {k: float(v) for k, v in json.loads(Path(args.baseline).read_text()).items()}
        columns = [(args.baseline_label, baseline_metrics)]

    columns.append((args.label, new_metrics))

    out_path = Path(
        args.out
        or f"tmp/inhouse_dter_report/{args.label.replace('/', '_').replace('@', '_')}.xlsx"
    )
    build_workbook(columns, out_path)
    print(out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
