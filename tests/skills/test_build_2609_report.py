import importlib.util
import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


SCRIPT = (
    Path(__file__).parents[2]
    / ".github/skills/eval-2609-benchmark-report/scripts/build_2609_report.py"
)


def _load_report_module():
    spec = importlib.util.spec_from_file_location("build_2609_report", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_inhouse_avg_rows_use_excel_arithmetic_formulas(tmp_path, monkeypatch):
    source = tmp_path / "measures"
    measures = {
        "enus_conv_fy21q1": {"dter": 0.10, "dter_n_err": 1, "dter_n_ref": 10},
        "enus_conv_om_fy25q3": {"dter": 0.30, "dter_n_err": 90, "dter_n_ref": 300},
    }
    for corpus, values in measures.items():
        corpus_dir = source / corpus
        corpus_dir.mkdir(parents=True)
        (corpus_dir / "measures.json").write_text(json.dumps(values))

    output = tmp_path / "report.xlsx"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            str(SCRIPT),
            "--label",
            "candidate",
            "--inhouse-dter",
            str(source),
            "--out",
            str(output),
        ],
    )

    report = _load_report_module()
    assert report.main() == 0

    workbook = load_workbook(output, data_only=False)
    sheet = workbook["inhouse_dter"]
    overall = next(row for row in sheet.iter_rows(values_only=True) if row[0] == "overall avg")
    baseline = report.BENCHMARKS["inhouse_dter"]["embedded_baseline"]
    baseline_values = [
        baseline[key]["dter"]
        for _, datasets in report.INHOUSE_GROUPS
        for key, _ in datasets
    ]
    assert sheet["B7"].value == "=AVERAGE(B4:B6)"
    assert sheet["C7"].value == "=AVERAGE(C4:C6)"
    assert sheet["D7"].value == "=1-C7/B7"
    assert overall[1] == "=AVERAGE(B4:B6,B8:B10,B12:B14,B16:B18,B20:B22,B24:B26)"
    assert overall[2] == "=AVERAGE(C4:C6,C8:C10,C12:C14,C16:C18,C20:C22,C24:C26)"
    assert overall[3] == "=1-C28/B28"
    assert workbook["summary"].cell(2, 3).value == pytest.approx(
        sum(baseline_values) / len(baseline_values)
    )
    assert workbook["summary"].cell(2, 4).value == pytest.approx(0.20)
    summary = workbook["summary"]
    assert [str(item.sqref) for item in summary.conditional_formatting] == ["E2"]
    rule = next(iter(summary.conditional_formatting)).rules[0]
    assert [value.val for value in rule.colorScale.cfvo] == [-0.1, 0.0, 0.1]
    assert [color.rgb for color in rule.colorScale.color] == [
        "FFF8696B",
        "FFFFFFFF",
        "FF63BE7B",
    ]
    assert len(summary._charts) == 1
    chart = summary._charts[0]
    assert len(chart._charts) == 1
    assert len(chart.series) == 1
    assert chart.title is None
    assert chart.visible_cells_only is False
    assert chart.anchor._from.col == 0
    assert chart.anchor._from.row == 4
    assert chart.series[0].tx.strRef.f.endswith("!X2")
    assert chart.series[0].cat.numRef.f.endswith("!$W$3")
    assert chart.series[0].invertIfNegative is False
    assert len(chart.series[0].dPt) == 1
    assert chart.series[0].dPt[0].invertIfNegative is False
    assert chart.series[0].dPt[0].graphicalProperties.solidFill.srgbClr == "4F81BD"
    assert chart.series[0].dPt[0].graphicalProperties.line.solidFill.srgbClr == "4F81BD"
    assert chart.y_axis.majorGridlines is None
    baseline_overall = sum(baseline_values) / len(baseline_values)
    delta = 1 - 0.20 / baseline_overall
    assert chart.y_axis.scaling.min == pytest.approx(delta - abs(delta) * 0.15)
    assert chart.y_axis.scaling.max == 0
    assert chart.x_axis.delete is False
    assert chart.x_axis.axPos == "b"
    assert chart.x_axis.title is None
    assert chart.x_axis.tickLblPos == "low"
    assert chart.x_axis.majorTickMark is None
    assert chart.x_axis.spPr.ln.solidFill.srgbClr == "595959"
    assert chart.x_axis.majorGridlines is not None
    assert chart.x_axis.majorGridlines.spPr.ln.solidFill.srgbClr == "D9D9D9"
    assert chart.y_axis.delete is False
    assert chart.y_axis.axPos == "l"
    assert chart.y_axis.title is None
    assert chart.y_axis.crossesAt == 0
    assert chart.legend is not None
    assert chart.legend.position == "t"
    assert chart.dLbls.showVal is True
    assert chart.dLbls.showLegendKey is False
    assert chart.dLbls.numFmt == "0.00%"
    assert chart.dLbls.position == "outEnd"
    assert chart.overlap == -40
    assert chart.gapWidth == 260
    assert summary["W3"].value == "inhouse_dter"


def test_default_report_path_uses_model_subdirectory():
    report = _load_report_module()

    assert report.default_report_path(
        "remax_2609@step560",
        "az://orngwus2cresco/data/boren/outputs/ver_2609/remax_2609/global_step_560/qwen_hf/",
    ) == Path(
        "tmp/eval_2609_reports/remax_2609/remax_2609_step560.xlsx"
    )
    assert report.default_report_path(
        "remax_2609_step560",
        "",
    ) == Path(
        "tmp/eval_2609_reports/remax_2609/remax_2609_step560.xlsx"
    )


def test_collect_reads_remote_json_metric_source(monkeypatch):
    report = _load_report_module()
    source = "az://orngwus2cresco/data/results/key_metrics.json"
    payload = {"de_fleurs": {"wer": 0.025}, "de_mcv": {"wer": 0.031}}

    def fake_run(command, **kwargs):
        assert command == ["bbb", "cat", source]
        return subprocess.CompletedProcess(command, 0, json.dumps(payload), "")

    monkeypatch.setattr(report.subprocess, "run", fake_run)

    assert report.collect(source, ["wer"]) == payload


def test_default_output_writes_checkpoint_report_under_model(tmp_path, monkeypatch):
    source = tmp_path / "measures"
    corpus_dir = source / "enus_conv_fy21q1"
    corpus_dir.mkdir(parents=True)
    (corpus_dir / "measures.json").write_text(
        json.dumps({"dter": 0.10, "dter_n_err": 1, "dter_n_ref": 10})
    )
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            str(SCRIPT),
            "--label",
            "remax_2609_step10",
            "--candidate-model-path",
            "az://orngwus2cresco/data/boren/outputs/ver_2609/remax_2609/global_step_10/qwen_hf/",
            "--inhouse-dter",
            str(source),
        ],
    )

    report = _load_report_module()
    assert report.main() == 0
    assert (
        tmp_path
        / "tmp/eval_2609_reports/remax_2609/remax_2609_step10.xlsx"
    ).is_file()


def test_merged_summary_chart_clusters_steps_by_dataset():
    report = _load_report_module()
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.append(["Model", "candidate"])
    summary.append([])
    summary.append(["Checkpoint", "Benchmark", "Metric", "Baseline", "Candidate", "Delta"])
    for checkpoint, values in (
        ("step10", (("inhouse_dter", 0.18, 0.17), ("openasr_ml", 0.03, 0.031))),
        ("step20", (("inhouse_dter", 0.18, 0.19), ("openasr_ml", 0.03, 0.028))),
    ):
        for benchmark, baseline, candidate in values:
            summary.append(
                [checkpoint, benchmark, "DTER", baseline, candidate, 1 - candidate / baseline]
            )

    assert report.apply_summary_charts(summary) == 1

    marker = next(cell.column for cell in summary[1] if cell.value == "__chart_data__")
    assert [summary.cell(row, marker).value for row in range(3, 5)] == [
        "inhouse_dter",
        "openasr_ml",
    ]
    assert [summary.cell(2, column).value for column in range(marker + 1, marker + 3)] == [
        "step10",
        "step20",
    ]
    chart = summary._charts[0]
    assert len(chart.series) == 2
    assert all(len(series.dPt) == 2 for series in chart.series)
    assert [series.graphicalProperties.solidFill.srgbClr for series in chart.series] == [
        "4F81BD",
        "C0504D",
    ]
    assert all(
        point.graphicalProperties.solidFill.srgbClr == "4F81BD"
        for point in chart.series[0].dPt
    )
    assert all(
        point.graphicalProperties.solidFill.srgbClr == "C0504D"
        for point in chart.series[1].dPt
    )
    assert chart.legend.position == "t"


def test_single_summary_chart_colors_each_dataset():
    report = _load_report_module()
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.append(["Benchmark", "Metric", "Baseline", "Candidate (step100)", "delta"])
    summary.append(["inhouse_dter", "DTER", 0.18, 0.17, 1 - 0.17 / 0.18])
    summary.append(["openasr_ml", "WER", 0.03, 0.029, 1 - 0.029 / 0.03])
    summary.append(["mixlang", "DTER", 0.21, 0.16, 1 - 0.16 / 0.21])

    assert report.apply_summary_charts(summary) == 1

    chart = summary._charts[0]
    assert len(chart.series) == 1
    assert [point.graphicalProperties.solidFill.srgbClr for point in chart.series[0].dPt] == [
        "4F81BD",
        "C0504D",
        "9BBB59",
    ]
    assert chart.legend.position == "t"